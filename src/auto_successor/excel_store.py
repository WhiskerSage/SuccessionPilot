from __future__ import annotations

import json
import re
import time
from datetime import datetime, timezone
from pathlib import Path
import csv

from openpyxl import Workbook, load_workbook

from .models import JobRecord, NoteRecord, SendLogRecord, SummaryRecord

RAW_HEADERS = [
    "run_id",
    "keyword",
    "note_id",
    "title",
    "author",
    "publish_time",
    "publish_timestamp",
    "publish_time_text",
    "publish_time_quality",
    "like_count",
    "comment_count",
    "share_count",
    "url",
    "xsec_token",
    "detail_text",
    "comments_preview",
    "fetched_at",
    "first_seen_at",
    "updated_at",
    "raw_json",
]

SUMMARY_HEADERS = [
    "run_id",
    "note_id",
    "keyword",
    "publish_time",
    "publish_timestamp",
    "title",
    "author",
    "summary",
    "confidence",
    "risk_flags",
    "url",
    "created_at",
]

SEND_HEADERS = [
    "run_id",
    "note_id",
    "channel",
    "send_status",
    "send_response",
    "sent_at",
]

JOB_HEADERS = [
    "run_id",
    "Company",
    "Position",
    "publish_time",
    "Location",
    "Requirements",
    "arrival_time",
    "application_method",
    "author",
    "risk_line",
    "match_score",
    "match_reason",
    "Link",
    "PostID",
    "mode",
    "comment_count",
    "comments_preview",
    "original_text",
    "source_title",
    "opportunity_point",
    "outreach_message",
]


class ExcelStore:
    def __init__(self, excel_path: str) -> None:
        self.path = Path(excel_path)
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def write(
        self,
        raw: list[NoteRecord],
        summaries: list[SummaryRecord],
        send_logs: list[SendLogRecord],
        jobs: list[JobRecord] | None = None,
    ) -> None:
        wb = self._load_or_create()

        raw_rows = self._merge_raw_rows(wb["raw_notes"], raw)
        summary_rows = self._merge_summary_rows(wb["succession_summary"], summaries)
        send_rows = self._merge_send_rows(wb["send_log"], send_logs)
        job_rows = self._merge_job_rows(wb["jobs"], jobs or [])

        self._rewrite_sheet(wb["raw_notes"], RAW_HEADERS, raw_rows)
        self._rewrite_sheet(wb["succession_summary"], SUMMARY_HEADERS, summary_rows)
        self._rewrite_sheet(wb["send_log"], SEND_HEADERS, send_rows)
        self._rewrite_sheet(wb["jobs"], JOB_HEADERS, job_rows)

        try:
            for attempt in range(3):
                try:
                    wb.save(self.path)
                    return
                except PermissionError:
                    if attempt < 2:
                        time.sleep(1.0)
                        continue
                    fallback = self.path.with_name(
                        f"{self.path.stem}.locked-{datetime.now().strftime('%Y%m%d-%H%M%S')}{self.path.suffix}"
                    )
                    wb.save(fallback)
                    return
        finally:
            wb.close()

    def _load_or_create(self):
        if self.path.exists():
            wb = load_workbook(self.path)
            for name, headers in [
                ("raw_notes", RAW_HEADERS),
                ("succession_summary", SUMMARY_HEADERS),
                ("send_log", SEND_HEADERS),
                ("jobs", JOB_HEADERS),
            ]:
                if name not in wb.sheetnames:
                    ws = wb.create_sheet(title=name)
                    ws.append(headers)
            return wb

        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for name, headers in [
            ("raw_notes", RAW_HEADERS),
            ("succession_summary", SUMMARY_HEADERS),
            ("send_log", SEND_HEADERS),
            ("jobs", JOB_HEADERS),
        ]:
            ws = wb.create_sheet(title=name)
            ws.append(headers)
        return wb

    def _rewrite_sheet(self, ws, headers: list[str], rows: list[dict]) -> None:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
        for row in rows:
            ws.append([row.get(col, "") for col in headers])

    def _read_sheet_rows(self, ws, headers: list[str]) -> list[dict]:
        rows = []
        for values in ws.iter_rows(min_row=2, values_only=True):
            if not values:
                continue
            row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
            rows.append(row)
        return rows

    def _merge_raw_rows(self, ws, raw: list[NoteRecord]) -> list[dict]:
        existing = self._read_sheet_rows(ws, RAW_HEADERS)
        by_note: dict[str, dict] = {str(r.get("note_id") or ""): r for r in existing if r.get("note_id")}

        for note in raw:
            prev = by_note.get(note.note_id, {})

            prev_publish_ts = self._safe_int(prev.get("publish_timestamp"))
            prev_publish_time = str(prev.get("publish_time") or "")
            prev_publish_text = str(prev.get("publish_time_text") or "")
            prev_publish_quality = str(prev.get("publish_time_quality") or "").strip().lower()

            incoming_publish_ts = int(note.publish_time.timestamp())
            incoming_publish_time = note.publish_time.isoformat()
            incoming_publish_text = note.publish_time_text
            incoming_publish_quality = str(getattr(note, "publish_time_quality", "parsed") or "").strip().lower()
            if incoming_publish_quality not in {"parsed", "fallback"}:
                incoming_publish_quality = "fallback"
            incoming_is_relative = self._is_relative_publish_text(incoming_publish_text)

            # Keep previous publish time when current parsing falls back.
            if incoming_publish_quality != "parsed" and prev_publish_ts > 0:
                publish_timestamp = prev_publish_ts
                publish_time = prev_publish_time or incoming_publish_time
                publish_time_text = prev_publish_text or incoming_publish_text
                publish_time_quality = prev_publish_quality or "fallback"
            elif incoming_publish_quality == "parsed" and incoming_is_relative and prev_publish_ts > 0:
                # Relative time text (e.g. "18分钟前") can drift across crawls; never move newer than previous value.
                if incoming_publish_ts > prev_publish_ts:
                    publish_timestamp = prev_publish_ts
                    publish_time = prev_publish_time or incoming_publish_time
                    publish_time_text = prev_publish_text or incoming_publish_text
                    publish_time_quality = prev_publish_quality or incoming_publish_quality
                else:
                    publish_timestamp = incoming_publish_ts
                    publish_time = incoming_publish_time
                    publish_time_text = incoming_publish_text
                    publish_time_quality = incoming_publish_quality
            else:
                publish_timestamp = incoming_publish_ts
                publish_time = incoming_publish_time
                publish_time_text = incoming_publish_text
                publish_time_quality = incoming_publish_quality

            first_seen_at = str(prev.get("first_seen_at") or prev.get("fetched_at") or self._to_iso_utc(note.fetched_at))
            updated_at = self._to_iso_utc(note.fetched_at)

            by_note[note.note_id] = {
                "run_id": note.run_id,
                "keyword": note.keyword or str(prev.get("keyword") or ""),
                "note_id": note.note_id,
                "title": note.title or str(prev.get("title") or ""),
                "author": note.author or str(prev.get("author") or ""),
                "publish_time": publish_time,
                "publish_timestamp": publish_timestamp,
                "publish_time_text": publish_time_text,
                "publish_time_quality": publish_time_quality,
                "like_count": self._prefer_latest_int(note.like_count, self._safe_int(prev.get("like_count"))),
                "comment_count": self._prefer_latest_int(note.comment_count, self._safe_int(prev.get("comment_count"))),
                "share_count": self._prefer_latest_int(note.share_count, self._safe_int(prev.get("share_count"))),
                "url": note.url or str(prev.get("url") or ""),
                "xsec_token": note.xsec_token or str(prev.get("xsec_token") or ""),
                "detail_text": note.detail_text or str(prev.get("detail_text") or ""),
                "comments_preview": note.comments_preview or str(prev.get("comments_preview") or ""),
                # Keep legacy fetched_at column for backward compatibility.
                "fetched_at": updated_at,
                "first_seen_at": first_seen_at,
                "updated_at": updated_at,
                "raw_json": note.raw_json or str(prev.get("raw_json") or ""),
            }

        rows = list(by_note.values())
        rows.sort(
            key=lambda x: (
                self._safe_int(x.get("publish_timestamp")),
                str(x.get("updated_at") or x.get("fetched_at") or ""),
                str(x.get("note_id") or ""),
            ),
            reverse=True,
        )
        return rows

    def _merge_summary_rows(self, ws, summaries: list[SummaryRecord]) -> list[dict]:
        existing = self._read_sheet_rows(ws, SUMMARY_HEADERS)
        by_note: dict[str, dict] = {str(r.get("note_id") or ""): r for r in existing if r.get("note_id")}

        for item in summaries:
            by_note[item.note_id] = {
                "run_id": item.run_id,
                "note_id": item.note_id,
                "keyword": item.keyword,
                "publish_time": item.publish_time.isoformat(),
                "publish_timestamp": int(item.publish_time.timestamp()),
                "title": item.title,
                "author": item.author,
                "summary": item.summary,
                "confidence": round(item.confidence, 4),
                "risk_flags": item.risk_flags,
                "url": item.url,
                "created_at": self._to_iso_utc(item.created_at),
            }

        rows = list(by_note.values())
        rows.sort(key=lambda x: int(x.get("publish_timestamp") or 0), reverse=True)
        return rows

    def _merge_send_rows(self, ws, send_logs: list[SendLogRecord]) -> list[dict]:
        existing = self._read_sheet_rows(ws, SEND_HEADERS)
        rows = list(existing)
        for log in send_logs:
            response = log.send_response
            if not isinstance(response, str):
                response = json.dumps(response, ensure_ascii=False)
            rows.append(
                {
                    "run_id": log.run_id,
                    "note_id": log.note_id,
                    "channel": log.channel,
                    "send_status": log.send_status,
                    "send_response": response[:20000],
                    "sent_at": self._to_iso_utc(log.sent_at),
                }
            )
        rows.sort(key=lambda x: str(x.get("sent_at") or ""), reverse=True)
        return rows

    def _merge_job_rows(self, ws, jobs: list[JobRecord]) -> list[dict]:
        existing = self._read_sheet_rows(ws, JOB_HEADERS)
        by_post: dict[str, dict] = {str(r.get("PostID") or ""): r for r in existing if r.get("PostID")}
        for job in jobs:
            by_post[job.post_id] = {
                "run_id": job.run_id,
                "Company": job.company,
                "Position": job.position,
                "publish_time": job.publish_time.isoformat(),
                "Location": job.location,
                "Requirements": job.requirements,
                "arrival_time": job.arrival_time,
                "application_method": job.application_method,
                "author": job.author,
                "risk_line": job.risk_line,
                "match_score": round(job.match_score, 2),
                "match_reason": job.match_reason,
                "Link": job.link,
                "PostID": job.post_id,
                "mode": job.mode,
                "comment_count": job.comment_count,
                "comments_preview": job.comments_preview,
                "original_text": job.original_text,
                "source_title": job.source_title,
                "opportunity_point": bool(job.opportunity_point),
                "outreach_message": job.outreach_message,
            }
        rows = list(by_post.values())
        rows.sort(
            key=lambda x: (
                self._iso_to_timestamp(x.get("publish_time")),
                str(x.get("PostID") or ""),
            ),
            reverse=True,
        )
        return rows

    @staticmethod
    def _safe_int(value) -> int:
        try:
            if value is None:
                return 0
            text = str(value).strip().replace(",", "")
            if not text:
                return 0
            return int(float(text))
        except Exception:
            return 0

    @staticmethod
    def _to_iso_utc(value: datetime) -> str:
        if not isinstance(value, datetime):
            return ""
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc).isoformat()
        return value.astimezone(timezone.utc).isoformat()

    @staticmethod
    def _is_relative_publish_text(text: str) -> bool:
        value = str(text or "").strip()
        if not value:
            return False
        patterns = (
            r"刚刚",
            r"刚才",
            r"\d+\s*秒前",
            r"\d+\s*分钟前",
            r"\d+\s*小时前",
            r"\d+\s*天前",
            r"\d+\s*澶╁墠",
            r"\d+\s*灏忔椂鍓?",
            r"\d+\s*鍒嗛挓鍓?",
        )
        return any(re.search(p, value) for p in patterns)

    @staticmethod
    def _prefer_latest_int(current: int, previous: int) -> int:
        if int(current or 0) > 0:
            return int(current)
        if int(previous or 0) > 0:
            return int(previous)
        return int(current or 0)

    @staticmethod
    def _iso_to_timestamp(value) -> int:
        if isinstance(value, datetime):
            return int(value.timestamp())
        text = str(value or "").strip()
        if not text:
            return 0
        try:
            return int(datetime.fromisoformat(text).timestamp())
        except Exception:
            return 0

    def export_jobs_csv(self, csv_path: str) -> None:
        if not self.path.exists():
            return
        wb = load_workbook(self.path, read_only=True)
        try:
            if "jobs" not in wb.sheetnames:
                return
            ws = wb["jobs"]
            rows = self._read_sheet_rows(ws, JOB_HEADERS)
        finally:
            wb.close()

        target = Path(csv_path)
        target.parent.mkdir(parents=True, exist_ok=True)
        with target.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(
                [
                    "Company",
                    "Position",
                    "publish_time",
                    "Location",
                    "Requirements",
                    "arrival_time",
                    "application_method",
                    "author",
                    "risk_line",
                    "match_score",
                    "Link",
                    "PostID",
                    "mode",
                    "opportunity_point",
                    "original_text",
                    "outreach_message",
                ]
            )
            for row in rows:
                writer.writerow(
                    [
                        row.get("Company", ""),
                        row.get("Position", ""),
                        row.get("publish_time", ""),
                        row.get("Location", ""),
                        row.get("Requirements", ""),
                        row.get("arrival_time", ""),
                        row.get("application_method", ""),
                        row.get("author", ""),
                        row.get("risk_line", ""),
                        row.get("match_score", ""),
                        row.get("Link", ""),
                        row.get("PostID", ""),
                        row.get("mode", ""),
                        row.get("opportunity_point", ""),
                        row.get("original_text", ""),
                        row.get("outreach_message", ""),
                    ]
                )
