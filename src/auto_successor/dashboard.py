from __future__ import annotations

import argparse
import json
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
import yaml


class DataBackend:
    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace
        self.data_dir = workspace / "data"
        self.runs_dir = self.data_dir / "runs"
        self.excel_path = self.data_dir / "output.xlsx"
        self.config_path = workspace / "config" / "config.yaml"

    def load_summary(self) -> dict[str, Any]:
        rows = self._load_workbook_rows()
        raw = rows.get("raw_notes", [])
        summaries = rows.get("succession_summary", [])
        jobs = rows.get("jobs", [])
        sends = rows.get("send_log", [])
        runs = self._load_runs(limit=1)

        latest_run_id = runs[0]["run_id"] if runs else ""
        latest_run_time = runs[0]["recorded_at"] if runs else ""

        return {
            "raw_count": len(raw),
            "summary_count": len(summaries),
            "jobs_count": len(jobs),
            "send_count": len(sends),
            "latest_run_id": latest_run_id,
            "latest_run_time": latest_run_time,
            "digest_interval_minutes": self._read_digest_interval(),
        }

    def load_leads(self, limit: int = 200, q: str = "") -> list[dict[str, Any]]:
        rows = self._load_workbook_rows()
        raw = rows.get("raw_notes", [])
        summaries = rows.get("succession_summary", [])
        jobs = rows.get("jobs", [])

        by_summary = {str(x.get("note_id") or ""): x for x in summaries}
        by_job = {str(x.get("PostID") or ""): x for x in jobs}

        merged: list[dict[str, Any]] = []
        for item in raw:
            note_id = str(item.get("note_id") or "")
            if not note_id:
                continue

            summary = by_summary.get(note_id, {})
            job = by_job.get(note_id, {})

            like_count = self._to_int(item.get("like_count"))
            comment_count = self._to_int(item.get("comment_count"))
            publish_ts = self._to_int(item.get("publish_timestamp"))
            status = self._resolve_status(
                like_count=like_count,
                comment_count=comment_count,
                has_summary=bool(summary),
                has_job=bool(job),
            )

            lead = {
                "note_id": note_id,
                "publish_time": str(item.get("publish_time") or ""),
                "publish_time_text": str(item.get("publish_time_text") or ""),
                "publish_timestamp": publish_ts,
                "title": str(item.get("title") or ""),
                "author": str(item.get("author") or ""),
                "like_count": like_count,
                "comment_count": comment_count,
                "share_count": self._to_int(item.get("share_count")),
                "url": str(item.get("url") or ""),
                "detail_text": str(item.get("detail_text") or ""),
                "comments_preview": str(item.get("comments_preview") or ""),
                "summary": str(summary.get("summary") or ""),
                "risk_flags": str(summary.get("risk_flags") or ""),
                "company": str(job.get("Company") or ""),
                "position": str(job.get("Position") or ""),
                "location": str(job.get("Location") or ""),
                "requirements": str(job.get("Requirements") or ""),
                "status": status,
            }
            merged.append(lead)

        merged.sort(key=lambda x: int(x.get("publish_timestamp") or 0), reverse=True)

        if q:
            key = q.strip().lower()
            merged = [
                row
                for row in merged
                if key
                in " ".join(
                    [
                        str(row.get("title") or ""),
                        str(row.get("author") or ""),
                        str(row.get("summary") or ""),
                        str(row.get("company") or ""),
                        str(row.get("position") or ""),
                        str(row.get("location") or ""),
                    ]
                ).lower()
            ]

        return merged[: max(1, limit)]

    def load_runs(self, limit: int = 20) -> list[dict[str, Any]]:
        return self._load_runs(limit=limit)

    def _load_runs(self, limit: int) -> list[dict[str, Any]]:
        if not self.runs_dir.exists():
            return []

        out: list[dict[str, Any]] = []
        files = sorted(self.runs_dir.glob("*.json"), reverse=True)
        for file in files[: max(1, limit)]:
            try:
                payload = json.loads(file.read_text(encoding="utf-8"))
            except Exception:
                continue
            out.append(
                {
                    "run_id": str(payload.get("run_id") or file.stem),
                    "recorded_at": str(payload.get("recorded_at") or ""),
                    "fetched": self._to_int(payload.get("fetched")),
                    "target_notes": self._to_int(payload.get("target_notes")),
                    "jobs": self._to_int(payload.get("jobs")),
                    "send_logs": self._to_int(payload.get("send_logs")),
                    "digest_sent": bool(payload.get("digest_sent")),
                }
            )
        return out

    def _load_workbook_rows(self) -> dict[str, list[dict[str, Any]]]:
        if not self.excel_path.exists():
            return {"raw_notes": [], "succession_summary": [], "jobs": [], "send_log": []}

        wb = load_workbook(self.excel_path, read_only=True)
        try:
            out: dict[str, list[dict[str, Any]]] = {}
            for name in ("raw_notes", "succession_summary", "jobs", "send_log"):
                if name not in wb.sheetnames:
                    out[name] = []
                    continue
                out[name] = self._read_sheet_rows(wb[name])
            return out
        finally:
            wb.close()

    @staticmethod
    def _read_sheet_rows(ws) -> list[dict[str, Any]]:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return []
        headers = [str(x or "").strip() for x in headers_row]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, Any] = {}
            for idx, value in enumerate(values):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if key:
                    row[key] = value
            rows.append(row)
        return rows

    @staticmethod
    def _resolve_status(like_count: int, comment_count: int, has_summary: bool, has_job: bool) -> str:
        score = like_count + comment_count * 2
        if has_job and score >= 20:
            return "高优先级"
        if has_job:
            return "可推送"
        if has_summary:
            return "待复核"
        return "新线索"

    @staticmethod
    def _to_int(value: Any) -> int:
        try:
            if value is None:
                return 0
            text = str(value).strip().replace(",", "")
            if not text:
                return 0
            return int(float(text))
        except Exception:
            return 0

    def _read_digest_interval(self) -> int:
        if not self.config_path.exists():
            return 60
        try:
            data = yaml.safe_load(self.config_path.read_text(encoding="utf-8")) or {}
            notification = data.get("notification", {}) if isinstance(data, dict) else {}
            raw = notification.get("digest_interval_minutes", 60)
            return max(1, int(raw))
        except Exception:
            return 60


def make_handler(backend: DataBackend, web_dir: Path):
    class DashboardHandler(SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=str(web_dir), **kwargs)

        def do_GET(self) -> None:
            parsed = urlparse(self.path)
            path = parsed.path
            if path.startswith("/api/"):
                self._handle_api(path=path, query=parse_qs(parsed.query))
                return
            if path == "/":
                self.path = "/index.html"
            return super().do_GET()

        def _handle_api(self, path: str, query: dict[str, list[str]]) -> None:
            try:
                if path == "/api/health":
                    self._json({"ok": True, "time": datetime.now().isoformat()})
                    return

                if path == "/api/summary":
                    self._json(backend.load_summary())
                    return

                if path == "/api/leads":
                    limit = int((query.get("limit") or ["200"])[0])
                    q = (query.get("q") or [""])[0]
                    self._json({"items": backend.load_leads(limit=limit, q=q)})
                    return

                if path == "/api/runs":
                    limit = int((query.get("limit") or ["20"])[0])
                    self._json({"items": backend.load_runs(limit=limit)})
                    return

                self._json({"error": "not_found", "path": path}, status=HTTPStatus.NOT_FOUND)
            except Exception as exc:
                self._json(
                    {"error": "server_error", "message": str(exc)},
                    status=HTTPStatus.INTERNAL_SERVER_ERROR,
                )

        def _json(self, payload: Any, status: HTTPStatus = HTTPStatus.OK) -> None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status.value)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type, Accept")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_OPTIONS(self) -> None:
            self.send_response(HTTPStatus.NO_CONTENT.value)
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type, Accept")
            self.end_headers()

    return DashboardHandler


def main() -> None:
    parser = argparse.ArgumentParser(description="SuccessionPilot dashboard server")
    parser.add_argument("--host", default="127.0.0.1", help="Bind host")
    parser.add_argument("--port", type=int, default=8787, help="Bind port")
    args = parser.parse_args()

    workspace = Path.cwd()
    web_dir = workspace / "web"
    if not web_dir.exists():
        raise FileNotFoundError(f"Web directory not found: {web_dir}")

    backend = DataBackend(workspace=workspace)
    handler = make_handler(backend=backend, web_dir=web_dir)
    server = ThreadingHTTPServer((args.host, args.port), handler)
    print(f"Dashboard running on http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
