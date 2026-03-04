from __future__ import annotations

import json
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml

from .config import ResumeConfig, load_settings
from .retry_queue import RetryQueue
from .resume_loader import ResumeLoader


class DashboardRepository:
    def _resolve_retry_queue_path(self) -> Path:
        default_path = self.data_dir / "retry_queue.json"
        try:
            settings = load_settings(str(self.config_path))
            candidate = Path(str(settings.storage.retry_queue_path or "").strip() or str(default_path))
        except Exception:
            candidate = default_path
        if not candidate.is_absolute():
            candidate = self.workspace / candidate
        return candidate

    def _load_retry_queue(self) -> RetryQueue:
        return RetryQueue(path=str(self.retry_queue_path))

    @staticmethod
    def _safe_run_id(run_id: str) -> str:
        text = str(run_id or "").strip()
        if not text:
            return ""
        if not re.fullmatch(r"[A-Za-z0-9._:-]+", text):
            return ""
        return text

    def _resolve_xhs_script_path(self, args: Any) -> Path:
        values: list[str] = []
        if isinstance(args, list):
            values = [str(item).strip() for item in args if str(item).strip()]
        elif isinstance(args, str):
            text = args.strip()
            if text:
                values = [text]
        first = values[0] if values else "vendor/xhs-mcp/dist/xhs-mcp.js"
        script = Path(first)
        if not script.is_absolute():
            script = self.workspace / script
        return script

    @staticmethod
    def _resolve_selected_xhs_cookie_file(*, account: str, account_cookies_dir: str) -> Path:
        account_name = str(account or "default").strip() or "default"
        if account_name.lower() == "default":
            return Path.home() / ".xhs-mcp" / "cookies.json"

        base = Path(str(account_cookies_dir or "~/.xhs-mcp/accounts")).expanduser()
        if account_name.lower().endswith(".json"):
            flat = base / account_name
            nested = base / account_name[: -len(".json")] / "cookies.json"
        else:
            flat = base / f"{account_name}.json"
            nested = base / account_name / "cookies.json"
        if flat.exists():
            return flat
        if nested.exists():
            return nested
        return flat

    def _read_env_values(self) -> dict[str, str]:
        env_path = self.workspace / ".env"
        out: dict[str, str] = {}
        if not env_path.exists():
            return out
        try:
            for line in env_path.read_text(encoding="utf-8").splitlines():
                text = str(line or "").strip()
                if not text or text.startswith("#") or "=" not in text:
                    continue
                key, value = text.split("=", 1)
                key = key.strip()
                value = value.strip().strip("\"'")
                if key:
                    out[key] = value
        except Exception:
            return out
        return out

    @staticmethod
    def _get_env_value(name: str, env_map: dict[str, str]) -> str:
        key = str(name or "").strip()
        if not key:
            return ""
        runtime = str(os.getenv(key, "") or "").strip()
        if runtime:
            return runtime
        return str(env_map.get(key, "") or "").strip()

    @staticmethod
    def _compact_json(value: Any) -> str:
        try:
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))[:280]
        except Exception:
            return str(value)[:280]

    def _load_runs(self, limit: int) -> list[dict[str, Any]]:
        if not self.runs_dir.exists():
            return []

        out: list[dict[str, Any]] = []
        files = sorted(self.runs_dir.glob("*.json"), reverse=True)
        for file in files[: max(1, limit)]:
            try:
                payload = json.loads(file.read_text(encoding="utf-8"))
            except Exception:
                continue
            stats = payload.get("stats")
            if not isinstance(stats, dict):
                stats = {}
            stage_records = payload.get("stage_records")
            if not isinstance(stage_records, list):
                stage_records = []

            slow_stages = stats.get("stage_top_slow")
            if not isinstance(slow_stages, list):
                normalized_stages: list[dict[str, Any]] = []
                for item in stage_records:
                    if not isinstance(item, dict):
                        continue
                    normalized_stages.append(
                        {
                            "name": str(item.get("name") or ""),
                            "duration_ms": self._to_int(item.get("duration_ms")),
                            "status": str(item.get("status") or ""),
                        }
                    )
                slow_stages = sorted(normalized_stages, key=lambda x: self._to_int(x.get("duration_ms")), reverse=True)[:3]

            stage_total_ms = self._to_int(stats.get("stage_total_ms"))
            if stage_total_ms <= 0 and stage_records:
                stage_total_ms = sum(self._to_int(item.get("duration_ms")) for item in stage_records if isinstance(item, dict))

            stage_avg_ms = self._to_int(stats.get("stage_avg_ms"))
            if stage_avg_ms <= 0 and stage_records:
                stage_avg_ms = int(stage_total_ms / max(1, len(stage_records)))

            stage_failed_count = self._to_int(stats.get("stage_failed_count"))
            if stage_failed_count <= 0 and stage_records:
                stage_failed_count = sum(
                    1
                    for item in stage_records
                    if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "failed"
                )

            llm_error_codes = stats.get("llm_error_codes")
            if not isinstance(llm_error_codes, dict):
                llm_error_codes = {}
            stage_error_codes = stats.get("stage_error_codes")
            if not isinstance(stage_error_codes, dict):
                stage_error_codes = {}
                for item in stage_records:
                    if not isinstance(item, dict):
                        continue
                    if str(item.get("status") or "").strip().lower() != "failed":
                        continue
                    code = str(item.get("error_code") or "").strip().lower() or "stage_failed"
                    stage_error_codes[code] = self._to_int(stage_error_codes.get(code)) + 1

            merged_error_codes: dict[str, int] = {}
            for source in (llm_error_codes, stage_error_codes):
                for key, value in source.items():
                    code = str(key or "").strip().lower()
                    if not code:
                        continue
                    merged_error_codes[code] = self._to_int(merged_error_codes.get(code)) + self._to_int(value)

            alerts_triggered_raw = stats.get("alerts_triggered")
            if isinstance(alerts_triggered_raw, list):
                alerts_triggered = [item for item in alerts_triggered_raw if isinstance(item, dict)]
            else:
                alerts_triggered = []
            alert_codes = [
                str(item.get("code") or "").strip().lower()
                for item in alerts_triggered
                if str(item.get("code") or "").strip()
            ]
            alerts_notified_raw = stats.get("alerts_notified")
            if isinstance(alerts_notified_raw, list):
                alerts_notified = [str(item).strip().lower() for item in alerts_notified_raw if str(item).strip()]
            else:
                alerts_notified = []

            out.append(
                {
                    "run_id": str(payload.get("run_id") or file.stem),
                    "recorded_at": str(payload.get("recorded_at") or ""),
                    "mode": str(payload.get("mode") or stats.get("mode") or ""),
                    "notification_mode": str(payload.get("notification_mode") or stats.get("notification_mode") or ""),
                    "fetched": self._to_int(stats.get("fetched") if stats else payload.get("fetched")),
                    "target_notes": self._to_int(stats.get("target_notes") if stats else payload.get("target_notes")),
                    "jobs": self._to_int(stats.get("jobs") if stats else payload.get("jobs")),
                    "send_logs": self._to_int(stats.get("send_logs") if stats else payload.get("send_logs")),
                    "digest_sent": bool(stats.get("digest_sent") if stats else payload.get("digest_sent")),
                    "llm_fail": self._to_int(stats.get("llm_fail")),
                    "llm_calls": self._to_int(stats.get("llm_calls")),
                    "process_workers": self._to_int(stats.get("process_workers")),
                    "detail_workers": self._to_int(stats.get("detail_workers")),
                    "stage_total_ms": stage_total_ms,
                    "stage_avg_ms": stage_avg_ms,
                    "stage_failed_count": stage_failed_count,
                    "slow_stages": slow_stages,
                    "error_codes": merged_error_codes,
                    "retry_pending": stats.get("retry_pending") if isinstance(stats.get("retry_pending"), dict) else {},
                    "retry_running": stats.get("retry_running") if isinstance(stats.get("retry_running"), dict) else {},
                    "retry_enqueued": self._to_int(stats.get("retry_enqueued")),
                    "retry_retried": self._to_int(stats.get("retry_retried")),
                    "retry_succeeded": self._to_int(stats.get("retry_succeeded")),
                    "retry_dropped": self._to_int(stats.get("retry_dropped")),
                    "fetch_fail_count_run": self._to_int(stats.get("fetch_fail_count_run")),
                    "fetch_fail_streak": self._to_int(stats.get("fetch_fail_streak")),
                    "xhs_data_empty": bool(stats.get("xhs_data_empty")),
                    "detail_attempted": self._to_int(stats.get("detail_attempted")),
                    "detail_target_notes": self._to_int(stats.get("detail_target_notes")),
                    "detail_success": self._to_int(stats.get("detail_success")),
                    "detail_failed": self._to_int(stats.get("detail_failed")),
                    "detail_filled": self._to_int(stats.get("detail_filled")),
                    "detail_missing": self._to_int(stats.get("detail_missing")),
                    "detail_missing_rate": float(stats.get("detail_missing_rate") or 0.0),
                    "llm_timeout_count": self._to_int(stats.get("llm_timeout_count")),
                    "llm_timeout_rate": float(stats.get("llm_timeout_rate") or 0.0),
                    "alerts_triggered_count": self._to_int(
                        stats.get("alerts_triggered_count")
                        if "alerts_triggered_count" in stats
                        else len(alert_codes)
                    ),
                    "alerts_notified_count": self._to_int(
                        stats.get("alerts_notified_count")
                        if "alerts_notified_count" in stats
                        else len(alerts_notified)
                    ),
                    "alert_codes": alert_codes,
                    "alerts_triggered": alerts_triggered,
                    "alerts_notified": alerts_notified,
                    "xhs_diagnosis": stats.get("xhs_diagnosis") if isinstance(stats.get("xhs_diagnosis"), dict) else {},
                }
            )
        return out

    def _load_workbook_rows(self) -> dict[str, list[dict[str, Any]]]:
        if not self.excel_path.exists():
            return {"raw_notes": [], "succession_summary": [], "jobs": [], "send_log": []}

        wb = load_workbook(self.excel_path, read_only=True)
        try:
            out: dict[str, list[dict[str, Any]]] = {}
            for name in ("raw_notes", "succession_summary", "jobs", "send_log"):
                if name not in wb.sheetnames:
                    out[name] = []
                    continue
                out[name] = self._read_sheet_rows(wb[name])
            return out
        finally:
            wb.close()

    @staticmethod
    def _read_sheet_rows(ws) -> list[dict[str, Any]]:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return []
        headers = [str(x or "").strip() for x in headers_row]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, Any] = {}
            for idx, value in enumerate(values):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if key:
                    row[key] = value
            rows.append(row)
        return rows

    @staticmethod
    def _normalize_edit_text(value: Any) -> str:
        text = str(value or "")
        return text.replace("\r\n", "\n").replace("\r", "\n").strip()

    @staticmethod
    def _append_row(ws) -> int:
        return max(2, ws.max_row + 1)

    @staticmethod
    def _header_map(ws) -> dict[str, int]:
        max_col = max(1, ws.max_column)
        mapping: dict[str, int] = {}
        for col in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col).value or "").strip()
            if header:
                mapping[header] = col
        return mapping

    def _ensure_sheet_with_headers(self, wb, sheet_name: str, expected_headers: list[str]):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        header_map = self._header_map(ws)
        if not header_map:
            for idx, key in enumerate(expected_headers, start=1):
                ws.cell(row=1, column=idx, value=key)
            return ws, {key: idx for idx, key in enumerate(expected_headers, start=1)}
        for key in expected_headers:
            if key in header_map:
                continue
            col = max(header_map.values(), default=0) + 1
            ws.cell(row=1, column=col, value=key)
            header_map[key] = col
        return ws, header_map

    @staticmethod
    def _find_row_index(ws, headers: dict[str, int], *, key_column: str, key_value: str) -> int:
        col = headers.get(key_column)
        if not col:
            return 0
        target = str(key_value or "").strip()
        if not target:
            return 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col).value
            if str(cell or "").strip() == target:
                return row
        return 0

    @staticmethod
    def _set_cell(ws, headers: dict[str, int], row: int, column_name: str, value: Any) -> None:
        col = headers.get(column_name)
        if not col:
            return
        ws.cell(row=row, column=col, value=value)

    @staticmethod
    def _row_snapshot(ws, headers: dict[str, int], row: int) -> dict[str, Any]:
        out: dict[str, Any] = {}
        for key, col in headers.items():
            out[key] = ws.cell(row=row, column=col).value
        return out

    @staticmethod
    def _resolve_status_key(like_count: int, comment_count: int, has_summary: bool, has_job: bool) -> str:
        score = like_count + comment_count * 2
        if has_job and score >= 20:
            return "high_priority"
        if has_job:
            return "actionable"
        if has_summary:
            return "pending_review"
        return "new_lead"

    @classmethod
    def _resolve_status(cls, like_count: int, comment_count: int, has_summary: bool, has_job: bool) -> str:
        key = cls._resolve_status_key(
            like_count=like_count,
            comment_count=comment_count,
            has_summary=has_summary,
            has_job=has_job,
        )
        labels = {
            "high_priority": "高优先级",
            "actionable": "可推进",
            "pending_review": "待复核",
            "new_lead": "新线索",
        }
        return labels.get(key, "新线索")

    @staticmethod
    def _to_int(value: Any) -> int:
        try:
            if value is None:
                return 0
            text = str(value).strip().replace(",", "")
            if not text:
                return 0
            return int(float(text))
        except Exception:
            return 0

    @staticmethod
    def _to_epoch(value: Any) -> int:
        if isinstance(value, datetime):
            try:
                if value.tzinfo is None:
                    value = value.replace(tzinfo=timezone.utc)
                return int(value.timestamp())
            except Exception:
                return 0
        text = str(value or "").strip()
        if not text:
            return 0
        try:
            dt = datetime.fromisoformat(text)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return int(dt.timestamp())
        except Exception:
            return 0

    @staticmethod
    def _format_time_from_epoch_or_iso(epoch: int, iso_text: str) -> str:
        try:
            if int(epoch) > 0:
                dt = datetime.fromtimestamp(int(epoch))
                return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass
        text = str(iso_text or "").strip()
        if text:
            try:
                dt = datetime.fromisoformat(text)
                return dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                return text
        return "-"

    def _read_digest_interval(self) -> int:
        if not self.config_path.exists():
            return 60
        try:
            data = yaml.safe_load(self.config_path.read_text(encoding="utf-8")) or {}
            notification = data.get("notification", {}) if isinstance(data, dict) else {}
            raw = notification.get("digest_interval_minutes", 60)
            return max(1, int(raw))
        except Exception:
            return 60

    def _read_config_data(self) -> dict[str, Any]:
        if not self.config_path.exists():
            return {}
        data = yaml.safe_load(self.config_path.read_text(encoding="utf-8")) or {}
        if isinstance(data, dict):
            return data
        return {}

    def _get_resume_loader(self) -> ResumeLoader:
        if self._resume_loader is not None:
            return self._resume_loader
        config = self._read_config_data()
        resume_cfg_raw = config.get("resume", {}) if isinstance(config, dict) else {}
        if not isinstance(resume_cfg_raw, dict):
            resume_cfg_raw = {}

        source_txt_path = str(resume_cfg_raw.get("source_txt_path") or "config/resume.txt").strip() or "config/resume.txt"
        resume_text_path = str(resume_cfg_raw.get("resume_text_path") or "data/resume_text.txt").strip() or "data/resume_text.txt"
        max_chars = self._coerce_int(resume_cfg_raw.get("max_chars"), 6000, minimum=500)

        source = Path(source_txt_path)
        if not source.is_absolute():
            source = self.workspace / source
        target = Path(resume_text_path)
        if not target.is_absolute():
            target = self.workspace / target

        self._resume_source_path = source
        self._resume_text_path = target
        cfg = ResumeConfig(source_txt_path=str(source), resume_text_path=str(target), max_chars=max_chars)
        self._resume_loader = ResumeLoader(cfg, logger=self)
        return self._resume_loader

    def warning(self, *args, **kwargs) -> None:
        return

    def _write_config_data(self, data: dict[str, Any]) -> None:
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        text = yaml.safe_dump(data, sort_keys=False, allow_unicode=True)
        self.config_path.write_text(text, encoding="utf-8")

    @staticmethod
    def _ensure_section(root: dict[str, Any], key: str) -> dict[str, Any]:
        item = root.get(key)
        if not isinstance(item, dict):
            item = {}
            root[key] = item
        return item

    @staticmethod
    def _coerce_int(value: Any, default: int, minimum: int = 0) -> int:
        try:
            parsed = int(str(value).strip())
        except Exception:
            parsed = int(default)
        return max(minimum, parsed)

    @staticmethod
    def _coerce_rate(value: Any, default: float) -> float:
        try:
            parsed = float(value)
        except Exception:
            parsed = float(default)
        if parsed > 1.0 and parsed <= 100.0:
            parsed = parsed / 100.0
        return max(0.0, min(1.0, parsed))

    @staticmethod
    def _coerce_float(value: Any, default: float, minimum: float = 0.0) -> float:
        try:
            parsed = float(value)
        except Exception:
            parsed = float(default)
        if parsed < float(minimum):
            parsed = float(minimum)
        return parsed

    @staticmethod
    def _normalize_mode(value: str) -> str:
        mode = (value or "auto").strip().lower()
        if mode == "smart":
            mode = "agent"
        if mode not in {"auto", "agent"}:
            return "auto"
        return mode

    @staticmethod
    def _normalize_notify_mode(value: str) -> str:
        mode = (value or "digest").strip().lower()
        if mode not in {"digest", "realtime", "off"}:
            return "digest"
        return mode

    @staticmethod
    def _to_name_list(value: Any, default: list[str]) -> list[str]:
        if isinstance(value, list):
            items = [str(item).strip() for item in value if str(item).strip()]
            return items or list(default)
        if isinstance(value, str):
            items = [x.strip() for x in value.split(",") if x.strip()]
            return items or list(default)
        return list(default)
