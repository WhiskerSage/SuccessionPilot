from __future__ import annotations

import hashlib
import json
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .agents import CommunicationAgent, IntelligenceAgent, PlannerAgent
from .config import Settings
from .email_sender import EmailSender
from .excel_store import ExcelStore, JOB_HEADERS
from .llm_client import LLMClient
from .llm_enricher import LLMEnricher
from .models import JobRecord, NoteRecord, SummaryRecord
from .notification_router import NotificationChannel, NotificationRouter
from .retry_queue import RetryQueue
from .resume_loader import ResumeLoader
from .run_journal import RunJournal
from .run_lock import RunLock
from .runtime_orchestrator import RuntimeOrchestrator
from .state_store import StateStore
from .wechat_service_sender import WeChatServiceSender
from .xhs_collector import XHSMcpCliCollector


class AutoSuccessorPipeline:
    """
    SuccessionPilot runtime:
    - collector layer
    - agent layer (planner/intelligence/communication)
    - storage + snapshot
    """

    def __init__(self, settings: Settings, logger) -> None:
        self.settings = settings
        self.logger = logger
        self.collector = XHSMcpCliCollector(settings.xhs, settings.app.timezone, logger)
        self.store = ExcelStore(settings.storage.excel_path)
        self.state = StateStore(settings.storage.state_path)
        self.journal = RunJournal()
        self.lock = RunLock()
        self.llm_client = LLMClient(settings=settings, logger=logger)
        self.llm_enricher = LLMEnricher(llm_client=self.llm_client, settings=settings, logger=logger)
        self.resume_loader = ResumeLoader(settings.resume, logger)
        self.retry_queue = RetryQueue(
            path=settings.storage.retry_queue_path,
            base_backoff_seconds=settings.retry.base_backoff_seconds,
            max_backoff_seconds=settings.retry.max_backoff_seconds,
            max_attempts_by_type={
                "fetch": settings.retry.fetch_max_attempts,
                "llm_timeout": settings.retry.llm_timeout_max_attempts,
                "email": settings.retry.email_max_attempts,
            },
        )
        self._retry_enabled = bool(settings.retry.enabled)
        self._retry_worker_interval = max(5, int(settings.retry.worker_interval_seconds))
        self._retry_batch_size = max(1, int(settings.retry.replay_batch_size))
        self._retry_stop = threading.Event()
        self._retry_worker: threading.Thread | None = None
        self._fetch_fail_streak = 0

        channels = [
            NotificationChannel(name="wechat_service", sender=WeChatServiceSender(settings, logger)),
            NotificationChannel(name="email", sender=EmailSender(settings, logger)),
        ]
        self.router = NotificationRouter(channels=channels, logger=logger)
        self.planner = PlannerAgent(settings=settings, logger=logger)
        self.intelligence = IntelligenceAgent(llm_enricher=self.llm_enricher, logger=logger)
        self.communication = CommunicationAgent(
            router=self.router,
            settings=settings,
            llm_enricher=self.llm_enricher,
            logger=logger,
        )
        if self._retry_enabled:
            self._retry_worker = threading.Thread(target=self._retry_worker_loop, name="retry-worker", daemon=True)
            self._retry_worker.start()

    def run_once(self, run_id: str, mode: str = "auto") -> dict:
        keyword = self.settings.xhs.keyword
        mode = self._normalize_mode(mode or self.settings.agent.mode or "auto")
        notification_mode = self._normalize_notification_mode(self.settings.notification.mode)
        orchestrator = RuntimeOrchestrator(runtime_name=self.settings.agent.runtime_name, logger=self.logger)

        if not self.lock.acquire():
            self.logger.warning(
                "跳过运行 | run=%s | 原因=检测到上一次运行锁未释放 | lock=%s",
                run_id,
                self.lock.path,
            )
            return {"skipped": True, "reason": "run_locked"}

        try:
            self.logger.info("运行开始 | run=%s | keyword=%s | mode=%s | notify=%s", run_id, keyword, mode, notification_mode)
            self._log_progress(run_id, 5, "初始化运行上下文")
            self.llm_enricher.reset_stats()
            self.llm_client.clear_error_counts()
            fetch_fail_events: list[dict[str, str]] = []
            detail_enrich_stats: dict[str, int] = {
                "target_notes": 0,
                "attempted": 0,
                "success": 0,
                "failed": 0,
                "skipped_no_token": 0,
                "detail_filled": 0,
                "detail_missing": 0,
                "blocked": 0,
            }
            xhs_failure_diagnosis: dict[str, Any] = {}
            xhs_data_empty = False

            resume_text = orchestrator.run_stage(
                "resume.load_text",
                lambda: self.resume_loader.load_resume_text(),
                meta={"resume_source": self.settings.resume.source_txt_path},
            )

            try:
                orchestrator.run_stage(
                    "collector.ensure_logged_in",
                    lambda: self.collector.ensure_logged_in(),
                    meta={"browser_path": self.settings.xhs.browser_path},
                )
            except Exception as exc:
                fetch_fail_events.append({"stage": "collector.ensure_logged_in", "error": str(exc)[:280]})
                self._enqueue_retry(
                    queue_type="fetch",
                    action="ensure_logged_in",
                    run_id=run_id,
                    error=str(exc),
                    payload={
                        "browser_path": self.settings.xhs.browser_path,
                        "account": self.settings.xhs.account,
                    },
                    dedupe_key=f"ensure-login:{self.settings.xhs.account}",
                )
            self._log_progress(run_id, 12, "登录态检查完成")

            try:
                notes = orchestrator.run_stage(
                    "collector.search_notes",
                    lambda: self.collector.search_notes(
                        run_id=run_id,
                        keyword=keyword,
                        max_results=self.settings.xhs.max_results,
                    ),
                    meta={"keyword": keyword, "max_results": self.settings.xhs.max_results},
                )
            except Exception as exc:
                notes = []
                fetch_fail_events.append({"stage": "collector.search_notes", "error": str(exc)[:280]})
                self._enqueue_retry(
                    queue_type="fetch",
                    action="search_notes",
                    run_id=run_id,
                    error=str(exc),
                    payload={
                        "keyword": keyword,
                        "max_results": self.settings.xhs.max_results,
                    },
                    dedupe_key=f"search:{keyword}:{self.settings.xhs.max_results}",
                )
                self.logger.warning("搜索失败已入重试队列 | run=%s | error=%s", run_id, exc)
            notes = sorted(notes, key=lambda n: (n.publish_time, n.note_id), reverse=True)
            xhs_data_empty = len(notes) == 0 and not fetch_fail_events
            self._log_progress(run_id, 28, f"搜索完成，抓取到 {len(notes)} 条")

            pre_new_notes = orchestrator.run_stage(
                "state.prefilter_new_notes",
                lambda: [note for note in notes if not self.state.has(note.note_id)],
                meta={"existing_state_size": len(self.state.processed_note_ids)},
            )
            updated_existing_notes = max(0, len(notes) - len(pre_new_notes))
            self._log_progress(run_id, 36, f"预过滤完成，候选新增 {len(pre_new_notes)} 条")

            initial_plan = self.planner.build_plan(mode=mode, fetched_count=len(notes), new_count=len(pre_new_notes))
            try:
                detail_enrich_stats = orchestrator.run_stage(
                    "collector.enrich_note_details",
                    lambda: self.collector.enrich_note_details(pre_new_notes, max_notes=initial_plan.detail_fetch_limit),
                    meta={
                        "max_detail_fetch": initial_plan.detail_fetch_limit,
                        "detail_workers": int(getattr(self.settings.xhs, "detail_workers", 1)),
                    },
                )
                if not isinstance(detail_enrich_stats, dict):
                    detail_enrich_stats = {
                        "target_notes": max(0, int(initial_plan.detail_fetch_limit)),
                        "attempted": 0,
                        "success": 0,
                        "failed": 0,
                        "skipped_no_token": 0,
                        "detail_filled": 0,
                        "detail_missing": 0,
                        "blocked": 0,
                    }
            except Exception as exc:
                fetch_fail_events.append({"stage": "collector.enrich_note_details", "error": str(exc)[:280]})
                queued = 0
                for note in pre_new_notes[: max(0, int(initial_plan.detail_fetch_limit))]:
                    if not str(getattr(note, "xsec_token", "") or "").strip():
                        continue
                    self._enqueue_retry(
                        queue_type="fetch",
                        action="enrich_note_detail",
                        run_id=run_id,
                        error=str(exc),
                        payload={"note_id": note.note_id, "xsec_token": note.xsec_token},
                        dedupe_key=f"detail:{note.note_id}",
                    )
                    queued += 1
                self.logger.warning("详情抓取失败已入队 | run=%s | queued=%s | error=%s", run_id, queued, exc)
                detail_enrich_stats = {
                    "target_notes": max(0, int(initial_plan.detail_fetch_limit)),
                    "attempted": 0,
                    "success": 0,
                    "failed": 1,
                    "skipped_no_token": 0,
                    "detail_filled": 0,
                    "detail_missing": 0,
                    "blocked": 0,
                }
            self._log_progress(
                run_id,
                46,
                (
                    f"正文补全完成，详情抓取上限 {initial_plan.detail_fetch_limit}，"
                    f"并行 {int(getattr(self.settings.xhs, 'detail_workers', 1))}"
                ),
            )

            if fetch_fail_events:
                self._fetch_fail_streak += 1
            else:
                self._fetch_fail_streak = 0

            if self._fetch_fail_streak >= 2:
                xhs_failure_diagnosis = self._probe_xhs_fetch_failure(run_id=run_id, fetch_fail_events=fetch_fail_events)

            new_notes = orchestrator.run_stage(
                "state.filter_new_notes",
                lambda: [note for note in pre_new_notes if not self.state.has(note.note_id)],
                meta={"prefiltered_count": len(pre_new_notes)},
            )
            new_notes = sorted(new_notes, key=lambda n: (n.publish_time, n.note_id), reverse=True)
            self._log_progress(run_id, 55, f"增量过滤完成，新增 {len(new_notes)} 条")

            plan = orchestrator.run_stage(
                "agent.planner.build_plan",
                lambda: self.planner.build_plan(mode=mode, fetched_count=len(notes), new_count=len(new_notes)),
                meta={"mode": mode},
            )
            process_workers = max(1, int(getattr(self.settings.pipeline, "process_workers", 1)))
            self.logger.info("提取并行配置 | run=%s | process_workers=%s", run_id, process_workers)
            use_single_pass_extract = bool(getattr(self.settings.llm, "single_pass_extract", True)) and bool(
                self.settings.llm.enabled
            )
            note_agent_stats: dict[str, int] = {}
            note_agent_details: list[dict[str, Any]] = []
            if use_single_pass_extract:
                extract_outcome = orchestrator.run_stage(
                    "agent.intelligence.process_notes_with_agents",
                    lambda: self.intelligence.process_notes_with_agents(
                        notes=new_notes,
                        resume_text=resume_text,
                        mode=mode,
                        workers=process_workers,
                    ),
                    meta={
                        "job_extract_llm_budget": "all",
                        "single_pass_extract": True,
                        "process_workers": process_workers,
                    },
                )
                target_notes = extract_outcome.targets
                filtered_out = extract_outcome.filtered_out
                jobs = extract_outcome.jobs
                note_agent_stats = dict(getattr(extract_outcome, "note_agent_stats", {}) or {})
                note_agent_details = list(getattr(extract_outcome, "note_agent_details", []) or [])
                llm_stage_calls = self.llm_enricher.stage_call_counts()
                llm_stage_fallbacks = self.llm_enricher.stage_fallback_counts()
                self._log_progress(
                    run_id,
                    66,
                    (
                        f"直接提取完成，命中 {len(target_notes)} 条，过滤 {len(filtered_out)} 条，"
                        f"LLM提取调用 {int(llm_stage_calls.get('job', 0))} 次，"
                        f"规则回退 {int(llm_stage_fallbacks.get('job', 0))} 次，"
                        f"每帖Agent回退 {int(note_agent_stats.get('worker_fallback', 0))} 次"
                    ),
                )
            else:
                filter_outcome = orchestrator.run_stage(
                    "agent.intelligence.filter_target_notes",
                    lambda: self.intelligence.filter_target_notes(
                        new_notes,
                        max_filter_items=plan.max_filter_items,
                        workers=process_workers,
                    ),
                    meta={"max_filter_items": plan.max_filter_items, "process_workers": process_workers},
                )
                target_notes = filter_outcome.targets
                filtered_out = filter_outcome.filtered_out
                llm_stage_calls = self.llm_enricher.stage_call_counts()
                llm_stage_fallbacks = self.llm_enricher.stage_fallback_counts()
                self._log_progress(
                    run_id,
                    66,
                    (
                        f"目标筛选完成，命中 {len(target_notes)} 条，过滤 {len(filtered_out)} 条，"
                        f"LLM筛选调用 {int(llm_stage_calls.get('filter', 0))} 次，"
                        f"规则回退 {int(llm_stage_fallbacks.get('filter', 0))} 次"
                    ),
                )

                jobs = orchestrator.run_stage(
                    "agent.intelligence.build_jobs",
                    lambda: self.intelligence.build_jobs(
                        target_notes,
                        max_job_items=plan.max_job_items,
                        resume_text=resume_text,
                        mode=mode,
                        workers=process_workers,
                    ),
                    meta={"max_job_items": plan.max_job_items, "process_workers": process_workers},
                )

            jobs = orchestrator.run_stage(
                "agent.intelligence.mark_opportunities",
                lambda: self.intelligence.mark_opportunities(jobs),
                meta={"jobs": len(jobs)},
            )

            jobs = orchestrator.run_stage(
                "agent.intelligence.attach_outreach_messages",
                lambda: self.intelligence.attach_outreach_messages(jobs, resume_text=resume_text),
                meta={"opportunity_jobs": sum(1 for item in jobs if item.opportunity_point)},
            )
            parse_details: list[dict[str, str]] = []
            job_parse_rule = 0
            job_parse_llm = 0
            for item in jobs:
                raw_parse_source = str(getattr(item, "parse_source", "") or "").strip().lower()[:20]
                parse_source = raw_parse_source if raw_parse_source in {"rule", "llm"} else "rule"
                if parse_source == "llm":
                    job_parse_llm += 1
                else:
                    job_parse_rule += 1
                parse_details.append({"post_id": item.post_id, "parse_source": parse_source})
                self.logger.info("岗位解析来源 | run=%s | post_id=%s | source=%s", run_id, item.post_id, parse_source)
            summaries = self._jobs_to_summary_records(run_id=run_id, jobs=jobs)
            llm_stage_calls = self.llm_enricher.stage_call_counts()
            llm_stage_fallbacks = self.llm_enricher.stage_fallback_counts()
            self._log_progress(
                run_id,
                80,
                (
                    f"岗位结构化完成，岗位 {len(jobs)} 条（LLM解析 {job_parse_llm}，规则解析 {job_parse_rule}），"
                    f"LLM提取调用 {int(llm_stage_calls.get('job', 0))} 次，"
                    f"规则回退 {int(llm_stage_fallbacks.get('job', 0))} 次"
                ),
            )

            orchestrator.run_stage(
                "storage.write_excel",
                lambda: self.store.write(notes, summaries, [], jobs=jobs),
                meta={
                    "excel_path": self.settings.storage.excel_path,
                    "upsert_notes": len(notes),
                    "new_notes": len(new_notes),
                    "updated_existing_notes": updated_existing_notes,
                },
            )
            orchestrator.run_stage(
                "storage.export_jobs_csv",
                lambda: self.store.export_jobs_csv(self.settings.storage.jobs_csv_path),
                meta={"jobs_csv_path": self.settings.storage.jobs_csv_path},
            )
            self._log_progress(
                run_id,
                88,
                f"本地存储已更新（新增 {len(new_notes)}，更新 {updated_existing_notes}，Excel/CSV）",
            )

            send_logs = []
            notify_note_ids: list[str] = []
            digest_due = False
            digest_sent = False
            digest_subject = ""
            digest_channels: list[str] = []
            digest_attachments: list[str] = []
            opportunity_post_ids = [item.post_id for item in jobs if item.opportunity_point]

            if notification_mode == "realtime":
                if jobs:
                    realtime_channels = list(self.settings.notification.realtime_channels)
                    try:
                        dispatch_result = orchestrator.run_stage(
                            "agent.communication.batch_dispatch_realtime",
                            lambda: self._dispatch_batch_with_compat(
                                run_id=run_id,
                                mode=mode,
                                jobs=jobs,
                                top_n=plan.top_n,
                                resume_text=resume_text,
                                channel_names=realtime_channels,
                                attachments=[],
                                digest_style=False,
                            ),
                            meta={
                                "mode": mode,
                                "channels": realtime_channels,
                                "job_count": len(jobs),
                            },
                        )
                    except Exception as exc:
                        dispatch_result = self._build_retry_dispatch_fallback(
                            run_id=run_id,
                            mode=mode,
                            jobs=jobs,
                            channels=realtime_channels,
                            attachments=[],
                            reason=f"dispatch_realtime_failed: {exc}",
                        )
                        self._enqueue_retry(
                            queue_type="email",
                            action="dispatch_digest",
                            run_id=run_id,
                            error=str(exc),
                            payload=dispatch_result,
                            dedupe_key=f"email:realtime:{run_id}",
                        )
                        self.logger.warning("实时通知失败已入队 | run=%s | error=%s", run_id, exc)
                    send_logs = dispatch_result["logs"]
                    notify_note_ids = sorted({log.note_id for log in send_logs})
                    digest_subject = dispatch_result["subject"]
                    self._enqueue_failed_email_dispatch(run_id=run_id, send_logs=send_logs, dispatch_result=dispatch_result)
                self._log_progress(run_id, 94, f"实时通知完成，发送日志 {len(send_logs)} 条")

            elif notification_mode == "digest":
                now = datetime.now(timezone.utc)
                digest_due = self._is_digest_due(now)
                enough_new = len(new_notes) >= max(0, int(self.settings.notification.digest_min_new_notes))
                allow_no_new = bool(self.settings.notification.digest_send_when_no_new)
                if digest_due and (enough_new or allow_no_new) and jobs:
                    digest_attachments = self._collect_digest_attachments()
                    digest_channels = list(self.settings.notification.digest_channels)
                    try:
                        dispatch_result = orchestrator.run_stage(
                            "agent.communication.batch_dispatch_digest",
                            lambda: self._dispatch_batch_with_compat(
                                run_id=run_id,
                                mode=mode,
                                jobs=jobs,
                                top_n=plan.top_n,
                                resume_text=resume_text,
                                channel_names=digest_channels,
                                attachments=digest_attachments,
                                digest_style=True,
                            ),
                            meta={
                                "channels": digest_channels,
                                "attachment_count": len(digest_attachments),
                                "new_notes": len(new_notes),
                                "target_notes": len(target_notes),
                            },
                        )
                    except Exception as exc:
                        dispatch_result = self._build_retry_dispatch_fallback(
                            run_id=run_id,
                            mode=mode,
                            jobs=jobs,
                            channels=digest_channels,
                            attachments=digest_attachments,
                            reason=f"dispatch_digest_failed: {exc}",
                        )
                        self._enqueue_retry(
                            queue_type="email",
                            action="dispatch_digest",
                            run_id=run_id,
                            error=str(exc),
                            payload=dispatch_result,
                            dedupe_key=f"email:digest:{run_id}",
                        )
                        self.logger.warning("摘要通知失败已入队 | run=%s | error=%s", run_id, exc)
                    send_logs = dispatch_result["logs"]
                    notify_note_ids = sorted({log.note_id for log in send_logs})
                    digest_subject = dispatch_result["subject"]
                    self._enqueue_failed_email_dispatch(run_id=run_id, send_logs=send_logs, dispatch_result=dispatch_result)
                    digest_sent = any(
                        str(getattr(log, "send_status", "")).strip().lower() == "success"
                        for log in send_logs
                    )
                    if digest_sent:
                        self._mark_digest_sent(now, run_id)
                if digest_due and (enough_new or allow_no_new):
                    self._log_progress(run_id, 94, f"摘要通知完成，发送日志 {len(send_logs)} 条")
                else:
                    self._log_progress(run_id, 94, "摘要通知按策略跳过（未到时间窗或新增不足）")
            else:
                self._log_progress(run_id, 94, "通知已关闭")

            if send_logs:
                orchestrator.run_stage(
                    "storage.append_send_logs",
                    lambda: self.store.write(raw=[], summaries=[], send_logs=send_logs, jobs=[]),
                    meta={"send_logs": len(send_logs)},
                )

            def _persist_state():
                for note in new_notes:
                    self.state.mark(note.note_id)
                self.state.save()

            orchestrator.run_stage("state.persist", _persist_state, meta={"marked_count": len(new_notes)})
            self._log_progress(run_id, 100, "本轮运行完成")

            stage_records = orchestrator.stage_records()
            stage_timing = self._summarize_stage_timing(stage_records)
            stage_error_codes = self._collect_stage_error_codes(stage_records)
            llm_error_codes = self.llm_client.error_counts()
            llm_stage_calls = self.llm_enricher.stage_call_counts()
            llm_stage_fallbacks = self.llm_enricher.stage_fallback_counts()
            llm_fallback_total = sum(int(v) for v in llm_stage_fallbacks.values())
            self._enqueue_llm_timeout_retries(run_id=run_id, llm_error_codes=llm_error_codes)
            llm_timeout_count = self._sum_timeout_error_counts(llm_error_codes)
            llm_timeout_rate = (
                float(llm_timeout_count) / float(self.llm_enricher.calls)
                if int(self.llm_enricher.calls) > 0
                else 0.0
            )
            detail_target_notes = int(detail_enrich_stats.get("target_notes", 0))
            detail_missing = int(detail_enrich_stats.get("detail_missing", 0))
            detail_missing_rate = (
                float(detail_missing) / float(detail_target_notes)
                if detail_target_notes > 0
                else 0.0
            )

            stats = {
                "runtime_name": self.settings.agent.runtime_name,
                "mode": mode,
                "notification_mode": notification_mode,
                "fetched": len(notes),
                "new_notes": len(new_notes),
                "updated_existing_notes": updated_existing_notes,
                "target_notes": len(target_notes),
                "filtered_out": len(filtered_out),
                "notify_note_count": len(notify_note_ids),
                "jobs": len(jobs),
                "opportunities": len(opportunity_post_ids),
                "summaries": len(jobs),
                "process_workers": process_workers,
                "send_logs": len(send_logs),
                "stages": len(stage_records),
                "stage_total_ms": stage_timing["total_ms"],
                "stage_avg_ms": stage_timing["avg_ms"],
                "stage_failed_count": stage_timing["failed_count"],
                "stage_slowest": stage_timing["slowest"],
                "stage_top_slow": stage_timing["top_slow"],
                "llm_enabled": self.llm_client.is_enabled(),
                "llm_available": self.llm_client.is_available(),
                "llm_calls": self.llm_enricher.calls,
                "llm_success": self.llm_enricher.success,
                "llm_fail": self.llm_enricher.fail,
                "llm_error_codes": llm_error_codes,
                "llm_timeout_count": int(llm_timeout_count),
                "llm_timeout_rate": llm_timeout_rate,
                "llm_stage_calls": llm_stage_calls,
                "llm_stage_fallbacks": llm_stage_fallbacks,
                "llm_fallback_total": llm_fallback_total,
                "stage_error_codes": stage_error_codes,
                "job_parse_rule": job_parse_rule,
                "job_parse_llm": job_parse_llm,
                "job_parse_details_count": len(parse_details),
                "digest_due": digest_due,
                "digest_sent": digest_sent,
                "resume_chars": len(resume_text),
                "fetch_fail_count_run": len(fetch_fail_events),
                "fetch_fail_streak": int(self._fetch_fail_streak),
                "xhs_data_empty": bool(xhs_data_empty),
                "detail_target_notes": detail_target_notes,
                "detail_attempted": int(detail_enrich_stats.get("attempted", 0)),
                "detail_success": int(detail_enrich_stats.get("success", 0)),
                "detail_failed": int(detail_enrich_stats.get("failed", 0)),
                "detail_skipped_no_token": int(detail_enrich_stats.get("skipped_no_token", 0)),
                "detail_filled": int(detail_enrich_stats.get("detail_filled", 0)),
                "detail_missing": detail_missing,
                "detail_missing_rate": detail_missing_rate,
                "detail_blocked": int(detail_enrich_stats.get("blocked", 0)),
                "detail_workers": int(getattr(self.settings.xhs, "detail_workers", 1)),
                "xhs_diagnosis": xhs_failure_diagnosis,
                "note_agent_total": int(note_agent_stats.get("total", 0)),
                "note_agent_llm_budgeted": int(note_agent_stats.get("llm_budgeted", 0)),
                "note_agent_rule_budgeted": int(note_agent_stats.get("rule_budgeted", 0)),
                "note_agent_worker_fallback": int(note_agent_stats.get("worker_fallback", 0)),
                "note_agent_errors": int(note_agent_stats.get("errors", 0)),
            }
            alert_eval = self._evaluate_threshold_alerts(run_id=run_id, stats=stats)
            stats["alerts_enabled"] = bool(alert_eval.get("enabled"))
            stats["alerts_triggered"] = list(alert_eval.get("triggered", []))
            stats["alerts_triggered_count"] = int(len(stats["alerts_triggered"]))
            stats["alerts_thresholds"] = dict(alert_eval.get("thresholds", {}))
            alert_dispatch = self._dispatch_threshold_alerts(
                run_id=run_id,
                mode=mode,
                stats=stats,
                triggered_alerts=list(stats["alerts_triggered"]),
            )
            stats["alerts_notified"] = list(alert_dispatch.get("notified_codes", []))
            stats["alerts_notified_count"] = int(len(stats["alerts_notified"]))
            stats["alerts_suppressed"] = list(alert_dispatch.get("suppressed_codes", []))
            stats["alerts_failed"] = list(alert_dispatch.get("failed_codes", []))
            stats["alerts_notification_channels"] = list(alert_dispatch.get("channels", []))
            stats["alerts_notification_logs"] = int(alert_dispatch.get("send_logs_count", 0))

            retry_snapshot = self.retry_queue.snapshot()
            stats["retry_pending"] = retry_snapshot.get("pending", {})
            stats["retry_running"] = retry_snapshot.get("running", {})
            stats["retry_enqueued"] = int(retry_snapshot.get("stats", {}).get("enqueued", 0))
            stats["retry_retried"] = int(retry_snapshot.get("stats", {}).get("retried", 0))
            stats["retry_succeeded"] = int(retry_snapshot.get("stats", {}).get("succeeded", 0))
            stats["retry_dropped"] = int(retry_snapshot.get("stats", {}).get("dropped", 0))
            self.journal.write(
                run_id,
                {
                    "stats": stats,
                    "keyword": keyword,
                    "mode": mode,
                    "notification_mode": notification_mode,
                    "new_note_ids": [note.note_id for note in new_notes],
                    "target_note_ids": [note.note_id for note in target_notes],
                    "notify_note_ids": notify_note_ids,
                    "opportunity_post_ids": opportunity_post_ids,
                    "filtered_out_samples": filtered_out[:50],
                    "fetch_fail_events": fetch_fail_events,
                    "xhs_diagnosis": xhs_failure_diagnosis,
                    "agent_plan": {
                        "detail_fetch_limit": plan.detail_fetch_limit,
                        "max_filter_items": plan.max_filter_items,
                        "max_job_items": plan.max_job_items,
                        "max_summary_items": plan.max_summary_items,
                        "top_n": plan.top_n,
                        "process_workers": process_workers,
                    },
                    "digest": {
                        "subject": digest_subject,
                        "channels": digest_channels,
                        "attachments": digest_attachments,
                    },
                    "job_parse": {
                        "rule": job_parse_rule,
                        "llm": job_parse_llm,
                        "details_count": len(parse_details),
                        "details": parse_details,
                    },
                    "note_agent": {
                        "stats": note_agent_stats,
                        "details_count": len(note_agent_details),
                        "details": note_agent_details[:200],
                    },
                    "alerts": {
                        "enabled": bool(stats.get("alerts_enabled")),
                        "triggered_count": int(stats.get("alerts_triggered_count", 0)),
                        "triggered": list(stats.get("alerts_triggered") or []),
                        "notified_count": int(stats.get("alerts_notified_count", 0)),
                        "notified_codes": list(stats.get("alerts_notified") or []),
                        "suppressed_codes": list(stats.get("alerts_suppressed") or []),
                        "failed_codes": list(stats.get("alerts_failed") or []),
                        "channels": list(stats.get("alerts_notification_channels") or []),
                    },
                    "retry": retry_snapshot,
                    "stage_records": stage_records,
                },
            )
            self.logger.info(
                "LLM统计 | run=%s | 调用=%s | 成功=%s | 失败=%s | 阶段调用=%s | 阶段回退=%s",
                run_id,
                self.llm_enricher.calls,
                self.llm_enricher.success,
                self.llm_enricher.fail,
                llm_stage_calls,
                llm_stage_fallbacks,
            )
            self.logger.info("运行结束 | run=%s | stats=%s", run_id, stats)
            return stats
        finally:
            self.lock.release()

    def _log_progress(self, run_id: str, percent: int, message: str) -> None:
        pct = max(0, min(100, int(percent)))
        width = 24
        done = max(0, min(width, int(round(pct * width / 100))))
        bar = f"{'#' * done}{'-' * (width - done)}"
        self.logger.info("运行进度 | run=%s | [%s] %3d%% | %s", run_id, bar, pct, message)

    def send_latest_stored(self, run_id: str, limit: int = 5) -> dict:
        mode = self._normalize_mode(self.settings.agent.mode or "auto")

        if not self.lock.acquire():
            self.logger.warning(
                "跳过发送 | run=%s | 原因=检测到上一次运行锁未释放 | lock=%s",
                run_id,
                self.lock.path,
            )
            return {"skipped": True, "reason": "run_locked"}

        try:
            top_n = max(1, int(limit))
            self._log_progress(run_id, 8, f"准备发送本地最新岗位，目标条数 {top_n}")
            summaries = list(self._load_latest_summaries_from_store(limit=top_n) or [])
            if summaries:
                jobs = self._summaries_to_jobs(summaries)
            else:
                jobs = self._load_latest_jobs_from_store(limit=top_n)
            self._log_progress(run_id, 32, f"已加载本地岗位 {len(jobs)} 条")
            resume_text = self.resume_loader.load_resume_text()
            digest_attachments = self._collect_digest_attachments()
            digest_channels = list(self.settings.notification.digest_channels)

            if not jobs:
                stats = {
                    "runtime_name": self.settings.agent.runtime_name,
                    "action": "send_latest_stored",
                    "requested_limit": top_n,
                    "loaded_jobs": 0,
                    "loaded_summaries": 0,
                    "send_logs": 0,
                    "digest_sent": False,
                }
                self.journal.write(
                    run_id,
                    {
                        "stats": stats,
                        "mode": mode,
                        "notification_mode": "digest",
                        "target_note_ids": [],
                        "notify_note_ids": [],
                        "digest": {
                            "subject": "",
                            "channels": digest_channels,
                            "attachments": digest_attachments,
                        },
                    },
                )
                self._log_progress(run_id, 100, "无可发送岗位，任务结束")
                self.logger.info("手动发送完成 | run=%s | stats=%s", run_id, stats)
                return stats

            try:
                dispatch_result = self._dispatch_batch_with_compat(
                    run_id=run_id,
                    mode=mode,
                    jobs=jobs,
                    top_n=top_n,
                    resume_text=resume_text,
                    channel_names=digest_channels,
                    attachments=digest_attachments,
                    digest_style=True,
                )
            except Exception as exc:
                dispatch_result = self._build_retry_dispatch_fallback(
                    run_id=run_id,
                    mode=mode,
                    jobs=jobs,
                    channels=digest_channels,
                    attachments=digest_attachments,
                    reason=f"send_latest_failed: {exc}",
                )
                self._enqueue_retry(
                    queue_type="email",
                    action="dispatch_digest",
                    run_id=run_id,
                    error=str(exc),
                    payload=dispatch_result,
                    dedupe_key=f"email:send-latest:{run_id}",
                )
                self.logger.warning("发送最新失败已入队 | run=%s | error=%s", run_id, exc)
            self._log_progress(run_id, 72, "通知分发执行完成，正在写入发送日志")
            send_logs = dispatch_result["logs"]
            self._enqueue_failed_email_dispatch(run_id=run_id, send_logs=send_logs, dispatch_result=dispatch_result)
            digest_sent = any(str(getattr(log, "send_status", "")).strip().lower() == "success" for log in send_logs)

            if send_logs:
                self.store.write(raw=[], summaries=[], send_logs=send_logs, jobs=[])

            if digest_sent:
                now = datetime.now(timezone.utc)
                self._mark_digest_sent(now, run_id)
                self.state.save()

            notify_note_ids = sorted({log.note_id for log in send_logs})
            stats = {
                "runtime_name": self.settings.agent.runtime_name,
                "action": "send_latest_stored",
                "requested_limit": top_n,
                "loaded_jobs": len(jobs),
                "loaded_summaries": len(summaries) if summaries else len(jobs),
                "send_logs": len(send_logs),
                "digest_sent": digest_sent,
            }
            self.journal.write(
                run_id,
                {
                    "stats": stats,
                    "mode": mode,
                    "notification_mode": "digest",
                    "target_note_ids": [item.post_id for item in jobs],
                    "notify_note_ids": notify_note_ids,
                    "digest": {
                        "subject": dispatch_result["subject"],
                        "channels": digest_channels,
                        "attachments": digest_attachments,
                    },
                },
            )
            self._log_progress(run_id, 100, f"发送任务完成，发送日志 {len(send_logs)} 条")
            self.logger.info("手动发送完成 | run=%s | stats=%s", run_id, stats)
            return stats
        finally:
            self.lock.release()

    @staticmethod
    def _normalize_mode(mode: str) -> str:
        value = (mode or "auto").strip().lower()
        if value == "smart":
            value = "agent"
        if value not in {"auto", "agent"}:
            return "auto"
        return value

    @staticmethod
    def _normalize_notification_mode(mode: str) -> str:
        value = (mode or "digest").strip().lower()
        if value not in {"digest", "realtime", "off"}:
            return "digest"
        return value

    def _collect_digest_attachments(self) -> list[str]:
        paths: list[str] = []
        if self.settings.notification.attach_excel:
            excel = Path(self.settings.storage.excel_path)
            if excel.exists():
                paths.append(str(excel))
        if self.settings.notification.attach_jobs_csv:
            jobs_csv = Path(self.settings.storage.jobs_csv_path)
            if jobs_csv.exists():
                paths.append(str(jobs_csv))
        return paths

    def _is_digest_due(self, now: datetime) -> bool:
        interval = int(self.settings.notification.digest_interval_minutes)
        if hasattr(self.state, "is_digest_due"):
            return bool(self.state.is_digest_due(now, interval))
        return True

    def _mark_digest_sent(self, now: datetime, run_id: str) -> None:
        if hasattr(self.state, "mark_digest_sent"):
            self.state.mark_digest_sent(now, run_id)

    @staticmethod
    def _sum_timeout_error_counts(llm_error_codes: dict[str, int] | None) -> int:
        timeout_keys = {"connect_timeout", "read_timeout", "timeout"}
        total = 0
        if not isinstance(llm_error_codes, dict):
            return 0
        for code, count in llm_error_codes.items():
            key = str(code or "").strip().lower()
            if key not in timeout_keys:
                continue
            total += max(0, int(count or 0))
        return total

    @staticmethod
    def _normalize_rate_threshold(value: Any, *, default: float) -> float:
        try:
            parsed = float(value)
        except Exception:
            return float(default)
        if parsed > 1.0 and parsed <= 100.0:
            parsed = parsed / 100.0
        if parsed < 0:
            return 0.0
        if parsed > 1.0:
            return 1.0
        return parsed

    @staticmethod
    def _normalize_numeric_threshold(value: Any, *, default: float, minimum: float = 0.0) -> float:
        try:
            parsed = float(value)
        except Exception:
            parsed = float(default)
        if parsed < float(minimum):
            parsed = float(minimum)
        return parsed

    @staticmethod
    def _normalize_window_runs(value: Any, *, default: int) -> int:
        try:
            parsed = int(value)
        except Exception:
            parsed = int(default)
        return max(1, parsed)

    @staticmethod
    def _normalize_min_samples(value: Any, *, default: int) -> int:
        try:
            parsed = int(value)
        except Exception:
            parsed = int(default)
        return max(1, parsed)

    def _load_recent_alert_stats(self, *, limit: int, exclude_run_id: str) -> list[dict[str, Any]]:
        max_items = max(0, int(limit))
        if max_items <= 0:
            return []
        runs_dir = getattr(self.journal, "base_dir", None)
        if runs_dir is None:
            return []
        if not isinstance(runs_dir, Path):
            runs_dir = Path(str(runs_dir))
        if not runs_dir.exists():
            return []

        history: list[dict[str, Any]] = []
        files = sorted(runs_dir.glob("*.json"), key=lambda item: item.name, reverse=True)
        for file in files:
            if str(file.stem) == str(exclude_run_id):
                continue
            try:
                payload = json.loads(file.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            stats = payload.get("stats")
            if not isinstance(stats, dict):
                continue
            history.append(stats)
            if len(history) >= max_items:
                break
        return history

    @staticmethod
    def _aggregate_window_fetch(window_stats: list[dict[str, Any]]) -> dict[str, Any]:
        values: list[float] = []
        for item in window_stats:
            try:
                values.append(max(0.0, float(item.get("fetch_fail_streak", 0) or 0)))
            except Exception:
                values.append(0.0)
        if not values:
            return {"runs": 0, "max": 0.0, "avg": 0.0}
        return {
            "runs": len(values),
            "max": max(values),
            "avg": float(sum(values)) / float(len(values)),
        }

    @staticmethod
    def _aggregate_window_rate(
        window_stats: list[dict[str, Any]],
        *,
        numerator_key: str,
        denominator_key: str,
    ) -> dict[str, Any]:
        numerator = 0
        denominator = 0
        for item in window_stats:
            numerator += max(0, int(item.get(numerator_key, 0) or 0))
            denominator += max(0, int(item.get(denominator_key, 0) or 0))
        rate = float(numerator) / float(denominator) if denominator > 0 else 0.0
        return {
            "runs": len(window_stats),
            "numerator": numerator,
            "denominator": denominator,
            "rate": rate,
        }

    def _evaluate_threshold_alerts(self, *, run_id: str, stats: dict[str, Any]) -> dict[str, Any]:
        cfg = getattr(getattr(self.settings, "observability", None), "alerts", None)
        if cfg is None or not bool(getattr(cfg, "enabled", True)):
            return {"enabled": False, "thresholds": {}, "triggered": []}

        # Legacy single-threshold values are kept as fallback defaults.
        legacy_fetch_threshold = max(1, int(getattr(cfg, "fetch_fail_streak_threshold", 2) or 2))
        legacy_llm_rate_threshold = self._normalize_rate_threshold(
            getattr(cfg, "llm_timeout_rate_threshold", 0.35),
            default=0.35,
        )
        legacy_llm_min_calls = max(1, int(getattr(cfg, "llm_timeout_min_calls", 6) or 6))
        legacy_detail_rate_threshold = self._normalize_rate_threshold(
            getattr(cfg, "detail_missing_rate_threshold", 0.45),
            default=0.45,
        )
        legacy_detail_min_samples = max(1, int(getattr(cfg, "detail_missing_min_samples", 6) or 6))

        fetch_cfg = getattr(cfg, "fetch_fail_streak", {}) or {}
        llm_cfg = getattr(cfg, "llm_timeout_rate", {}) or {}
        detail_cfg = getattr(cfg, "detail_missing_rate", {}) or {}
        if not isinstance(fetch_cfg, dict):
            fetch_cfg = {}
        if not isinstance(llm_cfg, dict):
            llm_cfg = {}
        if not isinstance(detail_cfg, dict):
            detail_cfg = {}

        fetch_short_window_runs = self._normalize_window_runs(fetch_cfg.get("short_window_runs"), default=1)
        fetch_short_threshold = self._normalize_numeric_threshold(
            fetch_cfg.get("short_threshold"),
            default=float(legacy_fetch_threshold),
            minimum=1.0,
        )
        fetch_short_min_runs = self._normalize_min_samples(fetch_cfg.get("short_min_runs"), default=1)
        fetch_long_window_runs = self._normalize_window_runs(fetch_cfg.get("long_window_runs"), default=6)
        fetch_long_threshold = self._normalize_numeric_threshold(
            fetch_cfg.get("long_threshold"),
            default=max(1.0, float(legacy_fetch_threshold) * 0.6),
            minimum=1.0,
        )
        fetch_long_min_runs = self._normalize_min_samples(fetch_cfg.get("long_min_runs"), default=3)

        llm_short_window_runs = self._normalize_window_runs(llm_cfg.get("short_window_runs"), default=1)
        llm_short_threshold = self._normalize_rate_threshold(
            llm_cfg.get("short_threshold"),
            default=legacy_llm_rate_threshold,
        )
        llm_short_min_calls = self._normalize_min_samples(
            llm_cfg.get("short_min_samples"),
            default=legacy_llm_min_calls,
        )
        llm_long_window_runs = self._normalize_window_runs(llm_cfg.get("long_window_runs"), default=8)
        llm_long_threshold = self._normalize_rate_threshold(
            llm_cfg.get("long_threshold"),
            default=min(1.0, max(0.0, legacy_llm_rate_threshold * 0.7)),
        )
        llm_long_min_calls = self._normalize_min_samples(
            llm_cfg.get("long_min_samples"),
            default=max(12, legacy_llm_min_calls * 3),
        )

        detail_short_window_runs = self._normalize_window_runs(detail_cfg.get("short_window_runs"), default=1)
        detail_short_threshold = self._normalize_rate_threshold(
            detail_cfg.get("short_threshold"),
            default=legacy_detail_rate_threshold,
        )
        detail_short_min_samples = self._normalize_min_samples(
            detail_cfg.get("short_min_samples"),
            default=legacy_detail_min_samples,
        )
        detail_long_window_runs = self._normalize_window_runs(detail_cfg.get("long_window_runs"), default=8)
        detail_long_threshold = self._normalize_rate_threshold(
            detail_cfg.get("long_threshold"),
            default=min(1.0, max(0.0, legacy_detail_rate_threshold * 0.7)),
        )
        detail_long_min_samples = self._normalize_min_samples(
            detail_cfg.get("long_min_samples"),
            default=max(12, legacy_detail_min_samples * 3),
        )

        max_window_runs = max(
            fetch_short_window_runs,
            fetch_long_window_runs,
            llm_short_window_runs,
            llm_long_window_runs,
            detail_short_window_runs,
            detail_long_window_runs,
        )
        historical_stats = self._load_recent_alert_stats(
            limit=max(0, max_window_runs - 1),
            exclude_run_id=run_id,
        )
        run_stats_window = [stats] + historical_stats

        thresholds = {
            "fetch_fail_streak": {
                "short_window_runs": fetch_short_window_runs,
                "short_threshold": fetch_short_threshold,
                "short_min_runs": fetch_short_min_runs,
                "long_window_runs": fetch_long_window_runs,
                "long_threshold": fetch_long_threshold,
                "long_min_runs": fetch_long_min_runs,
            },
            "llm_timeout_rate": {
                "short_window_runs": llm_short_window_runs,
                "short_threshold": llm_short_threshold,
                "short_min_samples": llm_short_min_calls,
                "long_window_runs": llm_long_window_runs,
                "long_threshold": llm_long_threshold,
                "long_min_samples": llm_long_min_calls,
            },
            "detail_missing_rate": {
                "short_window_runs": detail_short_window_runs,
                "short_threshold": detail_short_threshold,
                "short_min_samples": detail_short_min_samples,
                "long_window_runs": detail_long_window_runs,
                "long_threshold": detail_long_threshold,
                "long_min_samples": detail_long_min_samples,
            },
            "history_sample_runs": len(run_stats_window),
        }
        triggered: list[dict[str, Any]] = []

        fetch_short = self._aggregate_window_fetch(run_stats_window[:fetch_short_window_runs])
        fetch_long = self._aggregate_window_fetch(run_stats_window[:fetch_long_window_runs])
        fetch_ready = (
            int(fetch_short.get("runs", 0)) >= fetch_short_min_runs
            and int(fetch_long.get("runs", 0)) >= fetch_long_min_runs
        )
        fetch_short_value = float(fetch_short.get("max", 0.0))
        fetch_long_value = float(fetch_long.get("avg", 0.0))
        if fetch_ready and fetch_short_value >= fetch_short_threshold and fetch_long_value >= fetch_long_threshold:
            triggered.append(
                {
                    "code": "fetch_fail_streak",
                    "level": "critical",
                    "metric": "fetch_fail_streak",
                    "value": fetch_short_value,
                    "threshold": fetch_short_threshold,
                    "sample_size": int(fetch_short.get("runs", 0)),
                    "window_short": {
                        "runs": int(fetch_short.get("runs", 0)),
                        "value": fetch_short_value,
                        "threshold": fetch_short_threshold,
                        "mode": "max",
                    },
                    "window_long": {
                        "runs": int(fetch_long.get("runs", 0)),
                        "value": fetch_long_value,
                        "threshold": fetch_long_threshold,
                        "mode": "avg",
                    },
                    "reason": (
                        f"fetch_fail_streak 短窗max={fetch_short_value:.2f}/{fetch_short_threshold:.2f}，"
                        f"长窗avg={fetch_long_value:.2f}/{fetch_long_threshold:.2f}"
                    ),
                }
            )

        llm_short = self._aggregate_window_rate(
            run_stats_window[:llm_short_window_runs],
            numerator_key="llm_timeout_count",
            denominator_key="llm_calls",
        )
        llm_long = self._aggregate_window_rate(
            run_stats_window[:llm_long_window_runs],
            numerator_key="llm_timeout_count",
            denominator_key="llm_calls",
        )
        llm_short_calls = int(llm_short.get("denominator", 0))
        llm_short_timeout_count = int(llm_short.get("numerator", 0))
        llm_short_rate = float(llm_short.get("rate", 0.0))
        llm_long_calls = int(llm_long.get("denominator", 0))
        llm_long_timeout_count = int(llm_long.get("numerator", 0))
        llm_long_rate = float(llm_long.get("rate", 0.0))
        if (
            llm_short_calls >= llm_short_min_calls
            and llm_long_calls >= llm_long_min_calls
            and llm_short_rate >= llm_short_threshold
            and llm_long_rate >= llm_long_threshold
        ):
            triggered.append(
                {
                    "code": "llm_timeout_rate",
                    "level": "warning",
                    "metric": "llm_timeout_rate",
                    "value": llm_short_rate,
                    "threshold": llm_short_threshold,
                    "sample_size": llm_short_calls,
                    "window_short": {
                        "runs": int(llm_short.get("runs", 0)),
                        "value": llm_short_rate,
                        "threshold": llm_short_threshold,
                        "numerator": llm_short_timeout_count,
                        "denominator": llm_short_calls,
                    },
                    "window_long": {
                        "runs": int(llm_long.get("runs", 0)),
                        "value": llm_long_rate,
                        "threshold": llm_long_threshold,
                        "numerator": llm_long_timeout_count,
                        "denominator": llm_long_calls,
                    },
                    "reason": (
                        f"LLM 超时率短窗 {llm_short_rate:.1%}（{llm_short_timeout_count}/{llm_short_calls}）/"
                        f"{llm_short_threshold:.1%}，长窗 {llm_long_rate:.1%}（{llm_long_timeout_count}/{llm_long_calls}）/"
                        f"{llm_long_threshold:.1%}"
                    ),
                }
            )

        detail_short = self._aggregate_window_rate(
            run_stats_window[:detail_short_window_runs],
            numerator_key="detail_missing",
            denominator_key="detail_target_notes",
        )
        detail_long = self._aggregate_window_rate(
            run_stats_window[:detail_long_window_runs],
            numerator_key="detail_missing",
            denominator_key="detail_target_notes",
        )
        detail_short_target = int(detail_short.get("denominator", 0))
        detail_short_missing = int(detail_short.get("numerator", 0))
        detail_short_rate = float(detail_short.get("rate", 0.0))
        detail_long_target = int(detail_long.get("denominator", 0))
        detail_long_missing = int(detail_long.get("numerator", 0))
        detail_long_rate = float(detail_long.get("rate", 0.0))
        if (
            detail_short_target >= detail_short_min_samples
            and detail_long_target >= detail_long_min_samples
            and detail_short_rate >= detail_short_threshold
            and detail_long_rate >= detail_long_threshold
        ):
            triggered.append(
                {
                    "code": "detail_missing_rate",
                    "level": "warning",
                    "metric": "detail_missing_rate",
                    "value": detail_short_rate,
                    "threshold": detail_short_threshold,
                    "sample_size": detail_short_target,
                    "window_short": {
                        "runs": int(detail_short.get("runs", 0)),
                        "value": detail_short_rate,
                        "threshold": detail_short_threshold,
                        "numerator": detail_short_missing,
                        "denominator": detail_short_target,
                    },
                    "window_long": {
                        "runs": int(detail_long.get("runs", 0)),
                        "value": detail_long_rate,
                        "threshold": detail_long_threshold,
                        "numerator": detail_long_missing,
                        "denominator": detail_long_target,
                    },
                    "reason": (
                        f"详情缺失率短窗 {detail_short_rate:.1%}（{detail_short_missing}/{detail_short_target}）/"
                        f"{detail_short_threshold:.1%}，长窗 {detail_long_rate:.1%}（{detail_long_missing}/{detail_long_target}）/"
                        f"{detail_long_threshold:.1%}"
                    ),
                }
            )

        if triggered:
            codes = ",".join(str(item.get("code") or "") for item in triggered)
            self.logger.warning("阈值告警命中 | run=%s | count=%s | codes=%s", run_id, len(triggered), codes)
        return {"enabled": True, "thresholds": thresholds, "triggered": triggered}

    def _resolve_alert_channels(self) -> list[str]:
        cfg = getattr(getattr(self.settings, "observability", None), "alerts", None)
        configured = list(getattr(cfg, "channels", []) or [])
        channels = [str(item).strip() for item in configured if str(item).strip()]
        if channels:
            return channels
        fallback = [str(item).strip() for item in (self.settings.notification.digest_channels or []) if str(item).strip()]
        return fallback or ["email"]

    def _is_alert_due(self, alert_code: str, now: datetime, cooldown_minutes: int) -> bool:
        if hasattr(self.state, "is_alert_due"):
            return bool(self.state.is_alert_due(alert_code, now, cooldown_minutes))
        return True

    def _mark_alert_sent(self, alert_code: str, now: datetime) -> None:
        if hasattr(self.state, "mark_alert_sent"):
            self.state.mark_alert_sent(alert_code, now)

    def _build_threshold_alert_message(
        self,
        *,
        run_id: str,
        mode: str,
        stats: dict[str, Any],
        alerts: list[dict[str, Any]],
    ) -> str:
        lines = [
            "SuccessionPilot 阈值告警",
            "====================",
            "",
            f"Run ID: {run_id}",
            f"Mode: {mode}",
            f"Keyword: {self.settings.xhs.keyword}",
            f"Triggered: {len(alerts)}",
            "",
            "告警详情：",
        ]
        for index, item in enumerate(alerts, start=1):
            code = str(item.get("code") or "unknown")
            reason = str(item.get("reason") or "").strip()
            value = item.get("value")
            threshold = item.get("threshold")
            sample = int(item.get("sample_size") or 0)
            if isinstance(value, float):
                value_text = f"{value:.1%}" if "rate" in code else f"{value:.4f}"
            else:
                value_text = str(value)
            if isinstance(threshold, float):
                threshold_text = f"{threshold:.1%}" if "rate" in code else f"{threshold:.4f}"
            else:
                threshold_text = str(threshold)
            short_window = item.get("window_short") if isinstance(item.get("window_short"), dict) else {}
            long_window = item.get("window_long") if isinstance(item.get("window_long"), dict) else {}
            lines.extend(
                [
                    f"{index}. {code}",
                    f"   - value: {value_text}",
                    f"   - threshold: {threshold_text}",
                    f"   - sample_size: {sample}",
                    (
                        f"   - short_window: runs={int(short_window.get('runs', 0))} "
                        f"value={short_window.get('value')} threshold={short_window.get('threshold')}"
                    )
                    if short_window
                    else "   - short_window: -",
                    (
                        f"   - long_window: runs={int(long_window.get('runs', 0))} "
                        f"value={long_window.get('value')} threshold={long_window.get('threshold')}"
                    )
                    if long_window
                    else "   - long_window: -",
                    f"   - reason: {reason or '-'}",
                ]
            )
        diagnosis = stats.get("xhs_diagnosis")
        if isinstance(diagnosis, dict) and diagnosis:
            lines.extend(
                [
                    "",
                    "XHS 诊断：",
                    f"- category: {diagnosis.get('failure_category') or '-'}",
                    f"- reason: {diagnosis.get('reason') or '-'}",
                    f"- mcp_connect: {diagnosis.get('mcp_connect')}",
                    f"- login_status: {diagnosis.get('login_status')}",
                    f"- cookie_file_ready: {diagnosis.get('cookie_file_ready')}",
                ]
            )
        lines.extend(
            [
                "",
                "关键指标：",
                f"- fetch_fail_streak: {int(stats.get('fetch_fail_streak', 0))}",
                (
                    f"- llm_timeout_rate: {float(stats.get('llm_timeout_rate', 0.0) or 0.0):.1%}"
                    f" ({int(stats.get('llm_timeout_count', 0))}/{int(stats.get('llm_calls', 0))})"
                ),
                (
                    f"- detail_missing_rate: {float(stats.get('detail_missing_rate', 0.0) or 0.0):.1%}"
                    f" ({int(stats.get('detail_missing', 0))}/{int(stats.get('detail_target_notes', 0))})"
                ),
                "",
                "可在控制中心查看 run 详情与重试队列定位问题。",
            ]
        )
        return "\n".join(lines).strip()

    def _dispatch_threshold_alerts(
        self,
        *,
        run_id: str,
        mode: str,
        stats: dict[str, Any],
        triggered_alerts: list[dict[str, Any]],
    ) -> dict[str, Any]:
        if not triggered_alerts:
            return {
                "notified_codes": [],
                "suppressed_codes": [],
                "failed_codes": [],
                "channels": [],
                "send_logs_count": 0,
            }
        cfg = getattr(getattr(self.settings, "observability", None), "alerts", None)
        cooldown_minutes = max(1, int(getattr(cfg, "cooldown_minutes", 60) or 60))
        now = datetime.now(timezone.utc)
        due_alerts: list[dict[str, Any]] = []
        suppressed_codes: list[str] = []
        for item in triggered_alerts:
            code = str(item.get("code") or "").strip().lower()
            if not code:
                continue
            if self._is_alert_due(code, now, cooldown_minutes):
                due_alerts.append(item)
            else:
                suppressed_codes.append(code)

        if not due_alerts:
            if suppressed_codes:
                self.logger.info(
                    "阈值告警命中但处于冷却期 | run=%s | cooldown=%smin | codes=%s",
                    run_id,
                    cooldown_minutes,
                    ",".join(sorted(set(suppressed_codes))),
                )
            return {
                "notified_codes": [],
                "suppressed_codes": sorted(set(suppressed_codes)),
                "failed_codes": [],
                "channels": self._resolve_alert_channels(),
                "send_logs_count": 0,
            }

        channels = self._resolve_alert_channels()
        subject = f"SuccessionPilot | 阈值告警 | {run_id} | {len(due_alerts)}项"
        text = self._build_threshold_alert_message(run_id=run_id, mode=mode, stats=stats, alerts=due_alerts)
        logs = self.router.dispatch_digest(
            run_id=f"alert:{run_id}",
            subject=subject,
            text=text,
            attachments=[],
            channel_names=channels,
        )
        dispatched_ok = any(str(getattr(log, "send_status", "")).strip().lower() == "success" for log in logs)
        due_codes = sorted({str(item.get("code") or "").strip().lower() for item in due_alerts if str(item.get("code") or "").strip()})
        failed_codes: list[str] = []
        notified_codes: list[str] = []
        if dispatched_ok:
            for code in due_codes:
                self._mark_alert_sent(code, now)
            try:
                self.state.save()
            except Exception as exc:
                self.logger.warning("告警冷却状态保存失败 | run=%s | error=%s", run_id, exc)
            notified_codes = due_codes
            self.logger.warning(
                "阈值告警已发送 | run=%s | channels=%s | codes=%s",
                run_id,
                ",".join(channels),
                ",".join(notified_codes),
            )
        else:
            failed_codes = due_codes
            self.logger.warning(
                "阈值告警发送失败 | run=%s | channels=%s | codes=%s",
                run_id,
                ",".join(channels),
                ",".join(failed_codes),
            )

        dispatch_result = {
            "subject": subject,
            "body": text,
            "channels": channels,
            "attachments": [],
        }
        self._enqueue_failed_email_dispatch(run_id=run_id, send_logs=logs, dispatch_result=dispatch_result)
        return {
            "notified_codes": notified_codes,
            "suppressed_codes": sorted(set(suppressed_codes)),
            "failed_codes": failed_codes,
            "channels": channels,
            "send_logs_count": len(logs),
        }

    def _load_latest_jobs_from_store(self, limit: int) -> list[JobRecord]:
        excel_path = Path(self.settings.storage.excel_path)
        if not excel_path.exists():
            return []

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            if "jobs" not in wb.sheetnames:
                return []
            ws = wb["jobs"]
            rows = self._read_sheet_rows(ws)
        finally:
            wb.close()

        rows.sort(
            key=lambda item: (
                self._to_datetime(item.get("publish_time")),
                str(item.get("PostID") or ""),
            ),
            reverse=True,
        )

        jobs: list[JobRecord] = []
        for row in rows[: max(1, int(limit))]:
            post_id = str(row.get("PostID") or "").strip()
            if not post_id:
                continue
            jobs.append(
                JobRecord(
                    run_id=str(row.get("run_id") or f"from-store:{post_id[:12]}"),
                    post_id=post_id,
                    company=str(row.get("Company") or ""),
                    position=str(row.get("Position") or ""),
                    location=str(row.get("Location") or ""),
                    requirements=str(row.get("Requirements") or ""),
                    arrival_time=str(row.get("arrival_time") or ""),
                    application_method=str(row.get("application_method") or ""),
                    author=str(row.get("author") or ""),
                    risk_line=str(row.get("risk_line") or "low"),
                    match_score=float(row.get("match_score") or 0.0),
                    match_reason=str(row.get("match_reason") or ""),
                    link=str(row.get("Link") or ""),
                    mode=str(row.get("mode") or "auto"),
                    publish_time=self._to_datetime(row.get("publish_time")),
                    source_title=str(row.get("source_title") or ""),
                    comment_count=int(row.get("comment_count") or 0),
                    comments_preview=str(row.get("comments_preview") or ""),
                    original_text=str(row.get("original_text") or ""),
                    opportunity_point=bool(row.get("opportunity_point")),
                    outreach_message=str(row.get("outreach_message") or ""),
                )
            )
        return jobs

    def _dispatch_batch_with_compat(
        self,
        *,
        run_id: str,
        mode: str,
        jobs: list[JobRecord],
        top_n: int,
        resume_text: str,
        channel_names: list[str],
        attachments: list[str],
        digest_style: bool,
    ) -> dict:
        if hasattr(self.communication, "dispatch_batch"):
            result = self.communication.dispatch_batch(
                run_id=run_id,
                mode=mode,
                jobs=jobs,
                resume_text=resume_text,
                channel_names=channel_names,
                attachments=attachments,
            )
            return {
                "logs": result.logs,
                "subject": result.subject,
                "body": result.body,
                "attachments": list(attachments),
                "channels": list(channel_names),
            }

        realtime_jobs = jobs
        if not digest_style and mode == "agent":
            realtime_jobs = sorted(jobs, key=lambda item: float(item.match_score), reverse=True)[: max(1, int(top_n))]
        summaries = self._jobs_to_summary_records(run_id=run_id, jobs=realtime_jobs)
        if digest_style:
            result = self.communication.dispatch_digest(
                run_id=run_id,
                mode=mode,
                new_notes=[],
                target_notes=[],
                summaries=summaries,
                attachments=attachments,
                channel_names=channel_names,
            )
            return {
                "logs": result.logs,
                "subject": result.subject,
                "body": result.body,
                "attachments": list(attachments),
                "channels": list(channel_names),
            }

        logs = self.communication.dispatch_realtime(
            run_id=run_id,
            summaries=summaries,
            channel_names=channel_names,
        )
        body = "\n\n".join(str(item.summary or "").strip() for item in summaries if str(item.summary or "").strip())
        return {
            "logs": logs,
            "subject": f"SuccessionPilot | realtime | {run_id}",
            "body": body,
            "attachments": list(attachments),
            "channels": list(channel_names),
        }

    def _retry_worker_loop(self) -> None:
        self.logger.info("重试队列后台线程已启动 | interval=%ss | batch=%s", self._retry_worker_interval, self._retry_batch_size)
        while not self._retry_stop.wait(self._retry_worker_interval):
            try:
                self._process_retry_queue_once(limit=self._retry_batch_size)
            except Exception as exc:
                self.logger.warning("重试队列处理异常：%s", exc)

    def _process_retry_queue_once(self, *, limit: int) -> None:
        if not self._retry_enabled:
            return
        if self.lock.path.exists():
            return
        items = self.retry_queue.pop_due(limit=max(1, int(limit)))
        if not items:
            return
        for item in items:
            item_id = str(item.get("id") or "")
            queue_type = str(item.get("queue_type") or "").strip().lower()
            action = str(item.get("action") or "").strip().lower()
            attempt = max(0, int(item.get("attempt") or 0)) + 1
            trace_id = f"rq-{item_id[:8]}-a{attempt}" if item_id else f"rq-unknown-a{attempt}"
            started = time.perf_counter()
            self.logger.info(
                "重试执行开始 | trace=%s | queue=%s | action=%s | id=%s | attempt=%s/%s",
                trace_id,
                queue_type,
                action,
                item_id,
                attempt,
                int(item.get("max_attempts") or 0),
            )
            try:
                self._handle_retry_item(item)
                duration_ms = int((time.perf_counter() - started) * 1000)
                self.retry_queue.mark_success(item_id, result="ok", duration_ms=duration_ms, trace_id=trace_id)
                self.logger.info(
                    "重试成功 | trace=%s | queue=%s | action=%s | id=%s | duration_ms=%s",
                    trace_id,
                    queue_type,
                    action,
                    item_id,
                    duration_ms,
                )
            except Exception as exc:
                duration_ms = int((time.perf_counter() - started) * 1000)
                error_code = self._classify_retry_error_code(exc)
                self.retry_queue.mark_retry(
                    item_id,
                    error=str(exc),
                    duration_ms=duration_ms,
                    trace_id=trace_id,
                    error_code=error_code,
                )
                self.logger.warning(
                    "重试失败，已回退重排 | trace=%s | queue=%s | action=%s | id=%s | code=%s | duration_ms=%s | error=%s",
                    trace_id,
                    queue_type,
                    action,
                    item_id,
                    error_code,
                    duration_ms,
                    exc,
                )

    def _handle_retry_item(self, item: dict[str, Any]) -> None:
        queue_type = str(item.get("queue_type") or "").strip().lower()
        if queue_type == "fetch":
            self._handle_retry_fetch(item)
            return
        if queue_type == "llm_timeout":
            self._handle_retry_llm_timeout(item)
            return
        if queue_type == "email":
            self._handle_retry_email(item)
            return
        raise ValueError(f"unsupported queue type: {queue_type}")

    def _handle_retry_fetch(self, item: dict[str, Any]) -> None:
        payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
        action = str(item.get("action") or "").strip().lower()
        if action == "ensure_logged_in":
            self.collector.ensure_logged_in()
            return
        if action == "search_notes":
            keyword = str(payload.get("keyword") or self.settings.xhs.keyword).strip() or self.settings.xhs.keyword
            max_results = max(1, int(payload.get("max_results") or self.settings.xhs.max_results))
            self.collector.search_notes(
                run_id=f"retry-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}",
                keyword=keyword,
                max_results=max_results,
            )
            return
        if action == "enrich_note_detail":
            note_id = str(payload.get("note_id") or "").strip()
            xsec_token = str(payload.get("xsec_token") or "").strip()
            if not note_id or not xsec_token:
                raise ValueError("missing note_id/xsec_token")
            note = NoteRecord(
                run_id=f"retry-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}",
                keyword=self.settings.xhs.keyword,
                note_id=note_id,
                title="",
                author="",
                publish_time=datetime.now(timezone.utc),
                publish_time_text="",
                like_count=0,
                comment_count=0,
                share_count=0,
                url=f"https://www.xiaohongshu.com/explore/{note_id}",
                raw_json="{}",
                xsec_token=xsec_token,
            )
            self.collector.enrich_note_details([note], max_notes=1)
            return
        raise ValueError(f"unsupported fetch action: {action}")

    def _handle_retry_llm_timeout(self, item: dict[str, Any]) -> None:
        payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
        error_code = str(payload.get("error_code") or "timeout").strip().lower()
        scope = str(payload.get("scope") or "retry_llm_timeout").strip() or "retry_llm_timeout"
        result = self.llm_client.chat_text(
            system_prompt="You are a health checker. Reply with: ok",
            user_prompt=f"llm timeout replay probe, code={error_code}",
            temperature=0.0,
            max_tokens=8,
            scope=scope,
        )
        if not result:
            raise RuntimeError(f"llm unavailable: {self.llm_client.last_error_code(scope=scope)}")

    def _handle_retry_email(self, item: dict[str, Any]) -> None:
        payload = item.get("payload") if isinstance(item.get("payload"), dict) else {}
        action = str(item.get("action") or "").strip().lower()
        if action != "dispatch_digest":
            raise ValueError(f"unsupported email action: {action}")
        subject = str(payload.get("subject") or "").strip() or "SuccessionPilot | retry"
        text = str(payload.get("body") or payload.get("text") or "").strip()
        if not text:
            raise ValueError("missing email body")
        channels = payload.get("channels") if isinstance(payload.get("channels"), list) else ["email"]
        attachments = payload.get("attachments") if isinstance(payload.get("attachments"), list) else []
        idempotency_key = str(payload.get("idempotency_key") or item.get("idempotency_key") or "").strip()
        if not idempotency_key:
            idempotency_key = self._build_email_idempotency_key(subject=subject, body=text, channels=channels)

        if self.retry_queue.has_completed_idempotency(queue_type="email", idempotency_key=idempotency_key):
            self.logger.info("重试邮件幂等跳过 | key=%s | action=%s", idempotency_key[:48], action)
            return

        logs = self.router.dispatch_digest(
            run_id=f"retry:{item.get('id')}",
            subject=subject,
            text=text,
            attachments=[str(x) for x in attachments if str(x).strip()],
            channel_names=[str(x) for x in channels if str(x).strip()],
        )
        failed = [
            log
            for log in logs
            if str(getattr(log, "channel", "")).strip().lower() == "email"
            and str(getattr(log, "send_status", "")).strip().lower() != "success"
        ]
        if failed:
            errors = "; ".join(str(getattr(log, "send_response", "") or "") for log in failed)
            raise RuntimeError(errors or "email retry failed")

    def _enqueue_retry(
        self,
        *,
        queue_type: str,
        action: str,
        run_id: str,
        error: str,
        payload: dict[str, Any] | None = None,
        dedupe_key: str = "",
        idempotency_key: str = "",
    ) -> None:
        if not self._retry_enabled:
            return
        try:
            item = self.retry_queue.enqueue(
                queue_type=queue_type,
                action=action,
                payload=payload or {},
                run_id=run_id,
                error=error,
                dedupe_key=dedupe_key,
                idempotency_key=idempotency_key,
            )
            self.logger.info(
                "重试入队 | queue=%s | action=%s | id=%s | run=%s",
                queue_type,
                action,
                item.get("id"),
                run_id,
            )
        except Exception as exc:
            self.logger.warning("重试入队失败 | queue=%s | action=%s | run=%s | error=%s", queue_type, action, run_id, exc)

    def _enqueue_failed_email_dispatch(
        self,
        *,
        run_id: str,
        send_logs: list[Any],
        dispatch_result: dict[str, Any],
    ) -> None:
        if not self._retry_enabled:
            return
        if not send_logs:
            return
        failed_email = [
            log
            for log in send_logs
            if str(getattr(log, "channel", "")).strip().lower() == "email"
            and str(getattr(log, "send_status", "")).strip().lower() != "success"
        ]
        if not failed_email:
            return
        subject = str(dispatch_result.get("subject") or "").strip()
        body = str(dispatch_result.get("body") or "").strip()
        if not subject or not body:
            return
        channels = sorted(
            {
                str(getattr(log, "channel", "")).strip()
                for log in failed_email
                if str(getattr(log, "channel", "")).strip()
            }
        )
        attachments = dispatch_result.get("attachments")
        idempotency_key = self._build_email_idempotency_key(subject=subject, body=body, channels=channels or ["email"])
        payload = {
            "subject": subject,
            "body": body,
            "channels": channels or ["email"],
            "attachments": [str(x) for x in (attachments or []) if str(x).strip()],
            "idempotency_key": idempotency_key,
        }
        self._enqueue_retry(
            queue_type="email",
            action="dispatch_digest",
            run_id=run_id,
            error="email_send_failed",
            payload=payload,
            dedupe_key=f"email-failed:{run_id}:{subject[:80]}",
            idempotency_key=idempotency_key,
        )

    def _enqueue_llm_timeout_retries(self, *, run_id: str, llm_error_codes: dict[str, int]) -> None:
        if not self._retry_enabled:
            return
        timeout_keys = {"connect_timeout", "read_timeout", "timeout"}
        for code, count in (llm_error_codes or {}).items():
            key = str(code or "").strip().lower()
            num = max(0, int(count or 0))
            if key not in timeout_keys or num <= 0:
                continue
            self._enqueue_retry(
                queue_type="llm_timeout",
                action="probe",
                run_id=run_id,
                error=f"llm_{key}:{num}",
                payload={"error_code": key, "count": num, "scope": "retry_llm_timeout"},
                dedupe_key=f"llm-timeout:{run_id}:{key}",
                idempotency_key=f"llm-timeout:{run_id}:{key}",
            )

    @staticmethod
    def _classify_retry_error_code(exc: Exception) -> str:
        text = str(exc or "").strip().lower()
        if not text:
            return "retry_failed"
        if "timeout" in text:
            return "timeout"
        if "not found" in text or "404" in text:
            return "not_found"
        if "permission" in text or "denied" in text:
            return "permission_denied"
        if "auth" in text or "login" in text or "token" in text:
            return "auth_failed"
        if "network" in text or "connection" in text or "connect" in text:
            return "network_error"
        if "smtp" in text:
            return "smtp_error"
        return "retry_failed"

    @staticmethod
    def _build_email_idempotency_key(*, subject: str, body: str, channels: list[str]) -> str:
        joined_channels = ",".join(sorted(str(x or "").strip().lower() for x in channels if str(x or "").strip()))
        source = f"{subject.strip()}|{joined_channels}|{body.strip()}"
        digest = hashlib.sha1(source.encode("utf-8", errors="ignore")).hexdigest()
        return f"email:{digest}"

    def _probe_xhs_fetch_failure(self, *, run_id: str, fetch_fail_events: list[dict[str, str]]) -> dict[str, Any]:
        try:
            diagnosis = self.collector.probe_status_diagnostics()
            joined_errors = " | ".join(str(item.get("error") or "") for item in (fetch_fail_events or []))
            category = "unknown_fetch_failure"
            if any(k in joined_errors.lower() for k in ("risk", "风控", "访问受限", "网络环境")):
                category = "risk_control"
            elif not diagnosis.get("mcp_connect"):
                category = "mcp_unreachable"
            elif not diagnosis.get("login_status"):
                category = "not_logged_in"
            diagnosis["run_id"] = run_id
            diagnosis["fetch_fail_streak"] = int(self._fetch_fail_streak)
            diagnosis["fetch_fail_events"] = list(fetch_fail_events or [])
            diagnosis["failure_category"] = category
            self.logger.warning(
                "XHS 失败诊断 | run=%s | streak=%s | category=%s | mcp=%s | login=%s | cookie=%s | reason=%s",
                run_id,
                self._fetch_fail_streak,
                category,
                diagnosis.get("mcp_connect"),
                diagnosis.get("login_status"),
                diagnosis.get("cookie_file_ready"),
                diagnosis.get("reason"),
            )
            return diagnosis
        except Exception as exc:
            self.logger.warning("XHS 失败诊断执行异常 | run=%s | error=%s", run_id, exc)
            return {
                "run_id": run_id,
                "fetch_fail_streak": int(self._fetch_fail_streak),
                "fetch_fail_events": list(fetch_fail_events or []),
                "error": str(exc)[:280],
            }

    def _build_retry_dispatch_fallback(
        self,
        *,
        run_id: str,
        mode: str,
        jobs: list[JobRecord],
        channels: list[str],
        attachments: list[str],
        reason: str,
    ) -> dict[str, Any]:
        subject = self.communication._build_subject(
            run_id=run_id,
            mode=mode,
            jobs=jobs,
            headline="notification retry",
        )
        body = self.communication._build_body(
            run_id=run_id,
            mode=mode,
            jobs=jobs,
            headline="notification retry",
            overview=f"主流程通知失败，已写入重试队列。reason={reason}",
            attachments=attachments,
        )
        return {
            "logs": [],
            "subject": subject,
            "body": body,
            "attachments": list(attachments),
            "channels": list(channels),
        }

    def _jobs_to_summary_records(self, run_id: str, jobs: list[JobRecord]) -> list[SummaryRecord]:
        output: list[SummaryRecord] = []
        for item in jobs:
            requirements = (item.requirements or "").strip()
            original_text = self._build_original_text_summary(item)
            summary_text = (
                f"公司：{item.company}\n"
                f"岗位：{item.position}\n"
                f"地点：{item.location}\n"
                f"岗位要求：{requirements or '未提取到明确要求'}\n"
                f"到岗时间：{item.arrival_time}\n"
                f"投递方式：{item.application_method}\n"
                f"风险等级：{item.risk_line}\n"
                f"简历匹配度：{item.match_score:.2f}\n"
                f"原文：{original_text}"
            )
            output.append(
                SummaryRecord(
                    run_id=run_id,
                    note_id=item.post_id,
                    keyword=self.settings.xhs.keyword,
                    publish_time=item.publish_time,
                    title=item.source_title or item.position,
                    author=item.author,
                    summary=summary_text,
                    confidence=1.0,
                    risk_flags=item.risk_line,
                    url=item.link,
                )
            )
        return output

    @classmethod
    def _build_original_text_summary(cls, item: JobRecord) -> str:
        original_text = str(item.original_text or "").strip()
        requirements = str(item.requirements or "").strip()
        if original_text:
            if cls._is_duplicate_text(original_text, requirements):
                return "原文与岗位要求高度重合，建议查看原帖链接获取完整上下文。"
            return original_text

        fallbacks: list[str] = []
        title = str(item.source_title or "").strip()
        comments = str(item.comments_preview or "").strip()
        if title:
            fallbacks.append(f"标题：{title}")
        if comments:
            fallbacks.append(f"评论线索：{comments}")
        if fallbacks:
            return "；".join(fallbacks)[:700]
        return "未抓取到可用原文摘要，请查看原帖链接/图片。"

    @staticmethod
    def _is_duplicate_text(a: str, b: str) -> bool:
        left = AutoSuccessorPipeline._normalize_compare_text(a)
        right = AutoSuccessorPipeline._normalize_compare_text(b)
        if not left or not right:
            return False
        if left == right:
            return True
        short, long = (left, right) if len(left) <= len(right) else (right, left)
        if len(short) >= 24 and short in long:
            overlap = len(short) / max(1, len(long))
            return overlap >= 0.7
        return False

    @staticmethod
    def _normalize_compare_text(text: str) -> str:
        value = str(text or "").lower().strip()
        if not value:
            return ""
        value = "".join(ch for ch in value if ch.isalnum() or "\u4e00" <= ch <= "\u9fff")
        return value

    def _summaries_to_jobs(self, summaries: list[SummaryRecord]) -> list[JobRecord]:
        jobs: list[JobRecord] = []
        for item in summaries:
            jobs.append(
                JobRecord(
                    run_id=item.run_id,
                    post_id=item.note_id,
                    company="",
                    position=item.title or "",
                    location="",
                    requirements=item.summary or "",
                    link=item.url,
                    publish_time=item.publish_time,
                    source_title=item.title,
                    comment_count=0,
                    comments_preview="",
                    original_text=item.summary or "",
                    author=item.author,
                    risk_line=item.risk_flags or "low",
                    match_score=0.0,
                    mode=self._normalize_mode(self.settings.agent.mode or "auto"),
                )
            )
        return jobs

    # backward compatibility for tests/callers
    def _load_latest_summaries_from_store(self, limit: int):
        jobs = self._load_latest_jobs_from_store(limit=limit)
        return self._jobs_to_summary_records(run_id="from-store", jobs=jobs)

    @staticmethod
    def _read_sheet_rows(ws) -> list[dict[str, Any]]:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return []
        headers = [str(x or "").strip() for x in headers_row]
        if not any(headers):
            headers = list(JOB_HEADERS)

        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, Any] = {}
            for idx, value in enumerate(values):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if key:
                    row[key] = value
            rows.append(row)
        return rows

    @staticmethod
    def _to_datetime(value: Any) -> datetime:
        epoch = datetime(1970, 1, 1, tzinfo=timezone.utc)
        if isinstance(value, datetime):
            dt = value
        else:
            text = str(value or "").strip()
            if not text:
                return epoch
            try:
                dt = datetime.fromisoformat(text)
            except Exception:
                return epoch
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt

    @staticmethod
    def _summarize_stage_timing(stage_records: list[dict[str, Any]]) -> dict[str, Any]:
        normalized: list[dict[str, Any]] = []
        failed_count = 0
        total_ms = 0

        for item in stage_records or []:
            name = str(item.get("name") or "").strip()
            status = str(item.get("status") or "").strip().lower()
            try:
                duration_ms = int(item.get("duration_ms") or 0)
            except Exception:
                duration_ms = 0
            duration_ms = max(0, duration_ms)
            total_ms += duration_ms
            if status == "failed":
                failed_count += 1
            normalized.append(
                {
                    "name": name,
                    "duration_ms": duration_ms,
                    "status": status or "success",
                }
            )

        count = len(normalized)
        avg_ms = int(total_ms / count) if count > 0 else 0
        top = sorted(normalized, key=lambda x: int(x.get("duration_ms") or 0), reverse=True)[:3]
        slowest = top[0] if top else {"name": "", "duration_ms": 0, "status": ""}

        return {
            "total_ms": total_ms,
            "avg_ms": avg_ms,
            "failed_count": failed_count,
            "slowest": slowest,
            "top_slow": top,
        }

    @staticmethod
    def _collect_stage_error_codes(stage_records: list[dict[str, Any]]) -> dict[str, int]:
        counters: dict[str, int] = {}
        for item in stage_records or []:
            status = str(item.get("status") or "").strip().lower()
            if status != "failed":
                continue
            code = str(item.get("error_code") or "").strip().lower() or "stage_failed"
            counters[code] = int(counters.get(code, 0)) + 1
        return counters
