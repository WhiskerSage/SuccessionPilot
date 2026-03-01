from __future__ import annotations

import json
import base64
import os
import re
import shutil
import smtplib
import subprocess
import sys
import threading
from collections import deque
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import yaml

from .config import ResumeConfig, load_settings
from .excel_store import JOB_HEADERS, RAW_HEADERS, SUMMARY_HEADERS
from .retry_queue import RetryQueue
from .resume_loader import ResumeLoader
from .text_utils import repair_mojibake


class RuntimeManager:
    _PROGRESS_RE = re.compile(r"\[\s*([#\-]{6,})\s*\]\s*(\d{1,3})%\s*\|\s*(.+)$")
    _RUN_ID_RE = re.compile(r"\brun=([A-Za-z0-9._:-]+)")

    def __init__(self, workspace: Path, config_path: Path) -> None:
        self.workspace = workspace
        self.config_path = config_path
        self._lock = threading.Lock()

        self._job_thread: threading.Thread | None = None
        self._job_proc: subprocess.Popen[str] | None = None
        self._job_logs: deque[str] = deque(maxlen=240)
        self._job_state: dict[str, Any] = {
            "running": False,
            "name": "",
            "started_at": "",
            "finished_at": "",
            "ok": None,
            "exit_code": None,
            "message": "",
            "log_tail": [],
        }

        self._daemon_proc: subprocess.Popen[str] | None = None
        self._daemon_reader: threading.Thread | None = None
        self._daemon_logs: deque[str] = deque(maxlen=360)
        self._daemon_started_at: str = ""
        self._daemon_last_exit: int | None = None
        self._daemon_cmd: list[str] = []

    def _now_iso(self) -> str:
        return datetime.now().isoformat(timespec="seconds")

    def _resolve_python(self) -> str:
        if sys.platform.startswith("win"):
            venv_python = self.workspace / ".venv" / "Scripts" / "python.exe"
        else:
            venv_python = self.workspace / ".venv" / "bin" / "python"
        if venv_python.exists():
            return str(venv_python)
        return sys.executable

    @staticmethod
    def _resolve_powershell() -> str:
        if sys.platform.startswith("win"):
            return "powershell"
        return "pwsh"

    @staticmethod
    def _build_subprocess_env() -> dict[str, str]:
        env = os.environ.copy()
        env.setdefault("PYTHONIOENCODING", "utf-8")
        return env

    def _build_main_command(
        self,
        *,
        run_once: bool = False,
        daemon: bool = False,
        mode: str | None = None,
        interval_minutes: int | None = None,
        send_latest: int | None = None,
    ) -> list[str]:
        command = [self._resolve_python(), "-m", "auto_successor.main", "--config", str(self.config_path)]
        if run_once:
            command.append("--run-once")
        if daemon:
            command.append("--daemon")
        if mode:
            command.extend(["--mode", mode])
        if interval_minutes and interval_minutes > 0:
            command.extend(["--interval-minutes", str(interval_minutes)])
        if send_latest and send_latest > 0:
            command.extend(["--send-latest", str(send_latest)])
        return command

    def _read_xhs_account_settings(self) -> tuple[str, str]:
        try:
            settings = load_settings(str(self.config_path))
            account = str(getattr(settings.xhs, "account", "default") or "default").strip() or "default"
            account_dir = (
                str(getattr(settings.xhs, "account_cookies_dir", "~/.xhs-mcp/accounts") or "~/.xhs-mcp/accounts").strip()
                or "~/.xhs-mcp/accounts"
            )
            return account, account_dir
        except Exception:
            return "default", "~/.xhs-mcp/accounts"

    def _start_job(self, name: str, command: list[str]) -> dict[str, Any]:
        blocked_message = ""
        with self._lock:
            if self._job_state.get("running"):
                blocked_message = f"已有任务运行中: {self._job_state.get('name')}"
            else:
                self._job_logs.clear()
                self._job_state = {
                    "running": True,
                    "name": name,
                    "started_at": self._now_iso(),
                    "finished_at": "",
                    "ok": None,
                    "exit_code": None,
                    "message": f"{name} 已启动",
                    "log_tail": [],
                }
                self._job_thread = threading.Thread(target=self._run_job, args=(name, command), daemon=True)
                self._job_thread.start()

        if blocked_message:
            return {"ok": False, "message": blocked_message, "runtime": self.status()}

        return {"ok": True, "message": f"{name} 已启动", "runtime": self.status()}

    def _run_job(self, name: str, command: list[str]) -> None:
        ok = False
        exit_code: int | None = None
        message = ""

        try:
            proc = subprocess.Popen(
                command,
                cwd=str(self.workspace),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=self._build_subprocess_env(),
            )
        except Exception as exc:
            with self._lock:
                self._job_logs.append(f"任务启动失败: {exc}")
                self._job_state = {
                    "running": False,
                    "name": name,
                    "started_at": self._job_state.get("started_at", self._now_iso()),
                    "finished_at": self._now_iso(),
                    "ok": False,
                    "exit_code": -1,
                    "message": f"任务启动失败: {exc}",
                    "log_tail": list(self._job_logs),
                }
            return

        with self._lock:
            self._job_proc = proc
            self._job_logs.append("$ " + " ".join(command))
            self._job_state["log_tail"] = list(self._job_logs)

        try:
            if proc.stdout is not None:
                for raw in proc.stdout:
                    line = repair_mojibake(raw.rstrip())
                    if not line:
                        continue
                    with self._lock:
                        self._job_logs.append(line)
                        self._job_state["log_tail"] = list(self._job_logs)

            exit_code = proc.wait()
            ok = exit_code == 0
            message = "任务完成" if ok else f"任务失败，退出码 {exit_code}"
        except Exception as exc:
            exit_code = -1
            ok = False
            message = f"任务异常: {exc}"
            with self._lock:
                self._job_logs.append(message)
                self._job_state["log_tail"] = list(self._job_logs)
        finally:
            with self._lock:
                self._job_proc = None
                self._job_state = {
                    "running": False,
                    "name": name,
                    "started_at": self._job_state.get("started_at", self._now_iso()),
                    "finished_at": self._now_iso(),
                    "ok": ok,
                    "exit_code": exit_code,
                    "message": message,
                    "log_tail": list(self._job_logs),
                }

    def start_run_once(self, mode: str = "auto") -> dict[str, Any]:
        command = self._build_main_command(run_once=True, mode=mode)
        return self._start_job(name="单次抓取", command=command)

    def start_send_latest(self, limit: int = 5) -> dict[str, Any]:
        limit = max(1, int(limit))
        command = self._build_main_command(send_latest=limit)
        return self._start_job(name=f"发送最新 {limit} 条", command=command)

    def start_xhs_login(self, timeout_seconds: int = 180) -> dict[str, Any]:
        script = self.workspace / "scripts" / "xhs_login.ps1"
        account, account_dir = self._read_xhs_account_settings()
        command = [
            self._resolve_powershell(),
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-Timeout",
            str(max(30, int(timeout_seconds))),
            "-Account",
            account,
            "-AccountCookiesDir",
            account_dir,
        ]
        return self._start_job(name="XHS 扫码登录", command=command)

    def check_xhs_status(self) -> dict[str, Any]:
        script = self.workspace / "scripts" / "xhs_status.ps1"
        account, account_dir = self._read_xhs_account_settings()
        command = [
            self._resolve_powershell(),
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            str(script),
            "-Account",
            account,
            "-AccountCookiesDir",
            account_dir,
        ]
        try:
            result = subprocess.run(
                command,
                cwd=str(self.workspace),
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=90,
                env=self._build_subprocess_env(),
            )
            raw_lines = [repair_mojibake(x.strip()) for x in [result.stdout, result.stderr] if str(x or "").strip()]
            raw = "\n".join(raw_lines).strip()
            parsed: dict[str, Any] | None = None
            for line in reversed([x for x in raw.splitlines() if x.strip()]):
                text = line.strip()
                if not text.startswith("{"):
                    continue
                try:
                    maybe = json.loads(text)
                except Exception:
                    continue
                if isinstance(maybe, dict):
                    parsed = maybe
                    break
            return {
                "ok": result.returncode == 0,
                "exit_code": result.returncode,
                "message": "状态检查成功" if result.returncode == 0 else f"状态检查失败，退出码 {result.returncode}",
                "status": parsed or {},
                "output": raw[-1200:],
                "runtime": self.status(),
            }
        except Exception as exc:
            return {"ok": False, "message": f"状态检查异常: {exc}", "status": {}, "output": "", "runtime": self.status()}

    def start_daemon(self, mode: str = "auto", interval_minutes: int | None = None) -> dict[str, Any]:
        already_running = False
        with self._lock:
            if self._daemon_proc and self._daemon_proc.poll() is None:
                already_running = True
        if already_running:
            return {"ok": False, "message": "自动运行已在执行", "runtime": self.status()}

        command = self._build_main_command(daemon=True, mode=mode, interval_minutes=interval_minutes)
        try:
            proc = subprocess.Popen(
                command,
                cwd=str(self.workspace),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=self._build_subprocess_env(),
            )
        except Exception as exc:
            return {"ok": False, "message": f"启动自动运行失败: {exc}", "runtime": self.status()}

        with self._lock:
            self._daemon_proc = proc
            self._daemon_started_at = self._now_iso()
            self._daemon_last_exit = None
            self._daemon_cmd = list(command)
            self._daemon_logs.clear()
            self._daemon_logs.append("$ " + " ".join(command))
            self._daemon_reader = threading.Thread(target=self._drain_daemon_output, args=(proc,), daemon=True)
            self._daemon_reader.start()

        return {"ok": True, "message": "自动运行已启动", "runtime": self.status()}

    def _drain_daemon_output(self, proc: subprocess.Popen[str]) -> None:
        if proc.stdout is None:
            return
        for raw in proc.stdout:
            line = repair_mojibake(raw.rstrip())
            if not line:
                continue
            with self._lock:
                self._daemon_logs.append(line)
        with self._lock:
            if self._daemon_proc is proc:
                self._daemon_last_exit = proc.poll()

    def stop_daemon(self) -> dict[str, Any]:
        no_daemon = False
        with self._lock:
            proc = self._daemon_proc
            if not proc or proc.poll() is not None:
                self._daemon_proc = None
                no_daemon = True
        if no_daemon:
            return {"ok": False, "message": "自动运行未在执行", "runtime": self.status()}

        try:
            proc.terminate()
            proc.wait(timeout=8)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

        with self._lock:
            self._daemon_last_exit = proc.poll()
            self._daemon_logs.append(f"[{self._now_iso()}] 自动运行已停止")
            self._daemon_proc = None

        return {"ok": True, "message": "自动运行已停止", "runtime": self.status()}

    def stop_job(self) -> dict[str, Any]:
        no_job = False
        with self._lock:
            proc = self._job_proc
            if not proc or proc.poll() is not None:
                no_job = True
        if no_job:
            return {"ok": False, "message": "当前没有运行中的任务", "runtime": self.status()}

        try:
            proc.terminate()
        except Exception as exc:
            return {"ok": False, "message": f"停止任务失败: {exc}", "runtime": self.status()}
        return {"ok": True, "message": "已请求停止当前任务", "runtime": self.status()}

    @classmethod
    def _extract_progress(cls, logs: list[Any]) -> dict[str, Any]:
        if not isinstance(logs, list):
            return {}
        for raw in reversed(logs):
            line = str(raw or "").strip()
            if not line:
                continue
            match = cls._PROGRESS_RE.search(line)
            if not match:
                continue
            try:
                percent = max(0, min(100, int(match.group(2))))
            except Exception:
                percent = 0
            message = str(match.group(3) or "").strip()
            run_match = cls._RUN_ID_RE.search(line)
            run_id = str(run_match.group(1) if run_match else "").strip()
            return {
                "percent": percent,
                "message": message,
                "run_id": run_id,
                "raw_line": line,
            }
        return {}

    def _job_view(self) -> dict[str, Any]:
        with self._lock:
            out = dict(self._job_state)
            logs = out.get("log_tail")
            out["progress"] = self._extract_progress(logs if isinstance(logs, list) else [])
            return out

    def _daemon_view(self) -> dict[str, Any]:
        with self._lock:
            proc = self._daemon_proc
            running = bool(proc and proc.poll() is None)
            out = {
                "running": running,
                "pid": proc.pid if running else None,
                "started_at": self._daemon_started_at,
                "exit_code": None if running else self._daemon_last_exit,
                "command": list(self._daemon_cmd),
                "log_tail": list(self._daemon_logs),
            }
            out["progress"] = self._extract_progress(out["log_tail"])
            return out

    def status(self) -> dict[str, Any]:
        return {"job": self._job_view(), "daemon": self._daemon_view(), "updated_at": self._now_iso()}


class DataBackend:
    def __init__(self, workspace: Path) -> None:
        self.workspace = workspace
        self.data_dir = workspace / "data"
        self.runs_dir = self.data_dir / "runs"
        self.excel_path = self.data_dir / "output.xlsx"
        self.config_path = workspace / "config" / "config.yaml"
        self.retry_queue_path = self._resolve_retry_queue_path()
        self.runtime = RuntimeManager(workspace=workspace, config_path=self.config_path)
        self._resume_loader: ResumeLoader | None = None
        self._resume_source_path = workspace / "config" / "resume.txt"
        self._resume_text_path = workspace / "data" / "resume_text.txt"

    def load_summary(self) -> dict[str, Any]:
        rows = self._load_workbook_rows()
        raw = rows.get("raw_notes", [])
        summaries = rows.get("succession_summary", [])
        jobs = rows.get("jobs", [])
        sends = rows.get("send_log", [])
        runs = self._load_runs(limit=1)

        latest_run_id = runs[0]["run_id"] if runs else ""
        latest_run_time = runs[0]["recorded_at"] if runs else ""

        return {
            "raw_count": len(raw),
            "summary_count": len(summaries),
            "jobs_count": len(jobs),
            "send_count": len(sends),
            "latest_run_id": latest_run_id,
            "latest_run_time": latest_run_time,
            "digest_interval_minutes": self._read_digest_interval(),
        }

    def _build_merged_leads(
        self,
        q: str = "",
        summary_only: bool = False,
        status_filter: str = "all",
        dedupe_filter: str = "all",
    ) -> list[dict[str, Any]]:
        rows = self._load_workbook_rows()
        raw = rows.get("raw_notes", [])
        summaries = rows.get("succession_summary", [])
        jobs = rows.get("jobs", [])

        by_summary = {str(x.get("note_id") or ""): x for x in summaries}
        by_job = {str(x.get("PostID") or ""): x for x in jobs}

        merged: list[dict[str, Any]] = []
        for item in raw:
            note_id = str(item.get("note_id") or "")
            if not note_id:
                continue

            summary = by_summary.get(note_id, {})
            job = by_job.get(note_id, {})

            like_count = self._to_int(item.get("like_count"))
            comment_count = self._to_int(item.get("comment_count"))
            publish_ts = self._to_int(item.get("publish_timestamp"))
            if publish_ts <= 0:
                publish_ts = self._to_epoch(item.get("publish_time"))
            publish_time_value = str(item.get("publish_time") or "")
            publish_time_display = self._format_time_from_epoch_or_iso(publish_ts, publish_time_value)
            first_seen_at = str(item.get("first_seen_at") or item.get("fetched_at") or "").strip()
            updated_at = str(item.get("updated_at") or item.get("fetched_at") or first_seen_at).strip()
            first_seen_ts = self._to_epoch(first_seen_at)
            updated_ts = self._to_epoch(updated_at)
            dedupe_status = "updated" if first_seen_ts > 0 and updated_ts > first_seen_ts else "new"
            status_key = self._resolve_status_key(
                like_count=like_count,
                comment_count=comment_count,
                has_summary=bool(summary),
                has_job=bool(job),
            )
            status = self._resolve_status(
                like_count=like_count,
                comment_count=comment_count,
                has_summary=bool(summary),
                has_job=bool(job),
            )

            merged.append(
                {
                    "note_id": note_id,
                    "publish_time": publish_time_value,
                    "publish_time_text": str(item.get("publish_time_text") or ""),
                    "publish_time_quality": str(item.get("publish_time_quality") or ""),
                    "publish_timestamp": publish_ts,
                    "publish_time_display": publish_time_display,
                    "title": str(item.get("title") or ""),
                    "author": str(item.get("author") or ""),
                    "like_count": like_count,
                    "comment_count": comment_count,
                    "share_count": self._to_int(item.get("share_count")),
                    "url": str(item.get("url") or ""),
                    "detail_text": str(item.get("detail_text") or ""),
                    "comments_preview": str(item.get("comments_preview") or ""),
                    "summary": str(summary.get("summary") or ""),
                    "risk_flags": str(summary.get("risk_flags") or ""),
                    "company": str(job.get("Company") or ""),
                    "position": str(job.get("Position") or ""),
                    "location": str(job.get("Location") or ""),
                    "requirements": str(job.get("Requirements") or ""),
                    "status": status,
                    "status_key": status_key,
                    "first_seen_at": first_seen_at,
                    "updated_at": updated_at,
                    "dedupe_status": dedupe_status,
                }
            )

        merged.sort(key=lambda x: int(x.get("publish_timestamp") or 0), reverse=True)

        if summary_only:
            merged = [row for row in merged if str(row.get("summary") or "").strip()]

        status_key_filter = str(status_filter or "all").strip().lower() or "all"
        if status_key_filter != "all":
            merged = [row for row in merged if str(row.get("status_key") or "").strip().lower() == status_key_filter]

        dedupe_key_filter = str(dedupe_filter or "all").strip().lower() or "all"
        if dedupe_key_filter in {"new", "updated"}:
            merged = [row for row in merged if str(row.get("dedupe_status") or "").strip().lower() == dedupe_key_filter]

        if q:
            key = q.strip().lower()
            merged = [
                row
                for row in merged
                if key
                in " ".join(
                    [
                        str(row.get("title") or ""),
                        str(row.get("author") or ""),
                        str(row.get("summary") or ""),
                        str(row.get("company") or ""),
                        str(row.get("position") or ""),
                        str(row.get("location") or ""),
                    ]
                ).lower()
            ]

        return merged

    def load_leads_page(
        self,
        *,
        page: int = 1,
        page_size: int = 30,
        q: str = "",
        summary_only: bool = False,
        status_filter: str = "all",
        dedupe_filter: str = "all",
    ) -> dict[str, Any]:
        page = max(1, int(page))
        page_size = max(1, min(int(page_size), 200))

        merged = self._build_merged_leads(
            q=q,
            summary_only=summary_only,
            status_filter=status_filter,
            dedupe_filter=dedupe_filter,
        )
        total = len(merged)
        total_pages = max(1, (total + page_size - 1) // page_size)
        safe_page = min(page, total_pages)

        start = (safe_page - 1) * page_size
        end = start + page_size
        items = merged[start:end]

        return {
            "items": items,
            "total": total,
            "page": safe_page,
            "page_size": page_size,
            "total_pages": total_pages,
        }

    # Backward compatibility for existing callers.
    def load_leads(self, limit: int = 200, q: str = "") -> list[dict[str, Any]]:
        merged = self._build_merged_leads(q=q, summary_only=False)
        return merged[: max(1, int(limit))]

    def update_lead_fields(self, *, note_id: str, fields: dict[str, Any]) -> dict[str, Any]:
        target_note_id = str(note_id or "").strip()
        if not target_note_id:
            raise ValueError("missing note_id")
        if not isinstance(fields, dict):
            raise ValueError("fields must be an object")
        if not self.excel_path.exists():
            raise FileNotFoundError(f"excel not found: {self.excel_path}")

        editable_keys = {"title", "company", "position", "location", "requirements", "summary", "detail_text"}
        normalized: dict[str, str] = {}
        for key in editable_keys:
            if key in fields:
                normalized[key] = self._normalize_edit_text(fields.get(key))
        if not normalized:
            raise ValueError("no editable fields provided")

        wb = load_workbook(self.excel_path)
        try:
            ws_raw, raw_headers = self._ensure_sheet_with_headers(wb, "raw_notes", RAW_HEADERS)
            ws_summary, summary_headers = self._ensure_sheet_with_headers(wb, "succession_summary", SUMMARY_HEADERS)
            ws_jobs, job_headers = self._ensure_sheet_with_headers(wb, "jobs", JOB_HEADERS)

            raw_row = self._find_row_index(ws_raw, raw_headers, key_column="note_id", key_value=target_note_id)
            if raw_row <= 0:
                raise ValueError(f"note not found in raw_notes: {target_note_id}")

            now_iso = datetime.now().isoformat(timespec="seconds")
            raw_values = self._row_snapshot(ws_raw, raw_headers, raw_row)

            # raw_notes updates
            if "title" in normalized:
                self._set_cell(ws_raw, raw_headers, raw_row, "title", normalized["title"])
            if "detail_text" in normalized:
                self._set_cell(ws_raw, raw_headers, raw_row, "detail_text", normalized["detail_text"])
            self._set_cell(ws_raw, raw_headers, raw_row, "updated_at", now_iso)

            # jobs updates (create row when missing so manual fill can promote new leads to actionable jobs)
            if any(key in normalized for key in ("company", "position", "location", "requirements", "title")):
                jobs_row = self._find_row_index(ws_jobs, job_headers, key_column="PostID", key_value=target_note_id)
                if jobs_row <= 0:
                    jobs_row = self._append_row(ws_jobs)
                    self._set_cell(ws_jobs, job_headers, jobs_row, "PostID", target_note_id)
                    self._set_cell(ws_jobs, job_headers, jobs_row, "run_id", str(raw_values.get("run_id") or "manual"))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "publish_time", str(raw_values.get("publish_time") or now_iso))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "author", str(raw_values.get("author") or ""))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "Link", str(raw_values.get("url") or ""))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "source_title", str(raw_values.get("title") or ""))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "original_text", str(raw_values.get("detail_text") or ""))
                    self._set_cell(ws_jobs, job_headers, jobs_row, "mode", "manual_edit")
                if "company" in normalized:
                    self._set_cell(ws_jobs, job_headers, jobs_row, "Company", normalized["company"])
                if "position" in normalized:
                    self._set_cell(ws_jobs, job_headers, jobs_row, "Position", normalized["position"])
                if "location" in normalized:
                    self._set_cell(ws_jobs, job_headers, jobs_row, "Location", normalized["location"])
                if "requirements" in normalized:
                    self._set_cell(ws_jobs, job_headers, jobs_row, "Requirements", normalized["requirements"])
                if "title" in normalized:
                    self._set_cell(ws_jobs, job_headers, jobs_row, "source_title", normalized["title"])

            # succession_summary updates (create row when missing so summary page can render immediately)
            if any(key in normalized for key in ("summary", "title")):
                summary_row = self._find_row_index(
                    ws_summary,
                    summary_headers,
                    key_column="note_id",
                    key_value=target_note_id,
                )
                if summary_row <= 0:
                    summary_row = self._append_row(ws_summary)
                    publish_timestamp = self._to_int(raw_values.get("publish_timestamp"))
                    self._set_cell(ws_summary, summary_headers, summary_row, "run_id", str(raw_values.get("run_id") or "manual"))
                    self._set_cell(ws_summary, summary_headers, summary_row, "note_id", target_note_id)
                    self._set_cell(ws_summary, summary_headers, summary_row, "keyword", str(raw_values.get("keyword") or ""))
                    self._set_cell(ws_summary, summary_headers, summary_row, "publish_time", str(raw_values.get("publish_time") or now_iso))
                    self._set_cell(ws_summary, summary_headers, summary_row, "publish_timestamp", publish_timestamp)
                    self._set_cell(ws_summary, summary_headers, summary_row, "author", str(raw_values.get("author") or ""))
                    self._set_cell(ws_summary, summary_headers, summary_row, "url", str(raw_values.get("url") or ""))
                    self._set_cell(ws_summary, summary_headers, summary_row, "confidence", 0.0)
                    self._set_cell(ws_summary, summary_headers, summary_row, "risk_flags", "manual_edit")
                    self._set_cell(ws_summary, summary_headers, summary_row, "created_at", now_iso)
                if "summary" in normalized:
                    self._set_cell(ws_summary, summary_headers, summary_row, "summary", normalized["summary"])
                if "title" in normalized:
                    self._set_cell(ws_summary, summary_headers, summary_row, "title", normalized["title"])

            wb.save(self.excel_path)
        finally:
            wb.close()

        refreshed = self._build_merged_leads()
        lead = next((item for item in refreshed if str(item.get("note_id") or "") == target_note_id), None)
        return {
            "ok": True,
            "message": "线索字段已更新",
            "note_id": target_note_id,
            "updated_fields": sorted(normalized.keys()),
            "updated_at": now_iso,
            "lead": lead or {},
        }

    def load_runs(self, limit: int = 20) -> list[dict[str, Any]]:
        return self._load_runs(limit=limit)

    def load_performance(self, limit: int = 50) -> dict[str, Any]:
        safe_limit = max(1, min(int(limit), 300))
        runs = self._load_runs(limit=safe_limit)
        if not runs:
            return {
                "sample_size": 0,
                "stage_total_ms": {"avg": 0, "p50": 0, "p95": 0, "min": 0, "max": 0},
                "stage_avg_ms": {"avg": 0, "p50": 0, "p95": 0, "min": 0, "max": 0},
                "stage_failed_runs": 0,
                "stage_failed_rate": 0.0,
                "llm_fail_total": 0,
                "fetch_fail_total": 0,
                "detail_attempted_total": 0,
                "detail_success_total": 0,
                "detail_success_rate": 0.0,
                "llm_timeout_count_total": 0,
                "llm_timeout_rate": 0.0,
                "detail_missing_total": 0,
                "detail_missing_target_total": 0,
                "detail_missing_rate": 0.0,
                "alert_triggered_total": 0,
                "alert_notified_total": 0,
                "alert_triggered_runs": 0,
                "alert_triggered_rate": 0.0,
                "alert_codes": [],
                "error_codes": [],
                "slow_stages": [],
                "quality": self._empty_quality_metrics(),
                "generated_at": datetime.now().isoformat(timespec="seconds"),
            }

        totals = sorted(self._to_int(item.get("stage_total_ms")) for item in runs if self._to_int(item.get("stage_total_ms")) > 0)
        avgs = sorted(self._to_int(item.get("stage_avg_ms")) for item in runs if self._to_int(item.get("stage_avg_ms")) > 0)
        sample_size = len(runs)
        stage_failed_runs = sum(1 for item in runs if self._to_int(item.get("stage_failed_count")) > 0)
        llm_fail_total = sum(self._to_int(item.get("llm_fail")) for item in runs)
        llm_calls_total = sum(self._to_int(item.get("llm_calls")) for item in runs)
        llm_timeout_count_total = sum(self._to_int(item.get("llm_timeout_count")) for item in runs)
        llm_timeout_rate = float(llm_timeout_count_total) / float(llm_calls_total) if llm_calls_total > 0 else 0.0
        fetch_fail_total = sum(self._to_int(item.get("fetch_fail_count_run")) for item in runs)
        detail_attempted_total = sum(self._to_int(item.get("detail_attempted")) for item in runs)
        detail_success_total = sum(self._to_int(item.get("detail_success")) for item in runs)
        detail_success_rate = (
            float(detail_success_total) / float(detail_attempted_total) if detail_attempted_total > 0 else 0.0
        )
        detail_missing_total = sum(self._to_int(item.get("detail_missing")) for item in runs)
        detail_missing_target_total = sum(self._to_int(item.get("detail_target_notes")) for item in runs)
        detail_missing_rate = (
            float(detail_missing_total) / float(detail_missing_target_total)
            if detail_missing_target_total > 0
            else 0.0
        )
        alert_triggered_total = sum(self._to_int(item.get("alerts_triggered_count")) for item in runs)
        alert_notified_total = sum(self._to_int(item.get("alerts_notified_count")) for item in runs)
        alert_triggered_runs = sum(1 for item in runs if self._to_int(item.get("alerts_triggered_count")) > 0)

        error_codes: dict[str, int] = {}
        alert_codes: dict[str, int] = {}
        slow_stage_metrics: dict[str, dict[str, Any]] = {}
        for item in runs:
            codes = item.get("error_codes")
            if isinstance(codes, dict):
                for code, count in codes.items():
                    key = str(code or "").strip().lower()
                    if not key:
                        continue
                    error_codes[key] = self._to_int(error_codes.get(key)) + self._to_int(count)
            run_alert_codes = item.get("alert_codes")
            if isinstance(run_alert_codes, list):
                for code in run_alert_codes:
                    key = str(code or "").strip().lower()
                    if not key:
                        continue
                    alert_codes[key] = self._to_int(alert_codes.get(key)) + 1

            slow = item.get("slow_stages")
            if not isinstance(slow, list):
                continue
            for stage in slow:
                if not isinstance(stage, dict):
                    continue
                name = str(stage.get("name") or "").strip() or "unknown"
                duration = self._to_int(stage.get("duration_ms"))
                bucket = slow_stage_metrics.get(name)
                if bucket is None:
                    bucket = {"name": name, "count": 0, "sum_ms": 0, "max_ms": 0, "durations": []}
                    slow_stage_metrics[name] = bucket
                bucket["count"] = self._to_int(bucket.get("count")) + 1
                bucket["sum_ms"] = self._to_int(bucket.get("sum_ms")) + duration
                bucket["max_ms"] = max(self._to_int(bucket.get("max_ms")), duration)
                durations = bucket.get("durations")
                if isinstance(durations, list):
                    durations.append(duration)

        error_list = [
            {"code": code, "count": self._to_int(count)}
            for code, count in error_codes.items()
            if self._to_int(count) > 0
        ]
        error_list.sort(key=lambda item: self._to_int(item.get("count")), reverse=True)
        alert_code_list = [
            {"code": code, "count": self._to_int(count)}
            for code, count in alert_codes.items()
            if self._to_int(count) > 0
        ]
        alert_code_list.sort(key=lambda item: self._to_int(item.get("count")), reverse=True)

        slow_stage_list: list[dict[str, Any]] = []
        for value in slow_stage_metrics.values():
            durations = sorted(self._to_int(x) for x in value.get("durations", []) if self._to_int(x) > 0)
            count = self._to_int(value.get("count"))
            sum_ms = self._to_int(value.get("sum_ms"))
            slow_stage_list.append(
                {
                    "name": str(value.get("name") or "unknown"),
                    "count": count,
                    "avg_ms": int(sum_ms / max(1, count)),
                    "p95_ms": self._percentile(durations, 0.95),
                    "max_ms": self._to_int(value.get("max_ms")),
                }
            )
        slow_stage_list.sort(
            key=lambda item: (self._to_int(item.get("count")), self._to_int(item.get("p95_ms"))),
            reverse=True,
        )
        quality = self._build_quality_metrics(runs=runs)

        return {
            "sample_size": sample_size,
            "stage_total_ms": self._metric_summary(totals),
            "stage_avg_ms": self._metric_summary(avgs),
            "stage_failed_runs": stage_failed_runs,
            "stage_failed_rate": float(stage_failed_runs) / float(sample_size),
            "llm_fail_total": llm_fail_total,
            "fetch_fail_total": fetch_fail_total,
            "detail_attempted_total": detail_attempted_total,
            "detail_success_total": detail_success_total,
            "detail_success_rate": detail_success_rate,
            "llm_timeout_count_total": llm_timeout_count_total,
            "llm_timeout_rate": llm_timeout_rate,
            "detail_missing_total": detail_missing_total,
            "detail_missing_target_total": detail_missing_target_total,
            "detail_missing_rate": detail_missing_rate,
            "alert_triggered_total": alert_triggered_total,
            "alert_notified_total": alert_notified_total,
            "alert_triggered_runs": alert_triggered_runs,
            "alert_triggered_rate": float(alert_triggered_runs) / float(sample_size),
            "alert_codes": alert_code_list[:12],
            "error_codes": error_list[:12],
            "slow_stages": slow_stage_list[:10],
            "quality": quality,
            "generated_at": datetime.now().isoformat(timespec="seconds"),
        }

    @classmethod
    def _metric_summary(cls, values: list[int]) -> dict[str, int]:
        numbers = sorted(cls._to_int(item) for item in values if cls._to_int(item) > 0)
        if not numbers:
            return {"avg": 0, "p50": 0, "p95": 0, "min": 0, "max": 0}
        return {
            "avg": int(sum(numbers) / max(1, len(numbers))),
            "p50": cls._percentile(numbers, 0.50),
            "p95": cls._percentile(numbers, 0.95),
            "min": numbers[0],
            "max": numbers[-1],
        }

    @staticmethod
    def _percentile(values: list[int], q: float) -> int:
        if not values:
            return 0
        if len(values) == 1:
            return int(values[0])
        q = max(0.0, min(1.0, float(q)))
        pos = (len(values) - 1) * q
        lower = int(pos)
        upper = min(len(values) - 1, lower + 1)
        if upper == lower:
            return int(values[lower])
        weight = pos - lower
        return int(round(values[lower] + (values[upper] - values[lower]) * weight))

    @classmethod
    def _safe_rate(cls, numerator: int, denominator: int) -> float:
        num = cls._to_int(numerator)
        den = cls._to_int(denominator)
        return float(num) / float(den) if den > 0 else 0.0

    @classmethod
    def _empty_quality_metrics(cls) -> dict[str, Any]:
        return {
            "raw_total": 0,
            "jobs_total": 0,
            "detail_fill_rate": 0.0,
            "structured_complete_rate": 0.0,
            "company_fill_rate": 0.0,
            "position_fill_rate": 0.0,
            "location_fill_rate": 0.0,
            "requirements_fill_rate": 0.0,
            "recent_extraction_hit_rate": 0.0,
            "recent_llm_success_rate": 0.0,
            "recent_detail_fill_rate": 0.0,
            "missing_fields_top": [],
            "trend": [],
        }

    def _build_quality_metrics(self, *, runs: list[dict[str, Any]]) -> dict[str, Any]:
        try:
            rows = self._load_workbook_rows()
        except Exception:
            rows = {"raw_notes": [], "jobs": []}
        raw_rows = rows.get("raw_notes", []) if isinstance(rows.get("raw_notes"), list) else []
        jobs_rows = rows.get("jobs", []) if isinstance(rows.get("jobs"), list) else []
        raw_total = len(raw_rows)
        jobs_total = len(jobs_rows)

        detail_filled = sum(1 for item in raw_rows if str(item.get("detail_text") or "").strip())

        required_mapping = {
            "company": "Company",
            "position": "Position",
            "location": "Location",
            "requirements": "Requirements",
        }
        fill_counts: dict[str, int] = {}
        missing_counts: dict[str, int] = {}
        complete_count = 0
        for row in jobs_rows:
            present_flags = []
            for label, key in required_mapping.items():
                value = str(row.get(key) or "").strip()
                present = bool(value)
                present_flags.append(present)
                fill_counts[label] = self._to_int(fill_counts.get(label)) + (1 if present else 0)
                missing_counts[label] = self._to_int(missing_counts.get(label)) + (0 if present else 1)
            if all(present_flags):
                complete_count += 1

        recent_runs = runs[: max(1, min(20, len(runs)))]
        target_total = sum(self._to_int(item.get("target_notes")) for item in recent_runs)
        jobs_total_recent = sum(self._to_int(item.get("jobs")) for item in recent_runs)
        llm_calls_total = sum(self._to_int(item.get("llm_calls")) for item in recent_runs)
        llm_success_total = sum(self._to_int(item.get("llm_success")) for item in recent_runs)
        detail_target_total = sum(self._to_int(item.get("detail_target_notes")) for item in recent_runs)
        detail_missing_total = sum(self._to_int(item.get("detail_missing")) for item in recent_runs)

        trend_rows: list[dict[str, Any]] = []
        for item in recent_runs[:12]:
            target_notes = self._to_int(item.get("target_notes"))
            jobs = self._to_int(item.get("jobs"))
            llm_calls = self._to_int(item.get("llm_calls"))
            llm_success = self._to_int(item.get("llm_success"))
            detail_target = self._to_int(item.get("detail_target_notes"))
            detail_missing = self._to_int(item.get("detail_missing"))
            trend_rows.append(
                {
                    "run_id": str(item.get("run_id") or ""),
                    "recorded_at": str(item.get("recorded_at") or ""),
                    "extraction_hit_rate": self._safe_rate(jobs, target_notes),
                    "llm_success_rate": self._safe_rate(llm_success, llm_calls),
                    "detail_fill_rate": self._safe_rate(max(0, detail_target - detail_missing), detail_target),
                }
            )

        missing_top = [
            {"field": key, "missing": self._to_int(value)}
            for key, value in missing_counts.items()
            if self._to_int(value) > 0
        ]
        missing_top.sort(key=lambda item: self._to_int(item.get("missing")), reverse=True)

        return {
            "raw_total": raw_total,
            "jobs_total": jobs_total,
            "detail_fill_rate": self._safe_rate(detail_filled, raw_total),
            "structured_complete_rate": self._safe_rate(complete_count, jobs_total),
            "company_fill_rate": self._safe_rate(fill_counts.get("company", 0), jobs_total),
            "position_fill_rate": self._safe_rate(fill_counts.get("position", 0), jobs_total),
            "location_fill_rate": self._safe_rate(fill_counts.get("location", 0), jobs_total),
            "requirements_fill_rate": self._safe_rate(fill_counts.get("requirements", 0), jobs_total),
            "recent_extraction_hit_rate": self._safe_rate(jobs_total_recent, target_total),
            "recent_llm_success_rate": self._safe_rate(llm_success_total, llm_calls_total),
            "recent_detail_fill_rate": self._safe_rate(max(0, detail_target_total - detail_missing_total), detail_target_total),
            "missing_fields_top": missing_top[:4],
            "trend": trend_rows,
        }

    def load_runtime(self) -> dict[str, Any]:
        return self.runtime.status()

    def load_run_detail(self, run_id: str) -> dict[str, Any]:
        target = self._safe_run_id(run_id)
        if not target:
            raise ValueError("invalid run_id")
        file = self.runs_dir / f"{target}.json"
        if not file.exists():
            raise FileNotFoundError(f"run not found: {target}")
        payload = json.loads(file.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            raise ValueError("invalid run snapshot payload")

        stage_records = payload.get("stage_records") if isinstance(payload.get("stage_records"), list) else []
        failed_stages = [
            item
            for item in stage_records
            if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "failed"
        ]
        stats = payload.get("stats") if isinstance(payload.get("stats"), dict) else {}
        fetch_fail_events = payload.get("fetch_fail_events") if isinstance(payload.get("fetch_fail_events"), list) else []
        xhs_diagnosis = payload.get("xhs_diagnosis") if isinstance(payload.get("xhs_diagnosis"), dict) else {}
        retry = payload.get("retry") if isinstance(payload.get("retry"), dict) else {}
        stage_error_codes = stats.get("stage_error_codes") if isinstance(stats.get("stage_error_codes"), dict) else {}
        llm_error_codes = stats.get("llm_error_codes") if isinstance(stats.get("llm_error_codes"), dict) else {}
        alerts_triggered = stats.get("alerts_triggered") if isinstance(stats.get("alerts_triggered"), list) else []
        alerts_notified = stats.get("alerts_notified") if isinstance(stats.get("alerts_notified"), list) else []

        return {
            "run_id": str(payload.get("run_id") or target),
            "recorded_at": str(payload.get("recorded_at") or ""),
            "mode": str(payload.get("mode") or stats.get("mode") or ""),
            "notification_mode": str(payload.get("notification_mode") or stats.get("notification_mode") or ""),
            "stats": stats,
            "stage_records": stage_records,
            "failed_stages": failed_stages,
            "fetch_fail_events": fetch_fail_events,
            "xhs_diagnosis": xhs_diagnosis,
            "retry": retry,
            "stage_error_codes": stage_error_codes,
            "llm_error_codes": llm_error_codes,
            "alerts_triggered": alerts_triggered,
            "alerts_notified": alerts_notified,
        }

    def load_retry_queue_view(
        self,
        *,
        status: str = "all",
        queue_type: str = "all",
        limit: int = 120,
    ) -> dict[str, Any]:
        queue = self._load_retry_queue()
        items = queue.list_items(status=status, queue_type=queue_type, limit=limit)
        dead_letters = queue.list_dead_letters(queue_type=queue_type, limit=min(200, max(20, int(limit))))
        snapshot = queue.snapshot()
        return {
            "items": [
                {
                    "id": str(item.get("id") or ""),
                    "queue_type": str(item.get("queue_type") or ""),
                    "action": str(item.get("action") or ""),
                    "run_id": str(item.get("run_id") or ""),
                    "attempt": self._to_int(item.get("attempt")),
                    "max_attempts": self._to_int(item.get("max_attempts")),
                    "status": str(item.get("status") or ""),
                    "created_at": str(item.get("created_at") or ""),
                    "updated_at": str(item.get("updated_at") or ""),
                    "next_run_at": str(item.get("next_run_at") or ""),
                    "last_error": str(item.get("last_error") or ""),
                    "last_error_code": str(item.get("last_error_code") or ""),
                    "last_result": str(item.get("last_result") or ""),
                    "last_duration_ms": self._to_int(item.get("last_duration_ms")),
                    "last_trace_id": str(item.get("last_trace_id") or ""),
                    "dedupe_key": str(item.get("dedupe_key") or ""),
                    "idempotency_key": str(item.get("idempotency_key") or ""),
                }
                for item in items
            ],
            "dead_letters": [
                {
                    "id": str(item.get("id") or ""),
                    "queue_type": str(item.get("queue_type") or ""),
                    "action": str(item.get("action") or ""),
                    "run_id": str(item.get("run_id") or ""),
                    "attempt": self._to_int(item.get("attempt")),
                    "max_attempts": self._to_int(item.get("max_attempts")),
                    "dead_lettered_at": str(item.get("dead_lettered_at") or ""),
                    "reason": str(item.get("reason") or ""),
                    "error_code": str(item.get("error_code") or ""),
                    "dedupe_key": str(item.get("dedupe_key") or ""),
                    "idempotency_key": str(item.get("idempotency_key") or ""),
                }
                for item in dead_letters
            ],
            "summary": snapshot,
        }

    def retry_queue_requeue(self, item_id: str) -> dict[str, Any]:
        queue = self._load_retry_queue()
        item = queue.requeue(item_id)
        if item is None:
            raise ValueError("requeue failed: item missing or running")
        return {"ok": True, "item": item, "summary": queue.snapshot()}

    def retry_queue_drop(self, item_id: str) -> dict[str, Any]:
        queue = self._load_retry_queue()
        item = queue.drop(item_id, reason="dropped_by_user")
        if item is None:
            raise ValueError("drop failed: item missing")
        return {"ok": True, "item": item, "summary": queue.snapshot()}

    def retry_queue_kick(
        self,
        *,
        queue_type: str = "all",
        limit: int = 120,
    ) -> dict[str, Any]:
        queue = self._load_retry_queue()
        items = queue.list_items(status="pending", queue_type=queue_type, limit=limit)
        kicked = 0
        for item in items:
            item_id = str(item.get("id") or "").strip()
            if not item_id:
                continue
            if queue.kick(item_id) is not None:
                kicked += 1
        return {"ok": True, "kicked": kicked, "summary": queue.snapshot()}

    def load_resume_view(self) -> dict[str, Any]:
        loader = self._get_resume_loader()
        source = loader.source_path
        resume_text_path = loader.resume_text_path
        resume_text = loader.load_source_text()
        if not resume_text:
            resume_text = loader.load_resume_text()
        preview = resume_text[:400]
        return {
            "source_txt_path": str(source),
            "resume_text_path": str(resume_text_path),
            "source_exists": source.exists(),
            "resume_text_exists": resume_text_path.exists(),
            "resume_chars": len(resume_text),
            "resume_text": resume_text,
            "resume_preview": preview,
        }

    def save_resume_text(self, text: str) -> dict[str, Any]:
        loader = self._get_resume_loader()
        normalized = loader.save_resume_text(str(text or ""))
        return {
            "ok": True,
            "message": "resume text saved",
            "resume_chars": len(normalized),
            "resume_text_path": str(loader.resume_text_path),
        }

    def upload_resume_file(self, *, filename: str, content: bytes, mime_type: str = "") -> dict[str, Any]:
        loader = self._get_resume_loader()
        normalized = loader.update_from_upload_bytes(filename=filename, content=content, mime_type=mime_type)
        return {
            "ok": True,
            "message": "resume uploaded",
            "filename": filename,
            "mime_type": mime_type,
            "resume_chars": len(normalized),
            "resume_text_path": str(loader.resume_text_path),
        }

    def parse_resume_file(self, *, filename: str, content: bytes, mime_type: str = "") -> dict[str, Any]:
        loader = self._get_resume_loader()
        parsed_text = loader.parse_upload_bytes(filename=filename, content=content, mime_type=mime_type)
        return {
            "ok": True,
            "message": "resume file parsed",
            "filename": filename,
            "mime_type": mime_type,
            "resume_chars": len(parsed_text),
            "resume_text": parsed_text,
        }

    def upload_resume_base64(self, *, filename: str, content_base64: str, mime_type: str = "") -> dict[str, Any]:
        raw = str(content_base64 or "").strip()
        if not raw:
            raise ValueError("content_base64 is required")
        try:
            content = base64.b64decode(raw)
        except Exception as exc:
            raise ValueError(f"invalid base64 content: {exc}") from exc
        return self.upload_resume_file(filename=filename, content=content, mime_type=mime_type)

    def parse_resume_base64(self, *, filename: str, content_base64: str, mime_type: str = "") -> dict[str, Any]:
        raw = str(content_base64 or "").strip()
        if not raw:
            raise ValueError("content_base64 is required")
        try:
            content = base64.b64decode(raw)
        except Exception as exc:
            raise ValueError(f"invalid base64 content: {exc}") from exc
        return self.parse_resume_file(filename=filename, content=content, mime_type=mime_type)

    def load_xhs_accounts_view(self) -> dict[str, Any]:
        config = self._read_config_data()
        xhs = self._ensure_section(config, "xhs")
        selected = str(xhs.get("account") or "default").strip() or "default"
        account_cookies_dir = str(xhs.get("account_cookies_dir") or "~/.xhs-mcp/accounts").strip() or "~/.xhs-mcp/accounts"
        base = Path(account_cookies_dir).expanduser()

        options: dict[str, dict[str, Any]] = {
            "default": {
                "value": "default",
                "label": "default",
                "has_cookie": (Path.home() / ".xhs-mcp" / "cookies.json").exists(),
                "path": str((Path.home() / ".xhs-mcp" / "cookies.json")),
            }
        }
        if base.exists() and base.is_dir():
            for file in base.glob("*.json"):
                name = file.stem.strip()
                if not name:
                    continue
                options[name] = {
                    "value": name,
                    "label": name,
                    "has_cookie": file.exists(),
                    "path": str(file),
                }
            for folder in base.iterdir():
                if not folder.is_dir():
                    continue
                name = folder.name.strip()
                if not name:
                    continue
                nested = folder / "cookies.json"
                options[name] = {
                    "value": name,
                    "label": name,
                    "has_cookie": nested.exists(),
                    "path": str(nested),
                }

        if selected not in options:
            fallback_file = base / f"{selected}.json"
            options[selected] = {
                "value": selected,
                "label": selected,
                "has_cookie": fallback_file.exists(),
                "path": str(fallback_file),
            }

        ordered = [options["default"]]
        for key in sorted(k for k in options.keys() if k != "default"):
            ordered.append(options[key])
        return {
            "selected": selected,
            "account_cookies_dir": account_cookies_dir,
            "options": ordered,
        }

    def load_config_view(self) -> dict[str, Any]:
        config = self._read_config_data()
        app = self._ensure_section(config, "app")
        xhs = self._ensure_section(config, "xhs")
        pipeline = self._ensure_section(config, "pipeline")
        agent = self._ensure_section(config, "agent")
        notification = self._ensure_section(config, "notification")
        email = self._ensure_section(config, "email")
        wechat = self._ensure_section(config, "wechat_service")
        llm = self._ensure_section(config, "llm")
        resume = self._ensure_section(config, "resume")
        observability = self._ensure_section(config, "observability")
        alerts = self._ensure_section(observability, "alerts")
        fetch_alert = self._ensure_section(alerts, "fetch_fail_streak")
        llm_alert = self._ensure_section(alerts, "llm_timeout_rate")
        detail_alert = self._ensure_section(alerts, "detail_missing_rate")
        xhs_account = self.load_xhs_accounts_view()

        fetch_short_threshold = self._coerce_int(
            fetch_alert.get("short_threshold"),
            self._coerce_int(alerts.get("fetch_fail_streak_threshold"), 2, minimum=1),
            minimum=1,
        )
        llm_short_threshold = self._coerce_rate(
            llm_alert.get("short_threshold"),
            self._coerce_rate(alerts.get("llm_timeout_rate_threshold"), 0.35),
        )
        llm_short_min_samples = self._coerce_int(
            llm_alert.get("short_min_samples"),
            self._coerce_int(alerts.get("llm_timeout_min_calls"), 6, minimum=1),
            minimum=1,
        )
        detail_short_threshold = self._coerce_rate(
            detail_alert.get("short_threshold"),
            self._coerce_rate(alerts.get("detail_missing_rate_threshold"), 0.45),
        )
        detail_short_min_samples = self._coerce_int(
            detail_alert.get("short_min_samples"),
            self._coerce_int(alerts.get("detail_missing_min_samples"), 6, minimum=1),
            minimum=1,
        )

        return {
            "app": {"interval_minutes": self._coerce_int(app.get("interval_minutes"), 15, minimum=1)},
            "xhs": {
                "keyword": str(xhs.get("keyword") or "继任"),
                "search_sort": str(xhs.get("search_sort") or "time_descending"),
                "max_results": self._coerce_int(xhs.get("max_results"), 20, minimum=1),
                "max_detail_fetch": self._coerce_int(xhs.get("max_detail_fetch"), 5, minimum=1),
                "detail_workers": self._coerce_int(xhs.get("detail_workers"), 3, minimum=1),
                "account": xhs_account["selected"],
                "account_cookies_dir": xhs_account["account_cookies_dir"],
                "account_options": xhs_account["options"],
            },
            "pipeline": {
                "process_workers": self._coerce_int(pipeline.get("process_workers"), 4, minimum=1),
            },
            "agent": {"mode": self._normalize_mode(str(agent.get("mode") or "auto"))},
            "notification": {
                "mode": self._normalize_notify_mode(str(notification.get("mode") or "digest")),
                "digest_interval_minutes": self._coerce_int(notification.get("digest_interval_minutes"), 30, minimum=1),
                "digest_top_summaries": self._coerce_int(notification.get("digest_top_summaries"), 5, minimum=1),
                "digest_send_when_no_new": bool(notification.get("digest_send_when_no_new", False)),
                "digest_channels": self._to_name_list(notification.get("digest_channels"), default=["email"]),
                "realtime_channels": self._to_name_list(notification.get("realtime_channels"), default=["wechat_service", "email"]),
            },
            "email": {"enabled": bool(email.get("enabled", False))},
            "wechat_service": {"enabled": bool(wechat.get("enabled", False))},
            "llm": {
                "enabled": bool(llm.get("enabled", False)),
                "model": str(llm.get("model") or ""),
                "base_url": str(llm.get("base_url") or ""),
            },
            "observability": {
                "alerts": {
                    "enabled": bool(alerts.get("enabled", True)),
                    "cooldown_minutes": self._coerce_int(alerts.get("cooldown_minutes"), 60, minimum=1),
                    # Legacy fields are kept for existing UI controls.
                    "fetch_fail_streak_threshold": fetch_short_threshold,
                    "llm_timeout_rate_threshold": llm_short_threshold,
                    "llm_timeout_min_calls": llm_short_min_samples,
                    "detail_missing_rate_threshold": detail_short_threshold,
                    "detail_missing_min_samples": detail_short_min_samples,
                    # Dual-window fields are the runtime source of truth.
                    "fetch_fail_streak": {
                        "short_window_runs": self._coerce_int(fetch_alert.get("short_window_runs"), 1, minimum=1),
                        "short_threshold": fetch_short_threshold,
                        "short_min_runs": self._coerce_int(fetch_alert.get("short_min_runs"), 1, minimum=1),
                        "long_window_runs": self._coerce_int(fetch_alert.get("long_window_runs"), 6, minimum=1),
                        "long_threshold": self._coerce_float(
                            fetch_alert.get("long_threshold"),
                            max(1.0, float(fetch_short_threshold) * 0.6),
                            minimum=1.0,
                        ),
                        "long_min_runs": self._coerce_int(fetch_alert.get("long_min_runs"), 3, minimum=1),
                    },
                    "llm_timeout_rate": {
                        "short_window_runs": self._coerce_int(llm_alert.get("short_window_runs"), 1, minimum=1),
                        "short_threshold": llm_short_threshold,
                        "short_min_samples": llm_short_min_samples,
                        "long_window_runs": self._coerce_int(llm_alert.get("long_window_runs"), 8, minimum=1),
                        "long_threshold": self._coerce_rate(llm_alert.get("long_threshold"), llm_short_threshold * 0.7),
                        "long_min_samples": self._coerce_int(llm_alert.get("long_min_samples"), max(12, llm_short_min_samples * 3), minimum=1),
                    },
                    "detail_missing_rate": {
                        "short_window_runs": self._coerce_int(detail_alert.get("short_window_runs"), 1, minimum=1),
                        "short_threshold": detail_short_threshold,
                        "short_min_samples": detail_short_min_samples,
                        "long_window_runs": self._coerce_int(detail_alert.get("long_window_runs"), 8, minimum=1),
                        "long_threshold": self._coerce_rate(detail_alert.get("long_threshold"), detail_short_threshold * 0.7),
                        "long_min_samples": self._coerce_int(detail_alert.get("long_min_samples"), max(12, detail_short_min_samples * 3), minimum=1),
                    },
                    "channels": self._to_name_list(alerts.get("channels"), default=[]),
                }
            },
            "resume": {
                "source_txt_path": str(resume.get("source_txt_path") or "config/resume.txt"),
                "resume_text_path": str(resume.get("resume_text_path") or "data/resume_text.txt"),
                "max_chars": self._coerce_int(resume.get("max_chars"), 6000, minimum=500),
            },
        }

    def save_config_view(self, payload: dict[str, Any]) -> dict[str, Any]:
        if not isinstance(payload, dict):
            raise ValueError("config payload must be an object")

        config = self._read_config_data()
        app = self._ensure_section(config, "app")
        xhs = self._ensure_section(config, "xhs")
        pipeline = self._ensure_section(config, "pipeline")
        agent = self._ensure_section(config, "agent")
        notification = self._ensure_section(config, "notification")
        email = self._ensure_section(config, "email")
        wechat = self._ensure_section(config, "wechat_service")
        llm = self._ensure_section(config, "llm")
        resume = self._ensure_section(config, "resume")
        observability = self._ensure_section(config, "observability")
        alerts = self._ensure_section(observability, "alerts")

        app_in = payload.get("app")
        if isinstance(app_in, dict):
            app["interval_minutes"] = self._coerce_int(app_in.get("interval_minutes"), app.get("interval_minutes", 15), minimum=1)

        xhs_in = payload.get("xhs")
        if isinstance(xhs_in, dict):
            keyword = str(xhs_in.get("keyword") or xhs.get("keyword") or "继任").strip()
            xhs["keyword"] = keyword or "继任"
            xhs["search_sort"] = str(xhs_in.get("search_sort") or xhs.get("search_sort") or "time_descending").strip() or "time_descending"
            xhs["max_results"] = self._coerce_int(xhs_in.get("max_results"), xhs.get("max_results", 20), minimum=1)
            xhs["max_detail_fetch"] = self._coerce_int(xhs_in.get("max_detail_fetch"), xhs.get("max_detail_fetch", 5), minimum=1)
            xhs["detail_workers"] = self._coerce_int(xhs_in.get("detail_workers"), xhs.get("detail_workers", 3), minimum=1)
            xhs["account"] = str(xhs_in.get("account") or xhs.get("account") or "default").strip() or "default"
            xhs["account_cookies_dir"] = (
                str(xhs_in.get("account_cookies_dir") or xhs.get("account_cookies_dir") or "~/.xhs-mcp/accounts").strip()
                or "~/.xhs-mcp/accounts"
            )

        pipeline_in = payload.get("pipeline")
        if isinstance(pipeline_in, dict):
            pipeline["process_workers"] = self._coerce_int(
                pipeline_in.get("process_workers"),
                pipeline.get("process_workers", 4),
                minimum=1,
            )

        agent_in = payload.get("agent")
        if isinstance(agent_in, dict):
            agent["mode"] = self._normalize_mode(str(agent_in.get("mode") or agent.get("mode") or "auto"))

        notify_in = payload.get("notification")
        if isinstance(notify_in, dict):
            notification["mode"] = self._normalize_notify_mode(str(notify_in.get("mode") or notification.get("mode") or "digest"))
            notification["digest_interval_minutes"] = self._coerce_int(
                notify_in.get("digest_interval_minutes"),
                notification.get("digest_interval_minutes", 30),
                minimum=1,
            )
            notification["digest_top_summaries"] = self._coerce_int(
                notify_in.get("digest_top_summaries"),
                notification.get("digest_top_summaries", 5),
                minimum=1,
            )
            notification["digest_send_when_no_new"] = bool(notify_in.get("digest_send_when_no_new", False))
            if "digest_channels" in notify_in:
                notification["digest_channels"] = self._to_name_list(notify_in.get("digest_channels"), default=["email"])
            if "realtime_channels" in notify_in:
                notification["realtime_channels"] = self._to_name_list(
                    notify_in.get("realtime_channels"),
                    default=["wechat_service", "email"],
                )

        email_in = payload.get("email")
        if isinstance(email_in, dict) and "enabled" in email_in:
            email["enabled"] = bool(email_in.get("enabled"))

        wechat_in = payload.get("wechat_service")
        if isinstance(wechat_in, dict) and "enabled" in wechat_in:
            wechat["enabled"] = bool(wechat_in.get("enabled"))

        llm_in = payload.get("llm")
        if isinstance(llm_in, dict):
            if "enabled" in llm_in:
                llm["enabled"] = bool(llm_in.get("enabled"))
            if "model" in llm_in:
                llm["model"] = str(llm_in.get("model") or "").strip()
            if "base_url" in llm_in:
                llm["base_url"] = str(llm_in.get("base_url") or "").strip()

        observability_in = payload.get("observability")
        if isinstance(observability_in, dict):
            alerts_in = observability_in.get("alerts")
            if isinstance(alerts_in, dict):
                fetch_alert = self._ensure_section(alerts, "fetch_fail_streak")
                llm_alert = self._ensure_section(alerts, "llm_timeout_rate")
                detail_alert = self._ensure_section(alerts, "detail_missing_rate")
                if "enabled" in alerts_in:
                    alerts["enabled"] = bool(alerts_in.get("enabled"))
                if "cooldown_minutes" in alerts_in:
                    alerts["cooldown_minutes"] = self._coerce_int(
                        alerts_in.get("cooldown_minutes"),
                        alerts.get("cooldown_minutes", 60),
                        minimum=1,
                    )
                if "fetch_fail_streak_threshold" in alerts_in:
                    legacy_fetch_threshold = self._coerce_int(
                        alerts_in.get("fetch_fail_streak_threshold"),
                        alerts.get("fetch_fail_streak_threshold", 2),
                        minimum=1,
                    )
                    alerts["fetch_fail_streak_threshold"] = legacy_fetch_threshold
                    fetch_alert["short_threshold"] = legacy_fetch_threshold
                if "llm_timeout_rate_threshold" in alerts_in:
                    legacy_llm_threshold = self._coerce_rate(
                        alerts_in.get("llm_timeout_rate_threshold"),
                        alerts.get("llm_timeout_rate_threshold", 0.35),
                    )
                    alerts["llm_timeout_rate_threshold"] = legacy_llm_threshold
                    llm_alert["short_threshold"] = legacy_llm_threshold
                if "llm_timeout_min_calls" in alerts_in:
                    legacy_llm_min_calls = self._coerce_int(
                        alerts_in.get("llm_timeout_min_calls"),
                        alerts.get("llm_timeout_min_calls", 6),
                        minimum=1,
                    )
                    alerts["llm_timeout_min_calls"] = legacy_llm_min_calls
                    llm_alert["short_min_samples"] = legacy_llm_min_calls
                if "detail_missing_rate_threshold" in alerts_in:
                    legacy_detail_threshold = self._coerce_rate(
                        alerts_in.get("detail_missing_rate_threshold"),
                        alerts.get("detail_missing_rate_threshold", 0.45),
                    )
                    alerts["detail_missing_rate_threshold"] = legacy_detail_threshold
                    detail_alert["short_threshold"] = legacy_detail_threshold
                if "detail_missing_min_samples" in alerts_in:
                    legacy_detail_min_samples = self._coerce_int(
                        alerts_in.get("detail_missing_min_samples"),
                        alerts.get("detail_missing_min_samples", 6),
                        minimum=1,
                    )
                    alerts["detail_missing_min_samples"] = legacy_detail_min_samples
                    detail_alert["short_min_samples"] = legacy_detail_min_samples
                if "channels" in alerts_in:
                    alerts["channels"] = self._to_name_list(alerts_in.get("channels"), default=[])

                fetch_in = alerts_in.get("fetch_fail_streak")
                if isinstance(fetch_in, dict):
                    if "short_window_runs" in fetch_in:
                        fetch_alert["short_window_runs"] = self._coerce_int(
                            fetch_in.get("short_window_runs"),
                            fetch_alert.get("short_window_runs", 1),
                            minimum=1,
                        )
                    if "short_threshold" in fetch_in:
                        short_threshold = self._coerce_int(
                            fetch_in.get("short_threshold"),
                            fetch_alert.get("short_threshold", alerts.get("fetch_fail_streak_threshold", 2)),
                            minimum=1,
                        )
                        fetch_alert["short_threshold"] = short_threshold
                        alerts["fetch_fail_streak_threshold"] = short_threshold
                    if "short_min_runs" in fetch_in:
                        fetch_alert["short_min_runs"] = self._coerce_int(
                            fetch_in.get("short_min_runs"),
                            fetch_alert.get("short_min_runs", 1),
                            minimum=1,
                        )
                    if "long_window_runs" in fetch_in:
                        fetch_alert["long_window_runs"] = self._coerce_int(
                            fetch_in.get("long_window_runs"),
                            fetch_alert.get("long_window_runs", 6),
                            minimum=1,
                        )
                    if "long_threshold" in fetch_in:
                        fetch_alert["long_threshold"] = self._coerce_float(
                            fetch_in.get("long_threshold"),
                            fetch_alert.get("long_threshold", 1.2),
                            minimum=1.0,
                        )
                    if "long_min_runs" in fetch_in:
                        fetch_alert["long_min_runs"] = self._coerce_int(
                            fetch_in.get("long_min_runs"),
                            fetch_alert.get("long_min_runs", 3),
                            minimum=1,
                        )

                llm_in = alerts_in.get("llm_timeout_rate")
                if isinstance(llm_in, dict):
                    if "short_window_runs" in llm_in:
                        llm_alert["short_window_runs"] = self._coerce_int(
                            llm_in.get("short_window_runs"),
                            llm_alert.get("short_window_runs", 1),
                            minimum=1,
                        )
                    if "short_threshold" in llm_in:
                        short_threshold = self._coerce_rate(
                            llm_in.get("short_threshold"),
                            llm_alert.get("short_threshold", alerts.get("llm_timeout_rate_threshold", 0.35)),
                        )
                        llm_alert["short_threshold"] = short_threshold
                        alerts["llm_timeout_rate_threshold"] = short_threshold
                    if "short_min_samples" in llm_in:
                        short_min_samples = self._coerce_int(
                            llm_in.get("short_min_samples"),
                            llm_alert.get("short_min_samples", alerts.get("llm_timeout_min_calls", 6)),
                            minimum=1,
                        )
                        llm_alert["short_min_samples"] = short_min_samples
                        alerts["llm_timeout_min_calls"] = short_min_samples
                    if "long_window_runs" in llm_in:
                        llm_alert["long_window_runs"] = self._coerce_int(
                            llm_in.get("long_window_runs"),
                            llm_alert.get("long_window_runs", 8),
                            minimum=1,
                        )
                    if "long_threshold" in llm_in:
                        llm_alert["long_threshold"] = self._coerce_rate(
                            llm_in.get("long_threshold"),
                            llm_alert.get("long_threshold", 0.25),
                        )
                    if "long_min_samples" in llm_in:
                        llm_alert["long_min_samples"] = self._coerce_int(
                            llm_in.get("long_min_samples"),
                            llm_alert.get("long_min_samples", 18),
                            minimum=1,
                        )

                detail_in = alerts_in.get("detail_missing_rate")
                if isinstance(detail_in, dict):
                    if "short_window_runs" in detail_in:
                        detail_alert["short_window_runs"] = self._coerce_int(
                            detail_in.get("short_window_runs"),
                            detail_alert.get("short_window_runs", 1),
                            minimum=1,
                        )
                    if "short_threshold" in detail_in:
                        short_threshold = self._coerce_rate(
                            detail_in.get("short_threshold"),
                            detail_alert.get("short_threshold", alerts.get("detail_missing_rate_threshold", 0.45)),
                        )
                        detail_alert["short_threshold"] = short_threshold
                        alerts["detail_missing_rate_threshold"] = short_threshold
                    if "short_min_samples" in detail_in:
                        short_min_samples = self._coerce_int(
                            detail_in.get("short_min_samples"),
                            detail_alert.get("short_min_samples", alerts.get("detail_missing_min_samples", 6)),
                            minimum=1,
                        )
                        detail_alert["short_min_samples"] = short_min_samples
                        alerts["detail_missing_min_samples"] = short_min_samples
                    if "long_window_runs" in detail_in:
                        detail_alert["long_window_runs"] = self._coerce_int(
                            detail_in.get("long_window_runs"),
                            detail_alert.get("long_window_runs", 8),
                            minimum=1,
                        )
                    if "long_threshold" in detail_in:
                        detail_alert["long_threshold"] = self._coerce_rate(
                            detail_in.get("long_threshold"),
                            detail_alert.get("long_threshold", 0.3),
                        )
                    if "long_min_samples" in detail_in:
                        detail_alert["long_min_samples"] = self._coerce_int(
                            detail_in.get("long_min_samples"),
                            detail_alert.get("long_min_samples", 18),
                            minimum=1,
                        )

        resume_in = payload.get("resume")
        if isinstance(resume_in, dict):
            if "source_txt_path" in resume_in:
                resume["source_txt_path"] = str(resume_in.get("source_txt_path") or "config/resume.txt").strip() or "config/resume.txt"
            if "resume_text_path" in resume_in:
                resume["resume_text_path"] = str(resume_in.get("resume_text_path") or "data/resume_text.txt").strip() or "data/resume_text.txt"
            if "max_chars" in resume_in:
                resume["max_chars"] = self._coerce_int(resume_in.get("max_chars"), 6000, minimum=500)

        self._write_config_data(config)
        self._resume_loader = None  # Invalidate cached resume loader after config updates
        return self.load_config_view()

    def run_action(self, action: str, payload: dict[str, Any]) -> dict[str, Any]:
        value = str(action or "").strip().lower()
        if value == "run_once":
            mode = self._normalize_mode(str(payload.get("mode") or "auto"))
            return self.runtime.start_run_once(mode=mode)
        if value == "send_latest":
            limit = self._coerce_int(payload.get("limit"), 5, minimum=1)
            return self.runtime.start_send_latest(limit=limit)
        if value == "xhs_login":
            timeout_seconds = self._coerce_int(payload.get("timeout_seconds"), 180, minimum=30)
            return self.runtime.start_xhs_login(timeout_seconds=timeout_seconds)
        if value == "xhs_status":
            return self.runtime.check_xhs_status()
        if value == "start_daemon":
            mode = self._normalize_mode(str(payload.get("mode") or "auto"))
            interval = self._coerce_int(payload.get("interval_minutes"), 0, minimum=1)
            interval_arg = interval if interval > 0 else None
            return self.runtime.start_daemon(mode=mode, interval_minutes=interval_arg)
        if value == "stop_daemon":
            return self.runtime.stop_daemon()
        if value == "stop_job":
            return self.runtime.stop_job()
        raise ValueError(f"unsupported action: {action}")

    def run_setup_check(self, *, include_network: bool = True, include_xhs_status: bool = True) -> dict[str, Any]:
        checked_at = datetime.now().isoformat(timespec="seconds")
        items: list[dict[str, Any]] = []

        def push_item(
            key: str,
            name: str,
            status: str,
            message: str,
            detail: str = "",
            suggestion: str = "",
            reason: str = "",
            fix_command: str = "",
        ) -> None:
            resolved_reason = str(reason or message or "").strip()
            resolved_fix = str(fix_command or suggestion or "").strip()
            items.append(
                {
                    "key": key,
                    "name": name,
                    "status": status,
                    "message": message,
                    "reason": resolved_reason,
                    "detail": detail,
                    "suggestion": suggestion,
                    "fix_command": resolved_fix,
                }
            )

        config_data: dict[str, Any] = {}
        config_ok = False
        try:
            if not self.config_path.exists():
                raise FileNotFoundError(f"配置文件不存在: {self.config_path}")
            config_data = self._read_config_data()
            if not isinstance(config_data, dict):
                raise ValueError("配置文件内容不是对象")
            config_ok = True
            push_item(
                key="config_file",
                name="配置文件",
                status="pass",
                message="config/config.yaml 可读取",
                detail=str(self.config_path),
            )
        except Exception as exc:
            push_item(
                key="config_file",
                name="配置文件",
                status="fail",
                message=f"配置文件异常: {exc}",
                detail=str(self.config_path),
                suggestion="请先执行 scripts/bootstrap.ps1 生成并检查 config/config.yaml",
            )

        try:
            self.data_dir.mkdir(parents=True, exist_ok=True)
            probe = self.data_dir / ".setup_probe_write.tmp"
            probe.write_text("ok", encoding="utf-8")
            try:
                probe.unlink()
            except Exception:
                pass
            push_item(
                key="storage_write",
                name="本地写入权限",
                status="pass",
                message="data 目录可写",
                detail=str(self.data_dir),
            )
        except Exception as exc:
            push_item(
                key="storage_write",
                name="本地写入权限",
                status="fail",
                message=f"写入失败: {exc}",
                detail=str(self.data_dir),
                suggestion="检查 data 目录权限或关闭占用该目录/文件的程序",
            )

        settings = None
        if config_ok:
            try:
                settings = load_settings(str(self.config_path))
            except Exception as exc:
                push_item(
                    key="config_parse",
                    name="配置解析",
                    status="fail",
                    message=f"配置解析失败: {exc}",
                    suggestion="检查 config/config.yaml 字段类型和缩进格式",
                )

        env_map = self._read_env_values()

        xhs_command = "node"
        xhs_script = self.workspace / "vendor" / "xhs-mcp" / "dist" / "xhs-mcp.js"
        if settings is not None:
            xhs_command = str(settings.xhs.command or "node").strip() or "node"
            xhs_script = self._resolve_xhs_script_path(settings.xhs.args)
        else:
            xhs_cfg = config_data.get("xhs", {}) if isinstance(config_data, dict) else {}
            if isinstance(xhs_cfg, dict):
                xhs_command = str(xhs_cfg.get("command") or "node").strip() or "node"
                xhs_script = self._resolve_xhs_script_path(xhs_cfg.get("args"))

        cmd_path = shutil.which(xhs_command)
        xhs_script_exists = xhs_script.exists()
        account = "default"
        account_dir = "~/.xhs-mcp/accounts"
        if settings is not None:
            account = str(getattr(settings.xhs, "account", "default") or "default").strip() or "default"
            account_dir = str(getattr(settings.xhs, "account_cookies_dir", "~/.xhs-mcp/accounts") or "~/.xhs-mcp/accounts").strip() or "~/.xhs-mcp/accounts"
        elif isinstance(config_data.get("xhs"), dict):
            xhs_cfg = config_data.get("xhs") or {}
            account = str(xhs_cfg.get("account") or "default").strip() or "default"
            account_dir = str(xhs_cfg.get("account_cookies_dir") or "~/.xhs-mcp/accounts").strip() or "~/.xhs-mcp/accounts"
        selected_cookie_file = self._resolve_selected_xhs_cookie_file(account=account, account_cookies_dir=account_dir)
        account_args = f'-Account "{account}" -AccountCookiesDir "{account_dir}"'

        if cmd_path:
            push_item(
                key="xhs_cli_present",
                name="XHS CLI 可用性",
                status="pass",
                message=f"命令可用: {xhs_command}",
                detail=f"path={cmd_path}",
            )
        else:
            push_item(
                key="xhs_cli_present",
                name="XHS CLI 可用性",
                status="fail",
                message=f"命令不存在: {xhs_command}",
                suggestion="安装 Node.js 并确保 node 命令在 PATH 中",
                fix_command="powershell -ExecutionPolicy Bypass -File scripts/bootstrap.ps1",
            )

        if cmd_path and xhs_script_exists:
            push_item(
                key="xhs_runtime",
                name="XHS 运行依赖",
                status="pass",
                message="xhs-mcp 依赖就绪",
                detail=f"command={xhs_command} | script={xhs_script}",
            )
        else:
            missing_parts = []
            if not cmd_path:
                missing_parts.append(f"命令不存在: {xhs_command}")
            if not xhs_script_exists:
                missing_parts.append(f"脚本不存在: {xhs_script}")
            push_item(
                key="xhs_runtime",
                name="XHS 运行依赖",
                status="fail",
                message="；".join(missing_parts) or "xhs 运行依赖缺失",
                detail=f"command={xhs_command} | script={xhs_script}",
                suggestion="确认 Node.js 已安装，且 vendor/xhs-mcp 已完整放入项目目录",
                fix_command="cd vendor/xhs-mcp; npm install --no-fund --no-audit --cache .npm-cache",
            )

        if selected_cookie_file.exists():
            push_item(
                key="xhs_cookie_file_ready",
                name="XHS Cookie 文件",
                status="pass",
                message="当前账号 Cookie 文件已存在",
                detail=str(selected_cookie_file),
            )
        else:
            push_item(
                key="xhs_cookie_file_ready",
                name="XHS Cookie 文件",
                status="warn",
                message="当前账号 Cookie 文件不存在（首次扫码后会自动生成）",
                detail=str(selected_cookie_file),
                suggestion="在控制中心执行一次扫码登录",
                fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_login.ps1 -Timeout 180 {account_args}',
            )

        if include_xhs_status:
            status_resp = self.runtime.check_xhs_status()
            status_obj = status_resp.get("status") if isinstance(status_resp, dict) else {}
            if not isinstance(status_obj, dict):
                status_obj = {}
            logged_in = bool(
                status_obj.get("loggedIn")
                or status_obj.get("logged_in")
                or status_obj.get("isLogin")
                or status_obj.get("is_login")
            )
            if status_resp.get("ok"):
                push_item(
                    key="xhs_mcp_connect",
                    name="XHS MCP 连通性",
                    status="pass",
                    message="状态接口可访问",
                    detail=self._compact_json(status_obj),
                )
            else:
                push_item(
                    key="xhs_mcp_connect",
                    name="XHS MCP 连通性",
                    status="fail",
                    message=str(status_resp.get("message") or "状态检查失败"),
                    detail=str(status_resp.get("output") or "")[:280],
                    suggestion="检查 xhs-mcp 依赖和浏览器路径后重试",
                    fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_status.ps1 {account_args}',
                )

            if status_resp.get("ok") and logged_in:
                push_item(
                    key="xhs_login_status",
                    name="XHS 登录状态",
                    status="pass",
                    message="登录状态正常",
                    detail=self._compact_json(status_obj),
                )
                push_item(
                    key="xhs_login",
                    name="XHS 登录状态",
                    status="pass",
                    message="登录状态正常",
                    detail=self._compact_json(status_obj),
                )
            elif status_resp.get("ok"):
                push_item(
                    key="xhs_login_status",
                    name="XHS 登录状态",
                    status="fail",
                    message="未登录，请扫码登录",
                    detail=self._compact_json(status_obj),
                    suggestion="在控制中心点击“扫码登录”，完成后再次自检",
                    fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_login.ps1 -Timeout 180 {account_args}',
                )
                push_item(
                    key="xhs_login",
                    name="XHS 登录状态",
                    status="fail",
                    message="未登录，请扫码登录",
                    detail=self._compact_json(status_obj),
                    suggestion="在控制中心点击“扫码登录”，完成后再次自检",
                    fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_login.ps1 -Timeout 180 {account_args}',
                )
            else:
                push_item(
                    key="xhs_login_status",
                    name="XHS 登录状态",
                    status="fail",
                    message=str(status_resp.get("message") or "状态检查失败"),
                    detail=str(status_resp.get("output") or "")[:280],
                    suggestion="先执行扫码登录后重试",
                    fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_login.ps1 -Timeout 180 {account_args}',
                )
                push_item(
                    key="xhs_login",
                    name="XHS 登录状态",
                    status="fail",
                    message=str(status_resp.get("message") or "状态检查失败"),
                    detail=str(status_resp.get("output") or "")[:280],
                    suggestion="先执行 scripts/xhs_login.ps1 完成扫码后重试",
                    fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_login.ps1 -Timeout 180 {account_args}',
                )
        else:
            push_item(
                key="xhs_mcp_connect",
                name="XHS MCP 连通性",
                status="warn",
                message="已跳过连通性检查",
                suggestion="建议首次部署后执行一次状态检查",
                fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_status.ps1 {account_args}',
            )
            push_item(
                key="xhs_login_status",
                name="XHS 登录状态",
                status="warn",
                message="已跳过登录检查",
                suggestion="建议首次部署完成后执行一次登录检查",
                fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_status.ps1 {account_args}',
            )
            push_item(
                key="xhs_login",
                name="XHS 登录状态",
                status="warn",
                message="已跳过登录检查",
                suggestion="建议首次部署完成后执行一次登录检查",
                fix_command=f'powershell -ExecutionPolicy Bypass -File scripts/xhs_status.ps1 {account_args}',
            )

        email_enabled = False
        email_host = "smtp.126.com"
        email_port = 465
        email_use_ssl = True
        username_env = "EMAIL_SMTP_USERNAME"
        password_env = "EMAIL_SMTP_PASSWORD"
        from_env = "EMAIL_FROM"
        to_env = "EMAIL_TO"
        if settings is not None:
            email_enabled = bool(settings.email.enabled)
            email_host = str(settings.email.smtp_host or email_host)
            email_port = int(settings.email.smtp_port or email_port)
            email_use_ssl = bool(settings.email.use_ssl)
            username_env = str(settings.email.username_env or username_env)
            password_env = str(settings.email.password_env or password_env)
            from_env = str(settings.email.from_env or from_env)
            to_env = str(settings.email.to_env or to_env)
        elif isinstance(config_data.get("email"), dict):
            email_cfg = config_data.get("email") or {}
            email_enabled = bool(email_cfg.get("enabled", False))
            email_host = str(email_cfg.get("smtp_host") or email_host)
            email_port = self._coerce_int(email_cfg.get("smtp_port"), email_port, minimum=1)
            email_use_ssl = bool(email_cfg.get("use_ssl", True))
            username_env = str(email_cfg.get("username_env") or username_env)
            password_env = str(email_cfg.get("password_env") or password_env)
            from_env = str(email_cfg.get("from_env") or from_env)
            to_env = str(email_cfg.get("to_env") or to_env)

        email_creds = {
            username_env: self._get_env_value(username_env, env_map),
            password_env: self._get_env_value(password_env, env_map),
            from_env: self._get_env_value(from_env, env_map),
            to_env: self._get_env_value(to_env, env_map),
        }
        if not email_enabled:
            push_item(
                key="email_enabled",
                name="邮件通知开关",
                status="warn",
                message="email.enabled=false，邮件通知未启用",
                suggestion="如需邮件摘要，请在控制中心启用邮件并保存",
            )
        else:
            missing_email = [key for key, value in email_creds.items() if not str(value or "").strip()]
            if missing_email:
                push_item(
                    key="email_env",
                    name="邮件环境变量",
                    status="fail",
                    message=f"缺少: {', '.join(missing_email)}",
                    suggestion="请在 .env 中补齐邮件账号、授权码、收件人",
                )
            else:
                push_item(
                    key="email_env",
                    name="邮件环境变量",
                    status="pass",
                    message="邮件变量已配置",
                    detail=f"host={email_host}:{email_port} ssl={email_use_ssl}",
                )

            if include_network:
                if missing_email:
                    push_item(
                        key="email_connect",
                        name="邮件 SMTP 连接",
                        status="fail",
                        message="连接检查未执行（变量缺失）",
                    )
                else:
                    connect_ok, connect_msg = self._check_email_connectivity(
                        host=email_host,
                        port=email_port,
                        use_ssl=email_use_ssl,
                        username=str(email_creds.get(username_env) or ""),
                        password=str(email_creds.get(password_env) or ""),
                    )
                    push_item(
                        key="email_connect",
                        name="邮件 SMTP 连接",
                        status="pass" if connect_ok else "fail",
                        message=connect_msg,
                        suggestion="" if connect_ok else "检查 SMTP 主机/端口、授权码和网络连通性",
                    )
            else:
                push_item(
                    key="email_connect",
                    name="邮件 SMTP 连接",
                    status="warn",
                    message="已跳过网络连通检查",
                )

        llm_enabled = False
        llm_model = ""
        llm_base_url = "https://api.openai.com/v1"
        llm_api_key = ""
        llm_api_key_env = "OPENAI_API_KEY"
        if settings is not None:
            llm_enabled = bool(settings.llm.enabled)
            llm_model = str(settings.llm.model or "")
            llm_base_url = str(settings.llm.base_url or llm_base_url)
            llm_api_key = str(settings.llm_api_key or "")
            llm_api_key_env = str(settings.llm.api_key_env or llm_api_key_env)
        elif isinstance(config_data.get("llm"), dict):
            llm_cfg = config_data.get("llm") or {}
            llm_enabled = bool(llm_cfg.get("enabled", False))
            llm_model = str(llm_cfg.get("model") or "")
            llm_base_url = str(llm_cfg.get("base_url") or llm_base_url)
            llm_api_key_env = str(llm_cfg.get("api_key_env") or llm_api_key_env)
            llm_api_key = str(llm_cfg.get("api_key") or "").strip() or self._get_env_value(llm_api_key_env, env_map)

        if not llm_enabled:
            push_item(
                key="llm_enabled",
                name="LLM 开关",
                status="warn",
                message="llm.enabled=false，LLM 功能未启用",
                suggestion="如需 LLM 过滤/摘要，请开启 LLM 并配置 API Key",
            )
        else:
            llm_missing = []
            if not llm_model.strip():
                llm_missing.append("llm.model")
            if not llm_base_url.strip():
                llm_missing.append("llm.base_url")
            if not llm_api_key.strip():
                llm_missing.append(llm_api_key_env or "OPENAI_API_KEY")

            if llm_missing:
                push_item(
                    key="llm_config",
                    name="LLM 配置",
                    status="fail",
                    message=f"缺少: {', '.join(llm_missing)}",
                    suggestion="请在 config/.env 中补齐模型、地址与 API Key",
                )
            else:
                push_item(
                    key="llm_config",
                    name="LLM 配置",
                    status="pass",
                    message="LLM 基础配置已就绪",
                    detail=f"model={llm_model} | base={llm_base_url}",
                )

            if include_network:
                if llm_missing:
                    push_item(
                        key="llm_connect",
                        name="LLM 连接测试",
                        status="fail",
                        message="连接检查未执行（配置缺失）",
                    )
                else:
                    ok, status, detail = self._check_llm_connectivity(
                        base_url=llm_base_url,
                        api_key=llm_api_key,
                    )
                    if ok:
                        push_item(
                            key="llm_connect",
                            name="LLM 连接测试",
                            status="pass",
                            message=f"连接成功（HTTP {status}）",
                            detail=detail,
                        )
                    else:
                        push_item(
                            key="llm_connect",
                            name="LLM 连接测试",
                            status="fail" if status in {0, 401, 403} else "warn",
                            message=f"连接失败（HTTP {status}）" if status else "连接失败",
                            detail=detail,
                            suggestion="检查 API Key、base_url、代理和证书设置",
                        )
            else:
                push_item(
                    key="llm_connect",
                    name="LLM 连接测试",
                    status="warn",
                    message="已跳过网络连通检查",
                )

        passed = sum(1 for item in items if item.get("status") == "pass")
        warned = sum(1 for item in items if item.get("status") == "warn")
        failed = sum(1 for item in items if item.get("status") == "fail")

        return {
            "ok": failed == 0,
            "checked_at": checked_at,
            "summary": {
                "total": len(items),
                "passed": passed,
                "warned": warned,
                "failed": failed,
            },
            "items": items,
            "xhs_diagnostics": [
                {
                    "key": str(item.get("key") or ""),
                    "name": str(item.get("name") or ""),
                    "status": str(item.get("status") or "warn"),
                    "reason": str(item.get("reason") or item.get("message") or ""),
                    "detail": str(item.get("detail") or ""),
                    "fix_command": str(item.get("fix_command") or item.get("suggestion") or ""),
                }
                for item in items
                if str(item.get("key") or "").startswith("xhs_")
            ],
        }

    def _check_email_connectivity(self, *, host: str, port: int, use_ssl: bool, username: str, password: str) -> tuple[bool, str]:
        try:
            if use_ssl:
                with smtplib.SMTP_SSL(host, int(port), timeout=8) as server:
                    server.login(username, password)
            else:
                with smtplib.SMTP(host, int(port), timeout=8) as server:
                    server.starttls()
                    server.login(username, password)
            return True, "SMTP 登录成功"
        except Exception as exc:
            return False, f"SMTP 登录失败: {exc}"

    def _check_llm_connectivity(self, *, base_url: str, api_key: str) -> tuple[bool, int, str]:
        try:
            import requests

            target = base_url.rstrip("/") + "/models"
            resp = requests.get(
                target,
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=(3, 8),
            )
            detail = str(resp.text or "")[:220].strip()
            if resp.ok:
                return True, int(resp.status_code), detail
            return False, int(resp.status_code), detail
        except Exception as exc:
            return False, 0, str(exc)

    def _resolve_retry_queue_path(self) -> Path:
        default_path = self.data_dir / "retry_queue.json"
        try:
            settings = load_settings(str(self.config_path))
            candidate = Path(str(settings.storage.retry_queue_path or "").strip() or str(default_path))
        except Exception:
            candidate = default_path
        if not candidate.is_absolute():
            candidate = self.workspace / candidate
        return candidate

    def _load_retry_queue(self) -> RetryQueue:
        return RetryQueue(path=str(self.retry_queue_path))

    @staticmethod
    def _safe_run_id(run_id: str) -> str:
        text = str(run_id or "").strip()
        if not text:
            return ""
        if not re.fullmatch(r"[A-Za-z0-9._:-]+", text):
            return ""
        return text

    def _resolve_xhs_script_path(self, args: Any) -> Path:
        values: list[str] = []
        if isinstance(args, list):
            values = [str(item).strip() for item in args if str(item).strip()]
        elif isinstance(args, str):
            text = args.strip()
            if text:
                values = [text]
        first = values[0] if values else "vendor/xhs-mcp/dist/xhs-mcp.js"
        script = Path(first)
        if not script.is_absolute():
            script = self.workspace / script
        return script

    @staticmethod
    def _resolve_selected_xhs_cookie_file(*, account: str, account_cookies_dir: str) -> Path:
        account_name = str(account or "default").strip() or "default"
        if account_name.lower() == "default":
            return Path.home() / ".xhs-mcp" / "cookies.json"

        base = Path(str(account_cookies_dir or "~/.xhs-mcp/accounts")).expanduser()
        if account_name.lower().endswith(".json"):
            flat = base / account_name
            nested = base / account_name[: -len(".json")] / "cookies.json"
        else:
            flat = base / f"{account_name}.json"
            nested = base / account_name / "cookies.json"
        if flat.exists():
            return flat
        if nested.exists():
            return nested
        return flat

    def _read_env_values(self) -> dict[str, str]:
        env_path = self.workspace / ".env"
        out: dict[str, str] = {}
        if not env_path.exists():
            return out
        try:
            for line in env_path.read_text(encoding="utf-8").splitlines():
                text = str(line or "").strip()
                if not text or text.startswith("#") or "=" not in text:
                    continue
                key, value = text.split("=", 1)
                key = key.strip()
                value = value.strip().strip("\"'")
                if key:
                    out[key] = value
        except Exception:
            return out
        return out

    @staticmethod
    def _get_env_value(name: str, env_map: dict[str, str]) -> str:
        key = str(name or "").strip()
        if not key:
            return ""
        runtime = str(os.getenv(key, "") or "").strip()
        if runtime:
            return runtime
        return str(env_map.get(key, "") or "").strip()

    @staticmethod
    def _compact_json(value: Any) -> str:
        try:
            return json.dumps(value, ensure_ascii=False, separators=(",", ":"))[:280]
        except Exception:
            return str(value)[:280]

    def _load_runs(self, limit: int) -> list[dict[str, Any]]:
        if not self.runs_dir.exists():
            return []

        out: list[dict[str, Any]] = []
        files = sorted(self.runs_dir.glob("*.json"), reverse=True)
        for file in files[: max(1, limit)]:
            try:
                payload = json.loads(file.read_text(encoding="utf-8"))
            except Exception:
                continue
            stats = payload.get("stats")
            if not isinstance(stats, dict):
                stats = {}
            stage_records = payload.get("stage_records")
            if not isinstance(stage_records, list):
                stage_records = []

            slow_stages = stats.get("stage_top_slow")
            if not isinstance(slow_stages, list):
                normalized_stages: list[dict[str, Any]] = []
                for item in stage_records:
                    if not isinstance(item, dict):
                        continue
                    normalized_stages.append(
                        {
                            "name": str(item.get("name") or ""),
                            "duration_ms": self._to_int(item.get("duration_ms")),
                            "status": str(item.get("status") or ""),
                        }
                    )
                slow_stages = sorted(normalized_stages, key=lambda x: self._to_int(x.get("duration_ms")), reverse=True)[:3]

            stage_total_ms = self._to_int(stats.get("stage_total_ms"))
            if stage_total_ms <= 0 and stage_records:
                stage_total_ms = sum(self._to_int(item.get("duration_ms")) for item in stage_records if isinstance(item, dict))

            stage_avg_ms = self._to_int(stats.get("stage_avg_ms"))
            if stage_avg_ms <= 0 and stage_records:
                stage_avg_ms = int(stage_total_ms / max(1, len(stage_records)))

            stage_failed_count = self._to_int(stats.get("stage_failed_count"))
            if stage_failed_count <= 0 and stage_records:
                stage_failed_count = sum(
                    1
                    for item in stage_records
                    if isinstance(item, dict) and str(item.get("status") or "").strip().lower() == "failed"
                )

            llm_error_codes = stats.get("llm_error_codes")
            if not isinstance(llm_error_codes, dict):
                llm_error_codes = {}
            stage_error_codes = stats.get("stage_error_codes")
            if not isinstance(stage_error_codes, dict):
                stage_error_codes = {}
                for item in stage_records:
                    if not isinstance(item, dict):
                        continue
                    if str(item.get("status") or "").strip().lower() != "failed":
                        continue
                    code = str(item.get("error_code") or "").strip().lower() or "stage_failed"
                    stage_error_codes[code] = self._to_int(stage_error_codes.get(code)) + 1

            merged_error_codes: dict[str, int] = {}
            for source in (llm_error_codes, stage_error_codes):
                for key, value in source.items():
                    code = str(key or "").strip().lower()
                    if not code:
                        continue
                    merged_error_codes[code] = self._to_int(merged_error_codes.get(code)) + self._to_int(value)

            alerts_triggered_raw = stats.get("alerts_triggered")
            if isinstance(alerts_triggered_raw, list):
                alerts_triggered = [item for item in alerts_triggered_raw if isinstance(item, dict)]
            else:
                alerts_triggered = []
            alert_codes = [
                str(item.get("code") or "").strip().lower()
                for item in alerts_triggered
                if str(item.get("code") or "").strip()
            ]
            alerts_notified_raw = stats.get("alerts_notified")
            if isinstance(alerts_notified_raw, list):
                alerts_notified = [str(item).strip().lower() for item in alerts_notified_raw if str(item).strip()]
            else:
                alerts_notified = []

            out.append(
                {
                    "run_id": str(payload.get("run_id") or file.stem),
                    "recorded_at": str(payload.get("recorded_at") or ""),
                    "mode": str(payload.get("mode") or stats.get("mode") or ""),
                    "notification_mode": str(payload.get("notification_mode") or stats.get("notification_mode") or ""),
                    "fetched": self._to_int(stats.get("fetched") if stats else payload.get("fetched")),
                    "target_notes": self._to_int(stats.get("target_notes") if stats else payload.get("target_notes")),
                    "jobs": self._to_int(stats.get("jobs") if stats else payload.get("jobs")),
                    "send_logs": self._to_int(stats.get("send_logs") if stats else payload.get("send_logs")),
                    "digest_sent": bool(stats.get("digest_sent") if stats else payload.get("digest_sent")),
                    "llm_fail": self._to_int(stats.get("llm_fail")),
                    "llm_calls": self._to_int(stats.get("llm_calls")),
                    "process_workers": self._to_int(stats.get("process_workers")),
                    "detail_workers": self._to_int(stats.get("detail_workers")),
                    "stage_total_ms": stage_total_ms,
                    "stage_avg_ms": stage_avg_ms,
                    "stage_failed_count": stage_failed_count,
                    "slow_stages": slow_stages,
                    "error_codes": merged_error_codes,
                    "retry_pending": stats.get("retry_pending") if isinstance(stats.get("retry_pending"), dict) else {},
                    "retry_running": stats.get("retry_running") if isinstance(stats.get("retry_running"), dict) else {},
                    "retry_enqueued": self._to_int(stats.get("retry_enqueued")),
                    "retry_retried": self._to_int(stats.get("retry_retried")),
                    "retry_succeeded": self._to_int(stats.get("retry_succeeded")),
                    "retry_dropped": self._to_int(stats.get("retry_dropped")),
                    "fetch_fail_count_run": self._to_int(stats.get("fetch_fail_count_run")),
                    "fetch_fail_streak": self._to_int(stats.get("fetch_fail_streak")),
                    "xhs_data_empty": bool(stats.get("xhs_data_empty")),
                    "detail_attempted": self._to_int(stats.get("detail_attempted")),
                    "detail_target_notes": self._to_int(stats.get("detail_target_notes")),
                    "detail_success": self._to_int(stats.get("detail_success")),
                    "detail_failed": self._to_int(stats.get("detail_failed")),
                    "detail_filled": self._to_int(stats.get("detail_filled")),
                    "detail_missing": self._to_int(stats.get("detail_missing")),
                    "detail_missing_rate": float(stats.get("detail_missing_rate") or 0.0),
                    "llm_timeout_count": self._to_int(stats.get("llm_timeout_count")),
                    "llm_timeout_rate": float(stats.get("llm_timeout_rate") or 0.0),
                    "alerts_triggered_count": self._to_int(
                        stats.get("alerts_triggered_count")
                        if "alerts_triggered_count" in stats
                        else len(alert_codes)
                    ),
                    "alerts_notified_count": self._to_int(
                        stats.get("alerts_notified_count")
                        if "alerts_notified_count" in stats
                        else len(alerts_notified)
                    ),
                    "alert_codes": alert_codes,
                    "alerts_triggered": alerts_triggered,
                    "alerts_notified": alerts_notified,
                    "xhs_diagnosis": stats.get("xhs_diagnosis") if isinstance(stats.get("xhs_diagnosis"), dict) else {},
                }
            )
        return out

    def _load_workbook_rows(self) -> dict[str, list[dict[str, Any]]]:
        if not self.excel_path.exists():
            return {"raw_notes": [], "succession_summary": [], "jobs": [], "send_log": []}

        wb = load_workbook(self.excel_path, read_only=True)
        try:
            out: dict[str, list[dict[str, Any]]] = {}
            for name in ("raw_notes", "succession_summary", "jobs", "send_log"):
                if name not in wb.sheetnames:
                    out[name] = []
                    continue
                out[name] = self._read_sheet_rows(wb[name])
            return out
        finally:
            wb.close()

    @staticmethod
    def _read_sheet_rows(ws) -> list[dict[str, Any]]:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return []
        headers = [str(x or "").strip() for x in headers_row]
        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, Any] = {}
            for idx, value in enumerate(values):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if key:
                    row[key] = value
            rows.append(row)
        return rows

    @staticmethod
    def _normalize_edit_text(value: Any) -> str:
        text = str(value or "")
        return text.replace("\r\n", "\n").replace("\r", "\n").strip()

    @staticmethod
    def _append_row(ws) -> int:
        return max(2, ws.max_row + 1)

    @staticmethod
    def _header_map(ws) -> dict[str, int]:
        max_col = max(1, ws.max_column)
        mapping: dict[str, int] = {}
        for col in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col).value or "").strip()
            if header:
                mapping[header] = col
        return mapping

    def _ensure_sheet_with_headers(self, wb, sheet_name: str, expected_headers: list[str]):
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
        header_map = self._header_map(ws)
        if not header_map:
            for idx, key in enumerate(expected_headers, start=1):
                ws.cell(row=1, column=idx, value=key)
            return ws, {key: idx for idx, key in enumerate(expected_headers, start=1)}
        for key in expected_headers:
            if key in header_map:
                continue
            col = max(header_map.values(), default=0) + 1
            ws.cell(row=1, column=col, value=key)
            header_map[key] = col
        return ws, header_map

    @staticmethod
    def _find_row_index(ws, headers: dict[str, int], *, key_column: str, key_value: str) -> int:
        col = headers.get(key_column)
        if not col:
            return 0
        target = str(key_value or "").strip()
        if not target:
            return 0
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col).value
            if str(cell or "").strip() == target:
                return row
        return 0

    @staticmethod
    def _set_cell(ws, headers: dict[str, int], row: int, column_name: str, value: Any) -> None:
        col = headers.get(column_name)
        if not col:
            return
        ws.cell(row=row, column=col, value=value)

    @staticmethod
    def _row_snapshot(ws, headers: dict[str, int], row: int) -> dict[str, Any]:
        out: dict[str, Any] = {}
        for key, col in headers.items():
            out[key] = ws.cell(row=row, column=col).value
        return out

    @staticmethod
    def _resolve_status_key(like_count: int, comment_count: int, has_summary: bool, has_job: bool) -> str:
        score = like_count + comment_count * 2
        if has_job and score >= 20:
            return "high_priority"
        if has_job:
            return "actionable"
        if has_summary:
            return "pending_review"
        return "new_lead"

    @classmethod
    def _resolve_status(cls, like_count: int, comment_count: int, has_summary: bool, has_job: bool) -> str:
        key = cls._resolve_status_key(
            like_count=like_count,
            comment_count=comment_count,
            has_summary=has_summary,
            has_job=has_job,
        )
        labels = {
            "high_priority": "高优先级",
            "actionable": "可推进",
            "pending_review": "待复核",
            "new_lead": "新线索",
        }
        return labels.get(key, "新线索")

    @staticmethod
    def _to_int(value: Any) -> int:
        try:
            if value is None:
                return 0
            text = str(value).strip().replace(",", "")
            if not text:
                return 0
            return int(float(text))
        except Exception:
            return 0

    @staticmethod
    def _to_epoch(value: Any) -> int:
        if isinstance(value, datetime):
            try:
                if value.tzinfo is None:
                    value = value.replace(tzinfo=timezone.utc)
                return int(value.timestamp())
            except Exception:
                return 0
        text = str(value or "").strip()
        if not text:
            return 0
        try:
            dt = datetime.fromisoformat(text)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return int(dt.timestamp())
        except Exception:
            return 0

    @staticmethod
    def _format_time_from_epoch_or_iso(epoch: int, iso_text: str) -> str:
        try:
            if int(epoch) > 0:
                dt = datetime.fromtimestamp(int(epoch))
                return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            pass
        text = str(iso_text or "").strip()
        if text:
            try:
                dt = datetime.fromisoformat(text)
                return dt.strftime("%Y-%m-%d %H:%M")
            except Exception:
                return text
        return "-"

    def _read_digest_interval(self) -> int:
        if not self.config_path.exists():
            return 60
        try:
            data = yaml.safe_load(self.config_path.read_text(encoding="utf-8")) or {}
            notification = data.get("notification", {}) if isinstance(data, dict) else {}
            raw = notification.get("digest_interval_minutes", 60)
            return max(1, int(raw))
        except Exception:
            return 60

    def _read_config_data(self) -> dict[str, Any]:
        if not self.config_path.exists():
            return {}
        data = yaml.safe_load(self.config_path.read_text(encoding="utf-8")) or {}
        if isinstance(data, dict):
            return data
        return {}

    def _get_resume_loader(self) -> ResumeLoader:
        if self._resume_loader is not None:
            return self._resume_loader
        config = self._read_config_data()
        resume_cfg_raw = config.get("resume", {}) if isinstance(config, dict) else {}
        if not isinstance(resume_cfg_raw, dict):
            resume_cfg_raw = {}

        source_txt_path = str(resume_cfg_raw.get("source_txt_path") or "config/resume.txt").strip() or "config/resume.txt"
        resume_text_path = str(resume_cfg_raw.get("resume_text_path") or "data/resume_text.txt").strip() or "data/resume_text.txt"
        max_chars = self._coerce_int(resume_cfg_raw.get("max_chars"), 6000, minimum=500)

        source = Path(source_txt_path)
        if not source.is_absolute():
            source = self.workspace / source
        target = Path(resume_text_path)
        if not target.is_absolute():
            target = self.workspace / target

        self._resume_source_path = source
        self._resume_text_path = target
        cfg = ResumeConfig(source_txt_path=str(source), resume_text_path=str(target), max_chars=max_chars)
        self._resume_loader = ResumeLoader(cfg, logger=self)
        return self._resume_loader

    def warning(self, *args, **kwargs) -> None:
        return

    def _write_config_data(self, data: dict[str, Any]) -> None:
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        text = yaml.safe_dump(data, sort_keys=False, allow_unicode=True)
        self.config_path.write_text(text, encoding="utf-8")

    @staticmethod
    def _ensure_section(root: dict[str, Any], key: str) -> dict[str, Any]:
        item = root.get(key)
        if not isinstance(item, dict):
            item = {}
            root[key] = item
        return item

    @staticmethod
    def _coerce_int(value: Any, default: int, minimum: int = 0) -> int:
        try:
            parsed = int(str(value).strip())
        except Exception:
            parsed = int(default)
        return max(minimum, parsed)

    @staticmethod
    def _coerce_rate(value: Any, default: float) -> float:
        try:
            parsed = float(value)
        except Exception:
            parsed = float(default)
        if parsed > 1.0 and parsed <= 100.0:
            parsed = parsed / 100.0
        return max(0.0, min(1.0, parsed))

    @staticmethod
    def _coerce_float(value: Any, default: float, minimum: float = 0.0) -> float:
        try:
            parsed = float(value)
        except Exception:
            parsed = float(default)
        if parsed < float(minimum):
            parsed = float(minimum)
        return parsed

    @staticmethod
    def _normalize_mode(value: str) -> str:
        mode = (value or "auto").strip().lower()
        if mode == "smart":
            mode = "agent"
        if mode not in {"auto", "agent"}:
            return "auto"
        return mode

    @staticmethod
    def _normalize_notify_mode(value: str) -> str:
        mode = (value or "digest").strip().lower()
        if mode not in {"digest", "realtime", "off"}:
            return "digest"
        return mode

    @staticmethod
    def _to_name_list(value: Any, default: list[str]) -> list[str]:
        if isinstance(value, list):
            items = [str(item).strip() for item in value if str(item).strip()]
            return items or list(default)
        if isinstance(value, str):
            items = [x.strip() for x in value.split(",") if x.strip()]
            return items or list(default)
        return list(default)
