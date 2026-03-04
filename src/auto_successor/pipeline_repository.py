from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .excel_store import JOB_HEADERS
from .models import JobRecord


class PipelineRepositoryMixin:
    def _load_recent_alert_stats(self, *, limit: int, exclude_run_id: str) -> list[dict[str, Any]]:
        max_items = max(0, int(limit))
        if max_items <= 0:
            return []
        runs_dir = getattr(self.journal, "base_dir", None)
        if runs_dir is None:
            return []
        if not isinstance(runs_dir, Path):
            runs_dir = Path(str(runs_dir))
        if not runs_dir.exists():
            return []

        history: list[dict[str, Any]] = []
        files = sorted(runs_dir.glob("*.json"), key=lambda item: item.name, reverse=True)
        for file in files:
            if str(file.stem) == str(exclude_run_id):
                continue
            try:
                payload = json.loads(file.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not isinstance(payload, dict):
                continue
            stats = payload.get("stats")
            if not isinstance(stats, dict):
                continue
            history.append(stats)
            if len(history) >= max_items:
                break
        return history

    def _load_latest_jobs_from_store(self, limit: int) -> list[JobRecord]:
        excel_path = Path(self.settings.storage.excel_path)
        if not excel_path.exists():
            return []

        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            if "jobs" not in wb.sheetnames:
                return []
            ws = wb["jobs"]
            rows = self._read_sheet_rows(ws)
        finally:
            wb.close()

        rows.sort(
            key=lambda item: (
                self._to_datetime(item.get("publish_time")),
                str(item.get("PostID") or ""),
            ),
            reverse=True,
        )

        jobs: list[JobRecord] = []
        for row in rows[: max(1, int(limit))]:
            post_id = str(row.get("PostID") or "").strip()
            if not post_id:
                continue
            jobs.append(
                JobRecord(
                    run_id=str(row.get("run_id") or f"from-store:{post_id[:12]}"),
                    post_id=post_id,
                    company=str(row.get("Company") or ""),
                    position=str(row.get("Position") or ""),
                    location=str(row.get("Location") or ""),
                    requirements=str(row.get("Requirements") or ""),
                    arrival_time=str(row.get("arrival_time") or ""),
                    application_method=str(row.get("application_method") or ""),
                    author=str(row.get("author") or ""),
                    risk_line=str(row.get("risk_line") or "low"),
                    match_score=float(row.get("match_score") or 0.0),
                    match_reason=str(row.get("match_reason") or ""),
                    link=str(row.get("Link") or ""),
                    mode=str(row.get("mode") or "auto"),
                    publish_time=self._to_datetime(row.get("publish_time")),
                    source_title=str(row.get("source_title") or ""),
                    comment_count=int(row.get("comment_count") or 0),
                    comments_preview=str(row.get("comments_preview") or ""),
                    original_text=str(row.get("original_text") or ""),
                    opportunity_point=bool(row.get("opportunity_point")),
                    outreach_message=str(row.get("outreach_message") or ""),
                )
            )
        return jobs

    def _load_latest_summaries_from_store(self, limit: int):
        jobs = self._load_latest_jobs_from_store(limit=limit)
        return self._jobs_to_summary_records(run_id="from-store", jobs=jobs)

    @staticmethod
    def _read_sheet_rows(ws) -> list[dict[str, Any]]:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers_row = next(iterator)
        except StopIteration:
            return []
        headers = [str(x or "").strip() for x in headers_row]
        if not any(headers):
            headers = list(JOB_HEADERS)

        rows: list[dict[str, Any]] = []
        for values in iterator:
            if not values:
                continue
            row: dict[str, Any] = {}
            for idx, value in enumerate(values):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if key:
                    row[key] = value
            rows.append(row)
        return rows

    @staticmethod
    def _to_datetime(value: Any) -> datetime:
        epoch = datetime(1970, 1, 1, tzinfo=timezone.utc)
        if isinstance(value, datetime):
            dt = value
        else:
            text = str(value or "").strip()
            if not text:
                return epoch
            try:
                dt = datetime.fromisoformat(text)
            except Exception:
                return epoch
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt
