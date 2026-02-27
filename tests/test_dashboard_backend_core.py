from __future__ import annotations

import json
from pathlib import Path
import tempfile
import unittest

import yaml
from openpyxl import Workbook

from auto_successor.dashboard_backend import DataBackend
from auto_successor.excel_store import JOB_HEADERS, RAW_HEADERS, SUMMARY_HEADERS


def _write_config(path: Path) -> None:
    data = {
        "app": {"interval_minutes": 15},
        "xhs": {
            "keyword": "缁т换",
            "search_sort": "time_descending",
            "max_results": 20,
            "max_detail_fetch": 5,
            "account": "default",
            "account_cookies_dir": "~/.xhs-mcp/accounts",
        },
        "agent": {"mode": "auto"},
        "notification": {
            "mode": "digest",
            "digest_interval_minutes": 30,
            "digest_top_summaries": 5,
            "digest_send_when_no_new": False,
            "digest_channels": ["email"],
            "realtime_channels": ["wechat_service", "email"],
        },
        "email": {"enabled": True},
        "wechat_service": {"enabled": False},
        "llm": {"enabled": False, "model": "gpt-5-mini", "base_url": "https://api.openai.com/v1"},
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(yaml.safe_dump(data, sort_keys=False, allow_unicode=True), encoding="utf-8")


def _write_min_excel(path: Path, *, raw_rows: list[dict] | None = None, job_rows: list[dict] | None = None, summary_rows: list[dict] | None = None) -> None:
    raw_rows = raw_rows or []
    job_rows = job_rows or []
    summary_rows = summary_rows or []
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "raw_notes"
    ws.append(RAW_HEADERS)
    for row in raw_rows:
        ws.append([row.get(h, "") for h in RAW_HEADERS])

    ws_summary = wb.create_sheet("succession_summary")
    ws_summary.append(SUMMARY_HEADERS)
    for row in summary_rows:
        ws_summary.append([row.get(h, "") for h in SUMMARY_HEADERS])

    ws_send = wb.create_sheet("send_log")
    ws_send.append(["run_id", "note_id", "channel", "send_status", "send_response", "sent_at"])

    ws_job = wb.create_sheet("jobs")
    ws_job.append(JOB_HEADERS)
    for row in job_rows:
        ws_job.append([row.get(h, "") for h in JOB_HEADERS])

    wb.save(path)
    wb.close()


def test_config_roundtrip(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    backend = DataBackend(workspace=workspace)

    view = backend.load_config_view()
    assert view["xhs"]["keyword"] == "缁т换"
    assert view["notification"]["digest_interval_minutes"] == 30
    assert view["xhs"]["account"] == "default"

    saved = backend.save_config_view(
        {
            "xhs": {"keyword": "缁т换宀椾綅", "max_results": 18, "account": "acc-a"},
            "notification": {"digest_interval_minutes": 60},
            "agent": {"mode": "agent"},
        }
    )
    assert saved["xhs"]["keyword"] == "缁т换宀椾綅"
    assert saved["xhs"]["max_results"] == 18
    assert saved["xhs"]["account"] == "acc-a"
    assert saved["notification"]["digest_interval_minutes"] == 60
    assert saved["agent"]["mode"] == "agent"


def test_runtime_no_job_stop_is_safe(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    backend = DataBackend(workspace=workspace)

    result = backend.run_action("stop_job", {})
    assert result["ok"] is False
    assert "任务" in str(result.get("message", ""))


def test_run_action_invalid(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    backend = DataBackend(workspace=workspace)

    try:
        backend.run_action("unknown_action", {})
    except ValueError as exc:
        assert "unsupported action" in str(exc)
    else:
        raise AssertionError("expected ValueError")


def test_load_leads_page_empty(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    backend = DataBackend(workspace=workspace)

    result = backend.load_leads_page(page=1, page_size=20, q="", summary_only=False)
    assert result["items"] == []
    assert result["total"] == 0
    assert result["page"] == 1
    assert result["page_size"] == 20
    assert result["total_pages"] == 1


def test_load_runs_with_stage_observability_from_stats(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    runs_dir = workspace / "data" / "runs"
    runs_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "run_id": "r-observe",
        "recorded_at": "2026-02-25T12:00:00+00:00",
        "stats": {
            "mode": "agent",
            "notification_mode": "digest",
            "fetched": 9,
            "target_notes": 4,
            "jobs": 3,
            "send_logs": 2,
            "digest_sent": True,
            "stage_total_ms": 4800,
            "stage_avg_ms": 1200,
            "stage_failed_count": 1,
            "stage_top_slow": [
                {"name": "collector.search", "duration_ms": 2600, "status": "success"},
                {"name": "agent.filter", "duration_ms": 1400, "status": "failed"},
            ],
            "llm_error_codes": {"read_timeout": 2},
            "stage_error_codes": {"timeout": 1},
        },
    }
    (runs_dir / "20260225-120000.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    backend = DataBackend(workspace=workspace)

    result = backend.load_runs(limit=10)
    assert len(result) == 1
    item = result[0]
    assert item["mode"] == "agent"
    assert item["notification_mode"] == "digest"
    assert item["stage_total_ms"] == 4800
    assert item["stage_avg_ms"] == 1200
    assert item["stage_failed_count"] == 1
    assert item["slow_stages"][0]["name"] == "collector.search"
    assert item["error_codes"]["read_timeout"] == 2
    assert item["error_codes"]["timeout"] == 1


def test_load_runs_fallback_stage_metrics_from_stage_records(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    runs_dir = workspace / "data" / "runs"
    runs_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "run_id": "r-stage-fallback",
        "recorded_at": "2026-02-25T12:10:00+00:00",
        "stats": {
            "fetched": 2,
            "target_notes": 1,
            "jobs": 1,
        },
        "stage_records": [
            {"name": "collector.search", "duration_ms": 900, "status": "success"},
            {"name": "agent.filter", "duration_ms": 300, "status": "failed", "error_code": "network_error"},
        ],
    }
    (runs_dir / "20260225-121000.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    backend = DataBackend(workspace=workspace)

    result = backend.load_runs(limit=10)
    assert len(result) == 1
    item = result[0]
    assert item["stage_total_ms"] == 1200
    assert item["stage_avg_ms"] == 600
    assert item["stage_failed_count"] == 1
    assert item["slow_stages"][0]["name"] == "collector.search"
    assert item["error_codes"]["network_error"] == 1


def test_setup_check_offline_mode(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")

    cfg_path = workspace / "config" / "config.yaml"
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    cfg.setdefault("xhs", {})
    cfg["xhs"]["command"] = "python"
    cfg["xhs"]["args"] = ["vendor/xhs-mcp/dist/xhs-mcp.js"]
    cfg.setdefault("email", {})
    cfg["email"]["enabled"] = False
    cfg.setdefault("llm", {})
    cfg["llm"]["enabled"] = False
    cfg_path.write_text(yaml.safe_dump(cfg, sort_keys=False, allow_unicode=True), encoding="utf-8")

    script = workspace / "vendor" / "xhs-mcp" / "dist" / "xhs-mcp.js"
    script.parent.mkdir(parents=True, exist_ok=True)
    script.write_text("// fake script for setup check\n", encoding="utf-8")

    backend = DataBackend(workspace=workspace)
    result = backend.run_setup_check(include_network=False, include_xhs_status=False)
    assert result["ok"] is True
    assert result["summary"]["failed"] == 0
    assert isinstance(result.get("xhs_diagnostics"), list)
    keys = {item["key"]: item for item in result["items"]}
    assert keys["config_file"]["status"] == "pass"
    assert keys["storage_write"]["status"] == "pass"
    assert keys["xhs_cli_present"]["status"] in {"pass", "warn"}
    assert keys["xhs_runtime"]["status"] == "pass"
    assert keys["xhs_cookie_file_ready"]["status"] in {"pass", "warn"}
    assert keys["xhs_mcp_connect"]["status"] == "warn"
    assert keys["xhs_login_status"]["status"] == "warn"
    assert keys["xhs_login"]["status"] == "warn"
    assert keys["email_enabled"]["status"] == "warn"
    assert keys["llm_enabled"]["status"] == "warn"


def test_load_xhs_accounts_view_lists_available_accounts(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")

    cfg_path = workspace / "config" / "config.yaml"
    cfg = yaml.safe_load(cfg_path.read_text(encoding="utf-8")) or {}
    cfg.setdefault("xhs", {})
    cfg["xhs"]["account"] = "acc-a"
    cfg["xhs"]["account_cookies_dir"] = str(workspace / "accounts")
    cfg_path.write_text(yaml.safe_dump(cfg, sort_keys=False, allow_unicode=True), encoding="utf-8")

    account_dir = workspace / "accounts"
    account_dir.mkdir(parents=True, exist_ok=True)
    (account_dir / "acc-a.json").write_text("[]", encoding="utf-8")
    (account_dir / "acc-b" / "cookies.json").parent.mkdir(parents=True, exist_ok=True)
    (account_dir / "acc-b" / "cookies.json").write_text("[]", encoding="utf-8")

    backend = DataBackend(workspace=workspace)
    view = backend.load_xhs_accounts_view()
    values = {str(item.get("value") or "") for item in view.get("options", [])}
    assert view["selected"] == "acc-a"
    assert "default" in values
    assert "acc-a" in values
    assert "acc-b" in values


def test_load_run_detail(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    runs_dir = workspace / "data" / "runs"
    runs_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "run_id": "r-detail",
        "recorded_at": "2026-02-27T10:00:00+00:00",
        "mode": "auto",
        "notification_mode": "digest",
        "stats": {"fetched": 3, "llm_error_codes": {"read_timeout": 1}, "stage_error_codes": {"network": 2}},
        "stage_records": [
            {"name": "collector.search", "status": "success", "duration_ms": 1000},
            {"name": "agent.extract", "status": "failed", "duration_ms": 800, "error_code": "network"},
        ],
        "fetch_fail_events": [{"stage": "collector.search_notes", "error": "timeout"}],
        "xhs_diagnosis": {"failure_category": "mcp_unreachable"},
        "retry": {"pending": {"fetch": 1, "llm_timeout": 0, "email": 0}},
    }
    (runs_dir / "r-detail.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    backend = DataBackend(workspace=workspace)
    detail = backend.load_run_detail("r-detail")
    assert detail["run_id"] == "r-detail"
    assert len(detail["failed_stages"]) == 1
    assert detail["llm_error_codes"]["read_timeout"] == 1
    assert detail["stage_error_codes"]["network"] == 2


def test_retry_queue_view_and_mutations(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    backend = DataBackend(workspace=workspace)

    queue = backend._load_retry_queue()
    item = queue.enqueue(queue_type="fetch", action="search_notes", payload={"keyword": "继任"}, run_id="r-queue", error="x")
    view = backend.load_retry_queue_view(status="all", queue_type="fetch", limit=20)
    assert view["items"]
    assert view["items"][0]["id"] == item["id"]

    requeue_resp = backend.retry_queue_requeue(item["id"])
    assert requeue_resp["ok"] is True
    assert requeue_resp["item"]["status"] == "pending"

    drop_resp = backend.retry_queue_drop(item["id"])
    assert drop_resp["ok"] is True
    assert drop_resp["item"]["status"] == "dropped"


def test_load_leads_page_with_status_and_dedupe_filters(tmp_path: Path) -> None:
    workspace = tmp_path
    _write_config(workspace / "config" / "config.yaml")
    excel_path = workspace / "data" / "output.xlsx"
    _write_min_excel(
        excel_path,
        raw_rows=[
            {
                "run_id": "r1",
                "keyword": "继任",
                "note_id": "note_new",
                "title": "title a",
                "author": "author a",
                "publish_time": "2026-02-27T11:00:00",
                "publish_timestamp": 1772190000,
                "publish_time_text": "10分钟前",
                "publish_time_quality": "parsed",
                "like_count": 1,
                "comment_count": 0,
                "share_count": 0,
                "url": "u1",
                "xsec_token": "x1",
                "detail_text": "d1",
                "comments_preview": "",
                "fetched_at": "2026-02-27T11:05:00",
                "first_seen_at": "2026-02-27T11:05:00",
                "updated_at": "2026-02-27T11:05:00",
                "raw_json": "{}",
            },
            {
                "run_id": "r2",
                "keyword": "继任",
                "note_id": "note_updated",
                "title": "title b",
                "author": "author b",
                "publish_time": "2026-02-27T12:00:00",
                "publish_timestamp": 1772193600,
                "publish_time_text": "5分钟前",
                "publish_time_quality": "parsed",
                "like_count": 30,
                "comment_count": 2,
                "share_count": 0,
                "url": "u2",
                "xsec_token": "x2",
                "detail_text": "d2",
                "comments_preview": "",
                "fetched_at": "2026-02-27T12:10:00",
                "first_seen_at": "2026-02-27T12:01:00",
                "updated_at": "2026-02-27T12:10:00",
                "raw_json": "{}",
            },
        ],
        job_rows=[
            {
                "run_id": "r2",
                "Company": "Acme",
                "Position": "Intern",
                "publish_time": "2026-02-27T12:00:00",
                "Location": "Shanghai",
                "Requirements": "SQL",
                "PostID": "note_updated",
            }
        ],
    )

    backend = DataBackend(workspace=workspace)
    all_rows = backend.load_leads_page(page=1, page_size=20, status_filter="all", dedupe_filter="all")
    assert all_rows["total"] == 2

    updated_rows = backend.load_leads_page(page=1, page_size=20, status_filter="all", dedupe_filter="updated")
    assert updated_rows["total"] == 1
    assert updated_rows["items"][0]["note_id"] == "note_updated"

    high_rows = backend.load_leads_page(page=1, page_size=20, status_filter="high_priority", dedupe_filter="all")
    assert high_rows["total"] == 1
    assert high_rows["items"][0]["status_key"] == "high_priority"


class TestDashboardBackendCore(unittest.TestCase):
    def test_load_runs_with_stage_observability_from_stats_unittest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            test_load_runs_with_stage_observability_from_stats(Path(tmp_dir))

    def test_load_runs_fallback_stage_metrics_from_stage_records_unittest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            test_load_runs_fallback_stage_metrics_from_stage_records(Path(tmp_dir))

    def test_setup_check_offline_mode_unittest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            test_setup_check_offline_mode(Path(tmp_dir))

    def test_load_xhs_accounts_view_lists_available_accounts_unittest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            test_load_xhs_accounts_view_lists_available_accounts(Path(tmp_dir))

