from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from auto_successor.excel_store import ExcelStore
from auto_successor.models import NoteRecord


def _read_raw_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["raw_notes"]
        iterator = ws.iter_rows(values_only=True)
        headers = [str(x or "") for x in next(iterator)]
        rows: list[dict] = []
        for values in iterator:
            if not values:
                continue
            rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
        return rows
    finally:
        wb.close()


class TestExcelStoreIncremental(unittest.TestCase):
    def test_upsert_same_note_keeps_single_row_and_tracks_update_time(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel_path = Path(tmp) / "output.xlsx"
            store = ExcelStore(str(excel_path))

            first_fetch = datetime(2026, 2, 26, 8, 0, tzinfo=timezone.utc)
            first_publish = datetime(2026, 2, 25, 10, 30, tzinfo=timezone.utc)
            note_first = NoteRecord(
                run_id="run-1",
                keyword="继任",
                note_id="note-1",
                title="找继任：初版",
                author="tester",
                publish_time=first_publish,
                publish_time_text="昨天 10:30",
                publish_time_quality="parsed",
                like_count=10,
                comment_count=3,
                share_count=1,
                url="https://www.xiaohongshu.com/explore/note-1",
                raw_json='{"v":1}',
                xsec_token="tok-1",
                detail_text="旧正文详情",
                comments_preview="旧评论线索",
                fetched_at=first_fetch,
            )
            store.write([note_first], [], [], jobs=[])

            second_fetch = datetime(2026, 2, 26, 9, 10, tzinfo=timezone.utc)
            note_second = NoteRecord(
                run_id="run-2",
                keyword="继任",
                note_id="note-1",
                title="找继任：更新版",
                author="tester",
                publish_time=datetime(2026, 2, 26, 9, 0, tzinfo=timezone.utc),
                publish_time_text="",
                publish_time_quality="fallback",
                like_count=0,
                comment_count=0,
                share_count=0,
                url="https://www.xiaohongshu.com/explore/note-1",
                raw_json='{"v":2}',
                xsec_token="tok-1",
                detail_text="",
                comments_preview="",
                fetched_at=second_fetch,
            )
            store.write([note_second], [], [], jobs=[])

            rows = _read_raw_rows(excel_path)
            self.assertEqual(len(rows), 1)
            row = rows[0]

            self.assertEqual(row["note_id"], "note-1")
            self.assertEqual(row["run_id"], "run-2")
            self.assertEqual(row["title"], "找继任：更新版")
            self.assertEqual(row["publish_timestamp"], int(first_publish.timestamp()))
            self.assertEqual(row["publish_time_quality"], "parsed")

            # Preserve rich fields/counters when incremental update has sparse data.
            self.assertEqual(row["detail_text"], "旧正文详情")
            self.assertEqual(row["comments_preview"], "旧评论线索")
            self.assertEqual(row["like_count"], 10)
            self.assertEqual(row["comment_count"], 3)
            self.assertEqual(row["share_count"], 1)

            self.assertEqual(row["first_seen_at"], first_fetch.isoformat())
            self.assertEqual(row["updated_at"], second_fetch.isoformat())
            self.assertEqual(row["fetched_at"], second_fetch.isoformat())


if __name__ == "__main__":
    unittest.main()

